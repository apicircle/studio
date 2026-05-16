"""Helper for an automated test runner (e.g., Claude Cowork) to record
manual-test-case outcomes into a results workbook.

Subcommands:
  init       Copy the source test-plan workbook to a results workbook.
  record     Mark one test case's actual result + status + notes.
  status     Print a one-line summary of progress.
  summarize  Append a Summary sheet update + print a final report.

Layout (results workbook):
  Test Cases sheet has the same columns as the source.
  Columns of interest:
    J  Actual Result
    K  Status        (Pass/Fail/Blocked/Skipped/Not Run)
    M  Tester        (set to "Claude Cowork")
    N  Test Date     (ISO-8601 YYYY-MM-DD)
    O  Notes / Defect ID  (evidence paths, retry info, blockers)

Usage examples (call from Bash inside Cowork):

  python tc_results.py init \
      --source ../web-app-manual-test-cases.xlsx \
      --target ../results/web-run-2026-05-14.xlsx

  python tc_results.py record \
      --workbook ../results/web-run-2026-05-14.xlsx \
      --tc TC-WS-0001 \
      --status Pass \
      --actual "Workspace created with name 'QA-Smoke-WS'; top bar updated." \
      --notes "evidence: results/evidence/run-001/TC-WS-0001-after.png"

  python tc_results.py status --workbook ../results/web-run-2026-05-14.xlsx

  python tc_results.py summarize --workbook ../results/web-run-2026-05-14.xlsx
"""
from __future__ import annotations
import argparse
import datetime as dt
import shutil
import sys
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook

COL_ACTUAL = "J"
COL_STATUS = "K"
COL_TESTER = "M"
COL_DATE = "N"
COL_NOTES = "O"

VALID_STATUS = {"Pass", "Fail", "Blocked", "Skipped", "Not Run"}


def cmd_init(args: argparse.Namespace) -> int:
    src = Path(args.source)
    dst = Path(args.target)
    if not src.exists():
        print(f"ERROR: source not found: {src}", file=sys.stderr)
        return 2
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)
    print(f"Initialized results workbook at {dst}")
    return 0


def _find_row(ws, tc_id: str) -> int | None:
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == tc_id:
            return row
    return None


def cmd_record(args: argparse.Namespace) -> int:
    if args.status not in VALID_STATUS:
        print(f"ERROR: status must be one of {sorted(VALID_STATUS)}", file=sys.stderr)
        return 2
    wb_path = Path(args.workbook)
    if not wb_path.exists():
        print(f"ERROR: workbook not found: {wb_path}", file=sys.stderr)
        return 2
    wb = load_workbook(wb_path)
    if "Test Cases" not in wb.sheetnames:
        print("ERROR: workbook has no 'Test Cases' sheet", file=sys.stderr)
        return 2
    ws = wb["Test Cases"]
    row = _find_row(ws, args.tc)
    if row is None:
        print(f"ERROR: TC ID not found: {args.tc}", file=sys.stderr)
        return 3
    if args.actual is not None:
        ws[f"{COL_ACTUAL}{row}"] = args.actual
    ws[f"{COL_STATUS}{row}"] = args.status
    ws[f"{COL_TESTER}{row}"] = args.tester or "Claude Cowork"
    ws[f"{COL_DATE}{row}"] = args.date or dt.date.today().isoformat()
    if args.notes is not None:
        ws[f"{COL_NOTES}{row}"] = args.notes
    wb.save(wb_path)
    print(f"Recorded {args.tc}: {args.status}")
    return 0


def _scan_statuses(ws) -> Counter:
    counts = Counter()
    for row in range(2, ws.max_row + 1):
        status = ws[f"{COL_STATUS}{row}"].value or "Not Run"
        counts[status] += 1
    return counts


def cmd_status(args: argparse.Namespace) -> int:
    wb = load_workbook(Path(args.workbook))
    ws = wb["Test Cases"]
    counts = _scan_statuses(ws)
    total = sum(counts.values())
    pass_n = counts.get("Pass", 0)
    fail_n = counts.get("Fail", 0)
    blocked = counts.get("Blocked", 0)
    skipped = counts.get("Skipped", 0)
    not_run = counts.get("Not Run", 0)
    executed = total - not_run
    pct = (pass_n / (pass_n + fail_n) * 100) if (pass_n + fail_n) else 0.0
    print(
        f"{executed}/{total} executed | "
        f"Pass={pass_n} Fail={fail_n} Blocked={blocked} Skipped={skipped} "
        f"NotRun={not_run} | Pass%={pct:.1f}%"
    )
    return 0


def cmd_summarize(args: argparse.Namespace) -> int:
    wb = load_workbook(Path(args.workbook))
    ws = wb["Test Cases"]
    counts = _scan_statuses(ws)
    total = sum(counts.values())
    print(f"=== Test run summary: {args.workbook} ===")
    print(f"Total rows: {total}")
    for s in ("Pass", "Fail", "Blocked", "Skipped", "Not Run"):
        print(f"  {s}: {counts.get(s, 0)}")
    # Per-module breakdown
    by_mod: dict[str, Counter] = {}
    for row in range(2, ws.max_row + 1):
        mod = ws.cell(row=row, column=2).value or "?"
        status = ws[f"{COL_STATUS}{row}"].value or "Not Run"
        by_mod.setdefault(mod, Counter())[status] += 1
    print("\nPer-module breakdown:")
    for mod, c in sorted(by_mod.items(), key=lambda x: -sum(x[1].values())):
        print(
            f"  {mod}: total={sum(c.values())} "
            f"pass={c.get('Pass', 0)} fail={c.get('Fail', 0)} "
            f"blocked={c.get('Blocked', 0)} skipped={c.get('Skipped', 0)} "
            f"notrun={c.get('Not Run', 0)}"
        )
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(prog="tc_results")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_init = sub.add_parser("init", help="Copy source workbook to a results workbook.")
    p_init.add_argument("--source", required=True)
    p_init.add_argument("--target", required=True)
    p_init.set_defaults(func=cmd_init)

    p_rec = sub.add_parser("record", help="Record a single test outcome.")
    p_rec.add_argument("--workbook", required=True)
    p_rec.add_argument("--tc", required=True, help="TC ID, e.g. TC-WS-0001")
    p_rec.add_argument("--status", required=True,
                       choices=sorted(VALID_STATUS))
    p_rec.add_argument("--actual", default=None,
                       help="What was actually observed.")
    p_rec.add_argument("--notes", default=None,
                       help="Evidence paths, retry info, blockers, defect link.")
    p_rec.add_argument("--tester", default=None,
                       help="Defaults to 'Claude Cowork'.")
    p_rec.add_argument("--date", default=None,
                       help="ISO-8601 date; defaults to today.")
    p_rec.set_defaults(func=cmd_record)

    p_st = sub.add_parser("status", help="Print one-line progress summary.")
    p_st.add_argument("--workbook", required=True)
    p_st.set_defaults(func=cmd_status)

    p_sum = sub.add_parser("summarize", help="Print module-level summary.")
    p_sum.add_argument("--workbook", required=True)
    p_sum.set_defaults(func=cmd_summarize)

    args = parser.parse_args()
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
