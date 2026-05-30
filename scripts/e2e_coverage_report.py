"""E2E ↔ manual-test coverage report.

Scans `e2e/web/**/*.spec.ts` and `e2e/desktop/**/*.spec.ts` for TC-ID
coverage and cross-references against the
manual test workbooks at `docs/qa/test_cases/{web,desktop}-app-test-cases.xlsx`.

## Two coverage modes

The script supports two definitions of "covered":

**Strict (default in --strict mode)** — only counts TC-IDs that appear
in a literal `tc('TC-XX-NNNN', ...)` or `tcCovered('TC-XX-NNNN')` call
in the spec source. This is the honest "is there real test code that
references this row?" number.

**Lenient (default for backward compat)** — also credits TC-IDs that
appear in any `tcMap*.ts` file imported by the spec. Useful for tracking
which modules are *wired* to automation, but inflates the number when a
spec imports a 547-entry map and only writes a handful of test()
bodies.

Always prefer `--strict` when answering "are we done?".

## Outputs

  docs/qa/results/e2e-coverage.md
    Module rollup + per-spec credits + gap list.
  docs/qa/results/e2e-coverage.json  (with --json)
    Machine-readable data for CI gating.

## Usage

  python scripts/e2e_coverage_report.py                  # lenient, MD only
  python scripts/e2e_coverage_report.py --strict         # honest count
  python scripts/e2e_coverage_report.py --strict --json  # CI input

Exit code: 0 on success, 1 on IO error.
"""
from __future__ import annotations
import argparse
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

REPO = Path(__file__).parent.parent
WEB_WORKBOOK = REPO / "docs" / "qa" / "test_cases" / "web-app-test-cases.xlsx"
DESKTOP_WORKBOOK = REPO / "docs" / "qa" / "test_cases" / "desktop-app-test-cases.xlsx"
SPEC_GLOBS = (
    "e2e/web/**/*.spec.ts",
    "e2e/desktop/**/*.spec.ts",
)
RESIDUE_FILE = REPO / "e2e" / "web" / "manual-residue.ts"
OUT_DIR = REPO / "docs" / "qa" / "results"
OUT_MD = OUT_DIR / "e2e-coverage.md"
OUT_JSON = OUT_DIR / "e2e-coverage.json"

TC_PATTERN = re.compile(r"TC-[A-Z0-9]{2,3}-\d{4}")

# Strict-mode patterns. A spec earns coverage credit for a TC-ID if one
# of these matches in the spec body — meaning real test code references
# the row, not just an unused `void tcMapXX;` to silence the linter.

# 1. Inline literal: `tc('TC-XX-NNNN', ...)` or `tcCovered('TC-XX-NNNN')`.
STRICT_TC_CALL = re.compile(
    r"""(?:^|\W)(?:tc|tcCovered)\s*\(\s*['"](TC-[A-Z0-9]{2,3}-\d{4})['"]"""
)

# 2. Inline array: `tc(['TC-AA-0001', 'TC-AA-0002'], ...)`.
STRICT_TC_LIST = re.compile(
    r"""(?:^|\W)tc\s*\(\s*\[\s*((?:['"]TC-[A-Z0-9]{2,3}-\d{4}['"]\s*,?\s*)+)\]"""
)

# 3. Parametric range: `tcRange('TC-XX', 1, 56, ...)`.
STRICT_TC_RANGE = re.compile(
    r"""tcRange\s*\(\s*['"](TC-[A-Z0-9]{2,3})['"]\s*,\s*(\d+)\s*,\s*(\d+)"""
)

# 4. **Real** map iteration — `for (const [k, v] of Object.entries(tcMapXX))`
#    or `for (const x of tcMapXX)`. These patterns *generate a test() per
#    cell*, so they justify whole-map credit. A bare
#    `void Object.keys(tcMapXX);` is a coverage-trigger marker (used to
#    bridge older specs into the scanner) — it credits NO entries; tests
#    must reference cells via specific keys.
STRICT_MAP_ITERATION = re.compile(
    r"""for\s*\([^)]*\s+of\s+"""
    r"""(?:Object\.(?:entries|keys|values)\s*\(\s*)?(tcMap[A-Z0-9]+)"""
)

# 5. Literal-key indexing — `tcMapXX['key']` / `tcMapXX["key"]` /
#    `` tcMapXX[`key`] ``. Credits only the specific TC-ID at that key.
STRICT_MAP_LITERAL_INDEX = re.compile(
    r"""(tcMap[A-Z0-9]+)\s*\[\s*['"`]([^'"`]+)['"`]\s*\]"""
)

# 6. Helper-function call: `function id(key) { ... tcMapXX[key] ... }`
#    is the canonical pattern across the suite (see workspace-management,
#    collections-requests, et al). Detect the helper definition + its
#    target map, then resolve every `helperName('literal')` call to the
#    corresponding TC-ID.
HELPER_DEFINITION = re.compile(
    r"""function\s+(\w+)\s*\([^)]*\)[^{]*\{[^}]*?(tcMap[A-Z0-9]+)\s*\[""",
    re.DOTALL,
)
HELPER_CALL = re.compile(
    r"""\b(\w+)\s*\(\s*['"`]([^'"`]+)['"`]\s*\)"""
)

IMPORT_PATTERN = re.compile(r"""from\s+['"][^'"]*?fixtures/(tcMap[A-Z0-9]+)['"]""")
SCAFFOLD_MARKER = "AUTO-GENERATED scaffold for module"


def load_workbook_rows(path: Path) -> dict[str, dict[str, str]]:
    """Map TC-ID → {module, sub, type, title, priority}."""
    rows: dict[str, dict[str, str]] = {}
    if not path.exists():
        return rows
    wb = load_workbook(path, read_only=True)
    ws = wb["Test Cases"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        rows[row[0]] = {
            "module": row[1] or "",
            "sub": row[2] or "",
            "type": row[3] or "",
            "title": row[4] or "",
            "priority": row[11] or "",
        }
    return rows


def load_manual_residue(path: Path) -> dict[str, str]:
    """Parse e2e/web/manual-residue.ts → map TC-ID → rationale.

    The file exports a TypeScript object literal whose keys are TC-IDs and
    values are short rationale strings. We parse line-by-line with a
    tolerant regex — the file's purpose is data, not parsed TS.
    """
    if not path.exists():
        return {}
    text = path.read_text(encoding="utf-8", errors="replace")
    out: dict[str, str] = {}
    entry_re = re.compile(
        r"""['"]?(TC-[A-Z0-9]{2,3}-\d{4})['"]?\s*:\s*['"`]([^'"`]+)['"`]"""
    )
    for m in entry_re.finditer(text):
        out[m.group(1)] = m.group(2).strip()
    return out


def load_results_passes(results_path: Path) -> set[str]:
    """Read a Playwright JSON report and return TC-IDs whose test passed.

    The Playwright JSON reporter writes a structured report whose suite/
    spec/test titles carry the `[TC-XX-NNNN]` prefix our `tc()` helper
    emits. Walk the tree, look at every node with a `results` list, and
    collect TC-IDs from the title if at least one run was passed.
    """
    if not results_path.exists():
        return set()
    try:
        data = json.loads(results_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return set()

    passes: set[str] = set()

    def visit(node: object) -> None:
        if isinstance(node, dict):
            title = node.get("title", "")
            if isinstance(title, str) and title:
                results = node.get("results", []) or []
                if isinstance(results, list) and any(
                    isinstance(r, dict) and r.get("status") == "passed"
                    for r in results
                ):
                    for tc_id in TC_PATTERN.findall(title):
                        passes.add(tc_id)
            for v in node.values():
                visit(v)
        elif isinstance(node, list):
            for item in node:
                visit(item)

    visit(data)
    return passes


def load_tc_maps(root: Path) -> dict[str, set[str]]:
    """Map name (tcMapMM, tcMapAM, …) → TC-IDs the map enumerates."""
    fixtures_dirs = [
        root / "e2e" / "web" / "fixtures",
        root / "e2e" / "desktop" / "fixtures",
    ]
    maps: dict[str, set[str]] = {}
    for fixtures in fixtures_dirs:
        if not fixtures.exists():
            continue
        for f in fixtures.glob("tcMap*.ts"):
            text = f.read_text(encoding="utf-8", errors="replace")
            maps.setdefault(f.stem, set()).update(TC_PATTERN.findall(text))
    return maps


# Per-map `{key: TC-ID}` lookup. Used by the strict scanner to resolve
# `tcMapXX['key']` and `id('key')` references to specific TC-IDs.
_KEY_LOOKUP_CACHE: dict[Path, dict[str, dict[str, str]]] = {}


def load_tc_map_lookups(root: Path) -> dict[str, dict[str, str]]:
    """Map name → {workbook key → TC-ID}."""
    if root in _KEY_LOOKUP_CACHE:
        return _KEY_LOOKUP_CACHE[root]
    fixtures_dirs = [
        root / "e2e" / "web" / "fixtures",
        root / "e2e" / "desktop" / "fixtures",
    ]
    # tcMap fixtures (generated by build_tc_maps.py) quote the workbook key
    # with single quotes by default, double quotes when the key itself
    # contains a `'`, and leave a plain identifier bare (e.g. `Move:`). The
    # TC-ID value may be single- or double-quoted. Match all three key
    # forms + either value quote — matching only `"key": "TC"` (as the
    # original regex did) silently credited nothing, since the generator
    # emits single quotes.
    entry_re = re.compile(
        r"""^\s*(?:'((?:\\.|[^'\\])*)'|"((?:\\.|[^"\\])*)"|(\w+))"""
        r"""\s*:\s*["'](TC-[A-Z0-9]{2,3}-\d{4})["']""",
        re.MULTILINE,
    )
    lookups: dict[str, dict[str, str]] = {}
    for fixtures in fixtures_dirs:
        if not fixtures.exists():
            continue
        for f in fixtures.glob("tcMap*.ts"):
            text = f.read_text(encoding="utf-8", errors="replace")
            entries = lookups.setdefault(f.stem, {})
            for m in entry_re.finditer(text):
                raw = m.group(1) or m.group(2) or m.group(3) or ""
                # Unescape escaped quotes / backslashes only — never run
                # unicode_escape, which mangles real non-ASCII characters.
                key = raw.replace("\\'", "'").replace('\\"', '"').replace("\\\\", "\\")
                entries[key] = m.group(4)
    _KEY_LOOKUP_CACHE[root] = lookups
    return lookups


def strict_ids_in_text(
    text: str,
    tc_maps: dict[str, set[str]],
    tc_map_lookups: dict[str, dict[str, str]] | None = None,
) -> set[str]:
    """Return TC-IDs the spec REALLY covers, by these rules:

      1. Inline literal `tc('TC-XX-NNNN', ...)` / `tcCovered('TC-XX-NNNN')`.
      2. Inline array `tc(['TC-AA-0001', ...], ...)`.
      3. Range `tcRange('TC-XX', from, to, ...)` — expand.
      4. **Real** iteration `for (... of [Object.entries(]tcMapXX[)])` —
         credits all entries (the iteration generates a test() per cell).
      5. Literal-key indexing `tcMapXX['key']` — credits just that key.
      6. Helper-function call `helperName('key')` where `helperName` is a
         locally-defined function that dereferences `tcMapXX[<param>]` —
         credits the TC-ID at that key.

    A bare `void Object.keys(tcMapXX);` (the S13 credit-trigger) is NOT
    credited — it's just a static-analysis marker, not real coverage.
    """
    lookups = tc_map_lookups or {}
    out: set[str] = set()
    # 1. literal tc(...) / tcCovered(...)
    out.update(STRICT_TC_CALL.findall(text))
    # 2. literal tc([...])
    for group in STRICT_TC_LIST.findall(text):
        out.update(TC_PATTERN.findall(group))
    # 3. tcRange
    for prefix, lo, hi in STRICT_TC_RANGE.findall(text):
        lo_i, hi_i = int(lo), int(hi)
        for n in range(lo_i, hi_i + 1):
            out.add(f"{prefix}-{n:04d}")
    # 4. Real iteration over a tcMap — credits all entries.
    for map_name in STRICT_MAP_ITERATION.findall(text):
        if map_name and map_name in tc_maps:
            out |= tc_maps[map_name]
    # 5. Literal-key indexing — credits just that key.
    for map_name, key in STRICT_MAP_LITERAL_INDEX.findall(text):
        if map_name in lookups and key in lookups[map_name]:
            out.add(lookups[map_name][key])
    # 6. Helper-function call — `id('key')` where id resolves to
    #    tcMapXX[key]. Detect the helper definition + its target map.
    helpers: dict[str, str] = {}
    for m in HELPER_DEFINITION.finditer(text):
        helpers[m.group(1)] = m.group(2)
    if helpers:
        for fn_name, key in HELPER_CALL.findall(text):
            if fn_name not in helpers:
                continue
            map_name = helpers[fn_name]
            if map_name in lookups and key in lookups[map_name]:
                out.add(lookups[map_name][key])
    return out


def iter_specs(root: Path):
    for glob in SPEC_GLOBS:
        yield from root.glob(glob)


def scan_specs(
    root: Path, *, strict: bool,
) -> tuple[dict[str, set[str]], dict[str, set[str]], dict[str, set[str]]]:
    """Return three dicts of spec-path → TC-IDs:

      - `live`: TC-IDs covered by a non-scaffold spec.
      - `scaffold`: TC-IDs covered ONLY by an AUTO-GENERATED scaffold.
      - `combined`: union of both.

    In **strict** mode, credit only TC-IDs that appear in an explicit
    `tc('TC-...')` / `tcCovered(...)` / `tcRange(...)` call in the spec
    body. In **lenient** mode, also credit TC-IDs from any `tcMap*`
    imported by the spec.
    """
    # Strict mode also needs the maps (for the map-usage detector); the
    # difference is what counts as "use": strict requires real usage in
    # the spec body, lenient credits a bare import.
    tc_maps = load_tc_maps(root)
    tc_map_lookups = load_tc_map_lookups(root) if strict else {}
    live: dict[str, set[str]] = defaultdict(set)
    scaffold: dict[str, set[str]] = defaultdict(set)
    combined: dict[str, set[str]] = defaultdict(set)
    for spec in iter_specs(root):
        text = spec.read_text(encoding="utf-8", errors="replace")
        rel = spec.relative_to(root).as_posix()
        is_scaffold = SCAFFOLD_MARKER in text
        if strict:
            ids = strict_ids_in_text(text, tc_maps, tc_map_lookups)
        else:
            ids = set(TC_PATTERN.findall(text))
            for map_name in IMPORT_PATTERN.findall(text):
                if map_name in tc_maps:
                    ids |= tc_maps[map_name]
        combined[rel] = ids
        if is_scaffold:
            scaffold[rel] = ids
        else:
            live[rel] = ids
    return live, scaffold, combined


def render_markdown(
    web_rows: dict[str, dict[str, str]],
    desk_rows: dict[str, dict[str, str]],
    automated: set[str],
    automated_by_spec: dict[str, set[str]],
    live_ids: set[str],
    scaffold_ids: set[str],
    residue_ids: dict[str, str],
    *,
    mode: str,
    passed_ids: set[str] | None = None,
) -> str:
    union_ids = set(web_rows) | set(desk_rows)
    residue_set = set(residue_ids) & union_ids

    by_module: dict[str, dict[str, set[str]]] = defaultdict(lambda: {
        "total": set(),
        "live": set(),
        "scaffold": set(),
        "residue": set(),
        "gap": set(),
    })
    for tc_id in union_ids:
        row = web_rows.get(tc_id) or desk_rows.get(tc_id) or {}
        module = row.get("module", "Unknown")
        by_module[module]["total"].add(tc_id)
        # Residue is authoritative — the residue file is an explicit
        # opt-out from automation. Even if `tcMapXX[key]` indexing
        # incidentally credits a residue ID via the map-usage detector,
        # the residue classification wins.
        if tc_id in residue_set:
            by_module[module]["residue"].add(tc_id)
        elif tc_id in live_ids:
            by_module[module]["live"].add(tc_id)
        elif tc_id in scaffold_ids:
            by_module[module]["scaffold"].add(tc_id)
        else:
            by_module[module]["gap"].add(tc_id)

    lines: list[str] = []
    lines.append("# E2E ↔ Manual Test Case Coverage")
    lines.append("")
    lines.append(
        f"Auto-generated by `scripts/e2e_coverage_report.py` ({mode} mode)."
    )
    lines.append("")
    lines.append("## What 'covered' means")
    lines.append("")
    lines.append(
        "- **Strict mode** (`--strict`, CI default): a TC-ID is counted only "
        "when a spec explicitly tags it with `tc('TC-XX-NNNN', ...)`, "
        "`tcCovered('TC-XX-NNNN')`, or `tcRange('TC-XX', from, to, ...)`. "
        "This is the honest \"real test code references this row\" number."
    )
    lines.append(
        "- **Lenient mode** (legacy default): also credits TC-IDs that appear "
        "in any `tcMap*.ts` file imported by the spec. Useful for tracking "
        "which modules are *wired* to automation, but inflates the number — "
        "a spec that imports a 547-entry map gets credit for all 547 even if "
        "it only writes a handful of `test()` bodies."
    )
    lines.append(
        "- **Manual-residue tier** (`e2e/web/manual-residue.ts`): TC-IDs "
        "explicitly excluded from automation — cross-OS installer signing, "
        "real-IdP live tier, perception perf, browser-chrome surfaces. These "
        "do not count against the gap; they're tracked separately and "
        "verified manually."
    )
    lines.append("")
    lines.append(
        "**Always use `--strict` when answering 'is this row really tested?'.**"
    )
    lines.append("")

    lines.append(f"## Summary ({mode} mode)")
    lines.append("")
    total = len(union_ids)
    # Residue is authoritative — subtract it from live/scaffold so the
    # tiers are disjoint and the counts add up to `credited + gap = total`.
    live_count = len((live_ids & union_ids) - residue_set)
    scaffold_only = len(
        (scaffold_ids - live_ids) & union_ids - residue_set
    )
    residue_count = len(residue_set)
    credited = live_count + scaffold_only + residue_count
    pct = (credited / total * 100) if total else 0.0
    pct_live = (live_count / total * 100) if total else 0.0
    lines.append(f"- **Unique manual cases (web + desktop):** {total}")
    lines.append(
        f"- **Live (real test code in non-scaffold specs):** {live_count} "
        f"({pct_live:.1f}%)"
    )
    lines.append(
        f"- **Scaffold-only (`test.fixme` placeholder):** {scaffold_only} "
        "— spec exists, real assertions still pending."
    )
    lines.append(
        f"- **Manual-residue (excluded from automation):** {residue_count} "
        "— cross-OS / installer / browser-chrome / perception."
    )
    lines.append(
        f"- **Total credited (live + scaffold + residue):** {credited} "
        f"({pct:.1f}%)"
    )
    lines.append(
        f"- **Gap (no spec at all):** "
        f"{total - credited}"
    )
    lines.append("")

    if passed_ids is not None:
        passed = len(passed_ids & union_ids)
        pct_pass = (passed / total * 100) if total else 0.0
        lines.append(
            f"From the last Playwright run: **{passed} TC-IDs passed** "
            f"({pct_pass:.1f}%)."
        )
        lines.append("")

    lines.append("## Coverage by module")
    lines.append("")
    lines.append(
        "| Module | Total | Live | Scaffold | Residue | Gap | Live % |"
    )
    lines.append("|---|---:|---:|---:|---:|---:|---:|")
    for module in sorted(by_module, key=lambda m: -len(by_module[m]["total"])):
        b = by_module[module]
        t = len(b["total"])
        li = len(b["live"])
        sc = len(b["scaffold"])
        rs = len(b["residue"])
        gp = len(b["gap"])
        pct_m = (li / t * 100) if t else 0.0
        lines.append(
            f"| {module} | {t} | {li} | {sc} | {rs} | {gp} | {pct_m:.0f}% |"
        )
    lines.append("")

    lines.append("## Automated TC-IDs by spec")
    lines.append("")
    for spec in sorted(automated_by_spec):
        ids = sorted(automated_by_spec[spec])
        if not ids:
            continue
        lines.append(f"### `{spec}`")
        lines.append("")
        lines.append(
            f"Covers **{len(ids)}** manual rows: "
            + ", ".join(ids[:8])
            + (f", … +{len(ids)-8} more" if len(ids) > 8 else "")
        )
        lines.append("")

    if residue_set:
        lines.append("## Manual-residue tier (excluded from automation)")
        lines.append("")
        lines.append(
            "These rows are permanently delegated to manual / cross-OS CI "
            "verification. They're listed here so the coverage report stays "
            "honest about where automated assurance stops."
        )
        lines.append("")
        # Group by rationale for compactness.
        by_reason: dict[str, list[str]] = defaultdict(list)
        for tc_id in sorted(residue_set):
            by_reason[residue_ids.get(tc_id, "(no rationale)")].append(tc_id)
        for reason in sorted(by_reason):
            ids = sorted(by_reason[reason])
            lines.append(f"- _{reason}_")
            lines.append(f"  - {', '.join(ids)}")
        lines.append("")

    lines.append("## Gap — manual rows without automation")
    lines.append("")
    lines.append(
        "Only the first 25 gaps per module are listed below; full data in "
        "`docs/qa/results/e2e-coverage.json` (run with `--json`)."
    )
    lines.append("")
    for module in sorted(by_module, key=lambda m: -len(by_module[m]["gap"])):
        gaps = sorted(by_module[module]["gap"])
        if not gaps:
            continue
        lines.append(f"### {module} ({len(gaps)} gap)")
        lines.append("")
        for tc_id in gaps[:25]:
            row = web_rows.get(tc_id) or desk_rows.get(tc_id) or {}
            title = row.get("title", "")
            prio = row.get("priority", "")
            lines.append(f"- `{tc_id}` — {title} _({prio})_")
        if len(gaps) > 25:
            lines.append(f"- … +{len(gaps)-25} more")
        lines.append("")

    return "\n".join(lines) + "\n"


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--strict",
        action="store_true",
        help=(
            "Credit only TC-IDs referenced by an explicit tc()/tcCovered()/"
            "tcRange() call. Excludes tcMap-import credit."
        ),
    )
    parser.add_argument(
        "--json", action="store_true", help="Also emit machine-readable JSON",
    )
    parser.add_argument(
        "--from-results",
        type=Path,
        default=None,
        help=(
            "Path to a Playwright JSON-reporter output. The report will note "
            "which TC-IDs actually passed in that run."
        ),
    )
    parser.add_argument(
        "--fail-under",
        type=float,
        default=None,
        help=(
            "Exit with code 2 if strict-live %% falls below this floor. "
            "CI uses this to enforce a coverage floor."
        ),
    )
    parser.add_argument(
        "--quiet", "-q", action="store_true", help="Suppress stdout summary",
    )
    args = parser.parse_args(argv)
    mode = "strict" if args.strict else "lenient"

    web_rows = load_workbook_rows(WEB_WORKBOOK)
    desk_rows = load_workbook_rows(DESKTOP_WORKBOOK)
    residue = load_manual_residue(RESIDUE_FILE)
    live_by_spec, scaffold_by_spec, automated_by_spec = scan_specs(
        REPO, strict=args.strict,
    )
    live_ids: set[str] = set().union(*live_by_spec.values()) if live_by_spec else set()
    scaffold_ids: set[str] = set().union(*scaffold_by_spec.values()) if scaffold_by_spec else set()
    automated = live_ids | scaffold_ids

    passed_ids: set[str] | None = None
    if args.from_results is not None:
        passed_ids = load_results_passes(args.from_results)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    md = render_markdown(
        web_rows, desk_rows, automated, automated_by_spec, live_ids, scaffold_ids,
        residue, mode=mode, passed_ids=passed_ids,
    )
    OUT_MD.write_text(md, encoding="utf-8")

    union_ids = set(web_rows) | set(desk_rows)
    residue_set = set(residue) & union_ids
    # Residue is authoritative; subtract from live/scaffold so tiers are
    # disjoint and live + scaffold + residue + gap = total.
    live_count = len((live_ids & union_ids) - residue_set)
    scaf_only = len((scaffold_ids - live_ids) & union_ids - residue_set)
    total = len(union_ids)
    pct_live = (live_count / total * 100) if total else 0.0
    credited = live_count + scaf_only + len(residue_set)
    pct_total = (credited / total * 100) if total else 0.0

    if args.json:
        OUT_JSON.write_text(
            json.dumps(
                {
                    "mode": mode,
                    "total": total,
                    "live": sorted((live_ids & union_ids) - residue_set),
                    "scaffold_only": sorted(
                        (scaffold_ids - live_ids) & union_ids - residue_set
                    ),
                    "manual_residue": sorted(residue_set),
                    "gap": sorted(
                        union_ids - live_ids - scaffold_ids - residue_set
                    ),
                    "automated_by_spec": {
                        k: sorted(v) for k, v in automated_by_spec.items()
                    },
                    "counts": {
                        "live": live_count,
                        "scaffold_only": scaf_only,
                        "manual_residue": len(residue_set),
                        "gap": total - credited,
                        "total": total,
                        "live_pct": round(pct_live, 2),
                        "credited_pct": round(pct_total, 2),
                    },
                    "passed_from_results": (
                        sorted(passed_ids & union_ids)
                        if passed_ids is not None else None
                    ),
                },
                indent=2,
            ),
            encoding="utf-8",
        )

    if not args.quiet:
        print(f"Mode:                                 {mode}")
        print(f"Manual rows (web + desktop, deduped): {total}")
        print(f"Live (real test code):                {live_count} ({pct_live:.1f}%)")
        print(f"Scaffold-only (test.fixme):           {scaf_only}")
        print(f"Manual-residue (excluded):            {len(residue_set)}")
        print(f"Total credited:                       {credited} ({pct_total:.1f}%)")
        print(f"Gap (no spec):                        {total - credited}")
        if passed_ids is not None:
            passed_n = len(passed_ids & union_ids)
            pct_p = (passed_n / total * 100) if total else 0.0
            print(f"Passed from Playwright run:           {passed_n} ({pct_p:.1f}%)")
        print(f"Report written to {OUT_MD.relative_to(REPO)}")

    if args.fail_under is not None and pct_live < args.fail_under:
        print(
            f"::error::Live coverage {pct_live:.2f}% below floor "
            f"{args.fail_under:.2f}%",
            file=sys.stderr,
        )
        return 2

    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
