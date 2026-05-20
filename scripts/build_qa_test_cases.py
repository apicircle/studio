"""Build API Circle Studio Manual Test Cases workbooks (web + desktop) with
combinatorial coverage across every dimension the team cares about:

  - HTTP Method x Body type matrix
  - HTTP Method x Auth type matrix
  - Body content variations per type
  - Variable interpolation surfaces x scopes
  - Changes-to-Push: every entity bucket x every operation
  - Git conflict matrix: every entity bucket x conflict shape x resolution
  - Mock-server response matrix (desktop runtime)
  - Headers deep matrix
  - Workspace restore round-trips from Git
  - History snapshot replay matrix
  - JSON Schema reference scenarios
  - HTTP method edge cases (GET with body, etc.)
  - Security combinatorics
  - Performance scenarios

Outputs:
  docs/qa/web-app-manual-test-cases.xlsx
  docs/qa/desktop-app-manual-test-cases.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
import os

OUT_DIR = r"C:/Local Development/APICircle/studio/docs/qa"

# Styling
FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="4F46E5")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=14)
SECTION_FILL = PatternFill("solid", start_color="6D28D9")
SECTION_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=12)
BASE_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PRIORITY_HIGH_FILL = PatternFill("solid", start_color="FEE2E2")
PRIORITY_MED_FILL = PatternFill("solid", start_color="FEF3C7")
PRIORITY_LOW_FILL = PatternFill("solid", start_color="DCFCE7")
PASS_FILL = PatternFill("solid", start_color="BBF7D0")
FAIL_FILL = PatternFill("solid", start_color="FECACA")
BLOCKED_FILL = PatternFill("solid", start_color="FED7AA")
NOT_RUN_FILL = PatternFill("solid", start_color="E5E7EB")
SKIPPED_FILL = PatternFill("solid", start_color="DBEAFE")

# Modules (code, name)
MODULES = [
    ("WS", "Workspace Management"),
    ("CR", "Collections & Requests"),
    ("RE", "Request Editor"),
    ("BE", "Body Editor"),
    ("BC", "Body Content Variations"),
    ("MM", "Method x Body Matrix"),
    ("ME", "HTTP Method Edge Cases"),
    ("AU", "Authentication"),
    ("AM", "Auth x Method Matrix"),
    ("O2", "OAuth2 Flows"),
    ("VR", "Variables & Environments"),
    ("VI", "Variable Interpolation Matrix"),
    ("HD", "Headers Deep Matrix"),
    ("RP", "Response Panel"),
    ("SC", "Pre-request Scripts & Tests"),
    ("HS", "History"),
    ("HR", "History Replay Matrix"),
    ("CO", "Cookies"),
    ("MK", "Mock Servers"),
    ("MR", "Mock Response Matrix"),
    ("ST", "Settings & Theming"),
    ("IE", "Import / Export"),
    ("GT", "Git Integration"),
    ("CP", "Changes-to-Push View"),
    ("GC", "Git Conflict Matrix"),
    ("WR", "Workspace Restore from Git"),
    ("GQ", "GraphQL"),
    ("JS", "JSON Schema References"),
    ("AS", "Assertions & Execution Plans"),
    ("DC", "Documentation Viewer"),
    ("SE", "Search & Marketplace"),
    ("KB", "Keyboard Shortcuts"),
    ("AL", "Accessibility"),
    ("NW", "Network Conditions"),
    ("PE", "Performance"),
    ("SY", "Security"),
    ("CC", "Cross-Cutting UX"),
    ("WB", "Web-Specific (Browser)"),
    ("DS", "Desktop-Specific (Electron)"),
    ("CL", "CLI"),
    # Senior-QA additions
    ("MC", "MCP (Model Context Protocol)"),
    ("OI", "OAuth2 IdP Compatibility"),
    ("NS", "Network Security / TLS"),
    ("PR", "Proxy Configuration"),
    ("HV", "HTTP Protocol Versions"),
    ("CE", "Compression & Content Encoding"),
    ("SM", "Schema Migration & Versioning"),
    ("MU", "Multi-User Concurrency"),
    ("LO", "Locale & i18n"),
    ("OP", "OS / Platform Compatibility"),
    ("LV", "Linked Workspace Versioning"),
    ("WK", "WebSocket / SSE / Streaming"),
    ("CA", "Caching / ETag / Conditional"),
    ("TP", "Telemetry & Privacy"),
    ("BK", "Backup & Restore"),
    ("CG", "Code Generation"),
]
MODULE_NAME = dict(MODULES)

# Tests: list of (platforms_set, mod, sub, ttype, title, pre, steps, data, expected, prio)
TESTS = []
BOTH = ("web", "desktop")
WEB = ("web",)
DESK = ("desktop",)


def t(plats, mod, sub, ttype, title, pre, steps, data, expected, prio="Medium"):
    TESTS.append((set(plats), mod, sub, ttype, title, pre, steps, data, expected, prio))


# =====================================================================
# WORKSPACE MANAGEMENT
# =====================================================================
def add_workspace_tests():
    t(BOTH, "WS", "Create", "Functional", "Create new local workspace",
      "App is launched.", "1. Workspace switcher.\n2. Create new.\n3. Name 'QA-Smoke'.",
      "Name: QA-Smoke",
      "Workspace created and active; empty explorer; top bar shows name.", "High")
    t(BOTH, "WS", "Create", "Negative", "Reject blank name",
      "Create dialog open.", "1. Empty name.\n2. Create.", "—",
      "Create disabled or validation error; no workspace created.", "Medium")
    t(BOTH, "WS", "Create", "Edge Case", "Unicode + emoji name",
      "Create dialog open.", "1. Name 'Café 🚀 测试 αβγ'.\n2. Create.", "—",
      "Created; name renders correctly in switcher, top bar, persisted state.", "Low")
    t(BOTH, "WS", "Create", "Edge Case", "256-char name truncation/rejection",
      "Create dialog open.", "1. Paste 256-char name.\n2. Create.", "256 chars",
      "Either truncated to documented limit, or clear validation error; UI does not break.", "Low")
    t(BOTH, "WS", "Create", "Edge Case", "Whitespace-only name rejected",
      "Create dialog.", "1. Name '   '.", "—",
      "Validation error; create blocked.", "Low")
    t(BOTH, "WS", "Create", "Negative", "Duplicate workspace name allowed (UUID id)",
      "Workspace 'A' exists.", "1. Create another 'A'.", "—",
      "Allowed by id, but switcher disambiguates via secondary label or recent timestamp.", "Low")
    t(BOTH, "WS", "Switcher", "Functional", "Switch between two workspaces",
      "Workspaces A and B exist.", "1. Open switcher.\n2. Pick B.", "—",
      "Active workspace = B; explorer/envs/history reflect B.", "High")
    t(BOTH, "WS", "Switcher", "Functional", "Recent workspaces persist across restart",
      "3 workspaces opened.", "1. Restart app.\n2. Open switcher.", "—",
      "3 workspaces listed in last-opened order; last-active restored.", "High")
    t(BOTH, "WS", "Switcher", "UX/UI", "Recent list shows last-active first",
      "—", "1. Use 3 workspaces in order A→B→C.\n2. Open switcher.", "—",
      "Order: C, B, A.", "Low")
    t(BOTH, "WS", "Delete", "Functional", "Delete requires confirmation",
      "Workspace 'Disposable' active.", "1. Settings → Delete.\n2. Cancel.", "—",
      "Destructive confirm shown; on cancel, workspace remains.", "High")
    t(BOTH, "WS", "Delete", "Functional", "Confirm deletion removes from registry",
      "Disposable active.", "1. Delete → Confirm.", "—",
      "Removed; app switches to another workspace or empty state; persists across restart.", "High")
    t(BOTH, "WS", "Link to Git", "Functional", "Link unlinked workspace to GitHub repo",
      "Workspace unlinked; GitHub auth ready.", "1. Link.\n2. Authorize.\n3. Pick repo.", "—",
      "Working branch created on remote; Push/Pull enabled; status 'Linked'.", "High")
    t(BOTH, "WS", "Link to Git", "Negative", "OAuth scope denial blocks linking",
      "Unlinked.", "1. Begin link.\n2. Cancel on GitHub consent.", "—",
      "Toast 'authorization cancelled'; workspace unlinked.", "Medium")
    t(BOTH, "WS", "Link to Git", "Negative", "Token revoked surfaces re-auth prompt",
      "Linked; user revoked token on github.com.", "1. Push.", "—",
      "App detects 401; prompts re-authorize; push aborted; no data lost.", "High")
    t(BOTH, "WS", "Link to Git", "Edge Case", "Link to repo without write permission",
      "User has read-only access.", "1. Link to such a repo.", "—",
      "Push fails with clear permission error; workspace usable read-only.", "Medium")
    t(BOTH, "WS", "Push", "Functional", "Push edits to working branch",
      "Linked; local edits.", "1. Push.", "—",
      "Commit pushed; metadata correct; sync snapshot updates; button idle.", "High")
    t(BOTH, "WS", "Push", "Edge Case", "Push with no changes is no-op",
      "Linked; no edits.", "1. Push.", "—",
      "Completes silently or 'No changes to push'.", "Medium")
    t(BOTH, "WS", "Push", "Negative", "Push during offline shows clear error",
      "Network off.", "1. Push.", "—",
      "Toast 'Network unavailable'; local edits preserved.", "High")
    t(BOTH, "WS", "Pull", "Functional", "Pull updates from remote",
      "Remote commit exists.", "1. Pull.", "—",
      "Latest applied; sync snapshot updates; UI reflects new collections/envs.", "High")
    t(BOTH, "WS", "Refresh", "Functional", "Refresh detects retired branch",
      "PR merged; branch deleted.", "1. Refresh.", "—",
      "App detects retirement; prompts new branch or main; retired branch recorded.", "High")
    t(BOTH, "WS", "Reset", "Functional", "Reset discards local edits",
      "Uncommitted edits exist.", "1. Reset to last sync → Confirm.", "—",
      "Edits discarded; returns to last sync; history preserved.", "High")
    t(BOTH, "WS", "Offline", "Functional", "Offline workspace has no git affordances",
      "Switcher open.", "1. Create offline workspace 'Local-Only'.", "—",
      "Push/Pull hidden; Link-to-Git still offered; data in IndexedDB only.", "Medium")
    t(BOTH, "WS", "Hydrate", "Functional", "Passphrase prompt on workspace with secrets",
      "Workspace has secret var; new session.", "1. Open.\n2. Enter correct passphrase.", "—",
      "Modal shown; on correct entry secrets decrypt.", "High")
    t(BOTH, "WS", "Hydrate", "Negative", "Wrong passphrase keeps secrets locked",
      "Secrets + passphrase.", "1. Enter wrong passphrase.", "—",
      "'Incorrect passphrase' error; secrets stay encrypted; retry option.", "High")
    t(BOTH, "WS", "Hydrate", "Edge Case", "Skip passphrase keeps non-secret data usable",
      "Secrets + passphrase.", "1. Dismiss modal.", "—",
      "Workspace loads; secret rows marked locked; non-secret ops normal.", "Medium")
    t(BOTH, "WS", "Restore", "Functional", "Fresh clone of linked repo restores workspace fully",
      "Linked workspace with all features used; delete locally; relink.",
      "1. Delete workspace locally.\n2. Re-link to the repo.\n3. Wait for pull.",
      "—",
      "All collections/folders/requests/envs/plans/mocks/linked-WS/schemas restore identically; secret values empty until passphrase entered.", "High")

    # Web-specific
    t(WEB, "WS", "Multi-Tab", "Edge Case", "Two browser tabs editing same workspace",
      "Tab A open.", "1. Open Tab B with same URL.\n2. Edit in A.\n3. Refresh B.", "—",
      "B reflects A's persisted edits; no IndexedDB corruption; last-write wins on simultaneous same-field edits.", "High")
    t(WEB, "WS", "Refresh", "Functional", "Browser refresh preserves workspace state",
      "Unpushed edits.", "1. F5.", "—",
      "Workspace re-hydrates with edits intact; active request and tab preserved.", "High")
    t(WEB, "WS", "Refresh", "Edge Case", "Refresh during in-flight request",
      "Slow request sending.", "1. F5 mid-send.", "—",
      "Request aborts; no half-saved history corrupts state; can re-send.", "Medium")
    t(WEB, "WS", "Quota", "Edge Case", "IndexedDB quota near full surfaces warning",
      "Quota-limited storage.", "1. Add a large attachment.", "—",
      "Toast 'Storage limit reached'; guidance offered (clear history, export); no silent data loss.", "Medium")
    t(WEB, "WS", "Storage", "Edge Case", "Incognito session limited persistence",
      "Incognito.", "1. Create workspace.\n2. Close window.\n3. Reopen incognito.", "—",
      "Workspace persists for session; user informed data clears on window close.", "Low")
    t(WEB, "WS", "Browser Nav", "Edge Case", "Browser Back does not lose unsaved edits",
      "On workspace view.", "1. Browser Back.", "—",
      "In-app nav OR beforeunload warning; no silent loss.", "Medium")
    t(WEB, "WS", "Cross-Tab Sync", "Edge Case", "Tab B refresh picks up Tab A's push",
      "Tab A pushed.", "1. Tab B → Refresh button.", "—",
      "Tab B fetches latest; UI reflects.", "Medium")

    # Desktop-specific
    t(DESK, "WS", "App Quit", "Functional", "Quit with unsaved edits prompts user",
      "Unpushed edits.", "1. Cmd/Alt+F4 quit.", "—",
      "Warning dialog; on quit, edits persisted; recoverable on next launch.", "High")
    t(DESK, "WS", "Crash Recovery", "Functional", "Force-quit recovers state on launch",
      "Local edits.", "1. Force-kill app.\n2. Relaunch.", "—",
      "Last workspace + edits restored from IndexedDB; banner may note unclean shutdown.", "High")
    t(DESK, "WS", "MCP Path", "Functional", "MCP config path shown",
      "Desktop open.", "1. Settings → MCP / AI Clients.", "—",
      "Snippet displayed with OS-conventional path; copy-to-clipboard works.", "Medium")
    t(DESK, "WS", "Auto-Save", "Functional", "Window state persists across restart",
      "—", "1. Resize and move window.\n2. Quit and relaunch.", "—",
      "Bounds restored to userData/window.json.", "Medium")


add_workspace_tests()


# =====================================================================
# COLLECTIONS & REQUESTS
# =====================================================================
def add_collections_tests():
    t(BOTH, "CR", "Collection", "Functional", "Create collection at root",
      "Workspace open.", "1. + Add → Collection.\n2. Name 'Smoke Suite'.", "—",
      "Added at root, selected, empty; rename input focused.", "High")
    t(BOTH, "CR", "Collection", "Functional", "Rename collection inline",
      "Collection 'Old' exists.", "1. Double-click → 'New'.\n2. Enter.", "—",
      "Name updates in tree and breadcrumbs of open tabs.", "Medium")
    t(BOTH, "CR", "Collection", "Edge Case", "Duplicate name at same level",
      "Collection 'Users' exists.", "1. Create another 'Users' at root.", "—",
      "Either auto-suffix or warning; both exist with unique ids.", "Low")
    t(BOTH, "CR", "Collection", "Functional", "Delete empty collection",
      "Empty 'Trash' exists.", "1. Delete → Confirm.", "—", "Removed from tree.", "Medium")
    t(BOTH, "CR", "Collection", "Functional", "Delete collection cascades to children",
      "Collection with 2 folders + 5 requests.", "1. Delete → Confirm.", "—",
      "Warning mentions child counts; all descendants removed.", "High")
    t(BOTH, "CR", "Collection", "Functional", "Duplicate collection deep-copies tree",
      "Collection with 3 nested folders + 4 requests.", "1. Right-click → Duplicate.", "—",
      "New collection has identical tree; ids new; original unmodified.", "Medium")
    t(BOTH, "CR", "Folder", "Functional", "Create folder under collection",
      "Collection 'API v1' exists.", "1. Add folder 'Users'.", "—",
      "Added; expand/collapse works.", "High")
    t(BOTH, "CR", "Folder", "Edge Case", "5-level nesting",
      "—", "1. Create A→B→C→D→E.", "—",
      "All levels render with distinct indents.", "Low")
    t(BOTH, "CR", "Folder", "Functional", "Folder auth inherited by requests",
      "Folder with Bearer auth.", "1. Create child request set auth=Inherit.\n2. Send.", "—",
      "Child uses folder's Bearer.", "High")
    t(BOTH, "CR", "Request", "Functional", "Create via Ctrl+N",
      "Editor focused; collection selected.", "1. Ctrl/Cmd+N.", "—",
      "GET 'Untitled Request' created under selection; URL focused.", "High")
    t(BOTH, "CR", "Request", "Functional", "Duplicate clones all fields",
      "Request fully configured.", "1. Duplicate.", "—",
      "'(copy)' with cloned URL/method/headers/body/auth/scripts/tests.", "Medium")
    t(BOTH, "CR", "Request", "Functional", "Delete keeps history readable",
      "Request with prior runs.", "1. Delete → Confirm.", "—",
      "Tree removes it; History panel still shows old runs by URL/method.", "Medium")
    t(BOTH, "CR", "Reorder", "Functional", "Drag request between folders",
      "Folders A, B; request R in A.", "1. Drag R from A → B.", "—",
      "R under B; persists across push/pull.", "Medium")
    t(BOTH, "CR", "Reorder", "Functional", "Reorder within folder",
      "Folder with R1, R2, R3.", "1. Drag R3 above R1.", "—",
      "Order R3, R1, R2; persists.", "Medium")
    t(BOTH, "CR", "Reorder", "Edge Case", "Drag folder into descendant blocked",
      "Folder A contains B.", "1. Try to drag A into B.", "—",
      "Operation rejected with clear message; structure unchanged.", "Low")
    t(BOTH, "CR", "Search", "Functional", "Filter tree by substring",
      "Many requests.", "1. Search 'login'.", "—",
      "Matches highlighted; clear restores tree.", "Medium")
    t(BOTH, "CR", "Search", "Edge Case", "Search is case-insensitive",
      "Request 'GetUser'.", "1. Search 'getuser'.", "—", "Matches.", "Low")
    t(BOTH, "CR", "Move", "Functional", "Move request to a different collection",
      "Collections C1, C2; request R in C1.", "1. Cut R; paste under C2 (or drag).", "—",
      "R under C2; folderId/collectionId updated; refs in plans/history still resolve.", "Medium")


add_collections_tests()


# =====================================================================
# REQUEST EDITOR
# =====================================================================
def add_request_editor_tests():
    t(BOTH, "RE", "URL Bar", "Functional", "Send simple GET",
      "Empty request.", "1. URL https://httpbin.org/get.\n2. Send.", "—",
      "200; JSON body; time/size populated.", "High")
    t(BOTH, "RE", "URL Bar", "Functional", "Variable interpolation in URL",
      "Env baseUrl=https://httpbin.org", "1. URL '{{baseUrl}}/get'.\n2. Send.", "—",
      "Hover shows resolved value; Send returns 200.", "High")
    t(BOTH, "RE", "URL Bar", "Negative", "Undefined variable resolves empty",
      "No env.", "1. URL 'https://{{missing}}.example'.\n2. Send.", "—",
      "URL becomes 'https://.example'; clear network error.", "Medium")
    t(BOTH, "RE", "URL Bar", "Edge Case", "Very long URL > 2KB",
      "—", "1. Paste 4KB URL.\n2. Send.", "—",
      "Editor handles without freezing; server may reject — error shown.", "Low")
    t(BOTH, "RE", "URL Bar", "Edge Case", "URL with non-ASCII path",
      "—", "1. URL 'https://httpbin.org/anything/测试'.\n2. Send.", "—",
      "Path percent-encoded; 200.", "Medium")
    t(BOTH, "RE", "URL Bar", "Edge Case", "URL with userinfo (user:pass@host)",
      "—", "1. URL 'https://user:pass@httpbin.org/basic-auth/user/pass'.\n2. Send.", "—",
      "Credentials extracted; sent via Basic auth; 200.", "Low")
    t(BOTH, "RE", "URL Bar", "Edge Case", "Whitespace in URL trimmed",
      "—", "1. Paste '  https://httpbin.org/get  '.\n2. Send.", "—",
      "Leading/trailing whitespace trimmed; succeeds.", "Low")
    t(BOTH, "RE", "URL Bar", "Edge Case", "Empty URL on Send",
      "—", "1. URL empty.\n2. Click Send.", "—",
      "Validation prevents send or shows clear error.", "Medium")
    t(BOTH, "RE", "URL Bar", "Edge Case", "URL with invalid port (>65535)",
      "—", "1. URL 'https://example.com:99999/'.\n2. Send.", "—",
      "Parse error or network error surfaced clearly.", "Low")
    t(BOTH, "RE", "Method", "Functional", "Switch GET → POST",
      "GET request.", "1. Pick POST.", "—",
      "Method updates; Body tab becomes prominent.", "High")
    t(BOTH, "RE", "Method", "Functional", "All standard methods present",
      "Method dropdown.", "1. Inspect.", "—",
      "GET, POST, PUT, PATCH, DELETE, HEAD, OPTIONS all available.", "Medium")
    t(BOTH, "RE", "Params", "Functional", "Add param updates URL",
      "URL https://httpbin.org/get", "1. Params → name=alice.", "—",
      "URL shows ?name=alice; response echoes.", "High")
    t(BOTH, "RE", "Params", "Functional", "Disable a param removes it",
      "Params name=alice, age=30 enabled.", "1. Uncheck age.", "—",
      "URL only ?name=alice; row greyed.", "Medium")
    t(BOTH, "RE", "Params", "Functional", "URL paste populates Params",
      "Params empty.", "1. Paste '?status=open&limit=10'.", "—",
      "Params now status=open, limit=10.", "Medium")
    t(BOTH, "RE", "Params", "Edge Case", "Duplicate param keys preserved",
      "—", "1. Add foo=1, foo=2.\n2. Send.", "—",
      "Both transmitted ?foo=1&foo=2.", "Low")
    t(BOTH, "RE", "Params", "Edge Case", "Param value with special chars URL-encoded",
      "—", "1. Add q='hello world&x=1'.\n2. Send.", "—",
      "Value percent-encoded; server receives literal.", "Low")
    t(BOTH, "RE", "Headers", "Functional", "Add custom header",
      "—", "1. X-Trace-Id=abc123.\n2. Send.", "—",
      "Header sent; response echoes.", "High")
    t(BOTH, "RE", "Headers", "Functional", "Header autocomplete suggests standard names",
      "—", "1. Type 'auth' in key.", "—",
      "Dropdown: Authorization, Authentication-Info, etc.", "Low")
    t(BOTH, "RE", "Headers", "Edge Case", "Variable interpolation in header value",
      "Env token=abc", "1. Authorization: 'Bearer {{token}}'.\n2. Send.", "—",
      "Resolved 'Bearer abc' sent.", "Medium")
    t(BOTH, "RE", "Send", "Functional", "Send via Ctrl+Enter from anywhere",
      "Valid URL.", "1. Anywhere → Ctrl/Cmd+Enter.", "—",
      "Request sent regardless of focus.", "High")
    t(BOTH, "RE", "Send", "Functional", "Cancel in-flight",
      "Slow endpoint.", "1. Send.\n2. Cancel mid-flight.", "—",
      "Aborts; response panel shows Cancelled; no partial body in history.", "Medium")
    t(BOTH, "RE", "Tabs", "UX/UI", "Editor remembers active sub-tab per request",
      "—", "1. A on Body. Open B (any). Open A.", "—",
      "A reopens on Body.", "Low")
    t(BOTH, "RE", "Path Params", "Functional", "URL with :id offers Path Params editor",
      "—", "1. URL 'https://api/users/:id'.", "—",
      "Path Params section appears; setting id=42 yields '…/users/42' on send.", "Medium")
    t(BOTH, "RE", "Path Params", "Functional", "URL with {id} style placeholder",
      "—", "1. URL 'https://api/users/{id}'.", "—",
      "Same Path Params behavior.", "Low")
    # Web-specific
    t(WEB, "RE", "Browser Zoom", "Compatibility", "Ctrl++ zoom keeps layout",
      "—", "1. Ctrl/Cmd++ ×3.", "—", "Layout reflows; nothing clipped.", "Medium")
    t(WEB, "RE", "Tab Title", "UX/UI", "Browser tab title reflects workspace",
      "Smoke active.", "1. Switch to Production.", "—",
      "Tab title updates 'Production · API Circle'.", "Low")
    # Desktop-specific
    t(DESK, "RE", "Native Menu", "Functional", "Edit menu Cut/Copy/Paste in inputs",
      "Text input focused.", "1. Edit menu → Cut/Copy/Paste.", "—",
      "Native behavior on input.", "Medium")
    t(DESK, "RE", "Window Title", "UX/UI", "Window title reflects workspace",
      "Smoke active.", "1. Switch to Production.", "—",
      "Title updates.", "Low")


add_request_editor_tests()


# =====================================================================
# BODY EDITOR (BE)
# =====================================================================
def add_body_editor_tests():
    t(BOTH, "BE", "Type", "Functional", "GET default body is none",
      "New GET.", "1. Open Body.", "-", "Picker defaults 'none'.", "Medium")
    t(BOTH, "BE", "Form Data", "Functional", "Submit text field",
      "POST httpbin/post; form-data.", "1. username=alice. Send.", "-",
      "Multipart sent; response.form.username=alice.", "High")
    t(BOTH, "BE", "Form Data", "Functional", "Upload file row",
      "form-data.", "1. Type=File; pick 1KB file. Send.", "-",
      "files entry present; attachment persisted.", "High")
    t(BOTH, "BE", "Form Data", "Edge Case", "Upload 50MB file",
      "-", "1. Pick 50MB. Send.", "-",
      "Either uploads or shows size-limit error; UI does not freeze.", "Medium")
    t(BOTH, "BE", "Form Data", "Edge Case", "Multiple files in same row",
      "-", "1. Pick 3 files in one row.", "-",
      "All 3 sent if multi supported; else clear single-file behavior.", "Low")
    t(BOTH, "BE", "Form Data", "Edge Case", "Empty value row",
      "-", "1. Add foo='' enabled. Send.", "-",
      "Documented behavior: sent as empty or skipped; consistent.", "Low")
    t(BOTH, "BE", "Form Data", "Edge Case", "Unicode field name",
      "-", "1. Field '测试'='a'. Send.", "-",
      "Encoded per RFC; server echo confirms.", "Low")
    t(BOTH, "BE", "URL-encoded", "Functional", "Submit urlencoded",
      "POST.", "1. a=1, b=hello. Send.", "-",
      "Body 'a=1&b=hello'; Content-Type set.", "High")
    t(BOTH, "BE", "URL-encoded", "Edge Case", "Reserved chars encoded",
      "-", "1. q='a&b=c'. Send.", "-",
      "Value encoded 'a%26b%3Dc'.", "Low")
    t(BOTH, "BE", "Raw JSON", "Functional", "Submit JSON body",
      "POST raw JSON.", "1. Type valid JSON. Send.", "-",
      "Content-Type application/json; httpbin parses.", "High")
    t(BOTH, "BE", "Raw JSON", "Negative", "Invalid JSON shows squiggle",
      "-", "1. Type invalid JSON.", "-",
      "Monaco squiggle; Send still works.", "Medium")
    t(BOTH, "BE", "Raw JSON", "Edge Case", "JSON schema validates body",
      "bodySchemaId set.", "1. Type violating JSON.", "-",
      "Field marked; user not blocked.", "Low")
    t(BOTH, "BE", "Raw XML", "Functional", "Submit XML",
      "POST raw XML.", "1. Valid XML. Send.", "-",
      "Content-Type application/xml.", "Medium")
    t(BOTH, "BE", "Raw HTML", "Functional", "Submit HTML",
      "POST raw HTML.", "1. Valid HTML. Send.", "-", "Content-Type text/html.", "Low")
    t(BOTH, "BE", "Raw Text", "Functional", "Submit plain text",
      "POST raw text.", "1. Hello. Send.", "-", "Content-Type text/plain.", "Low")
    t(BOTH, "BE", "Binary", "Functional", "Upload binary",
      "POST binary.", "1. Pick 10KB PNG. Send.", "-",
      "Bytes sent; Content-Type inferred.", "Medium")
    t(BOTH, "BE", "GraphQL", "Functional", "Send GraphQL query",
      "POST GraphQL.", "1. Query. Send.", "-",
      "Body wraps as query; response returned.", "High")
    t(BOTH, "BE", "GraphQL", "Functional", "Variables sent",
      "Query uses var.", "1. Variables JSON. Send.", "-",
      "Body has query+variables; data returned.", "Medium")
    t(BOTH, "BE", "GraphQL", "Edge Case", "Mutation operation",
      "-", "1. mutation block. Send.", "-",
      "Mutation executed; response correct.", "Low")
    t(BOTH, "BE", "GraphQL", "Edge Case", "Fragments and directives",
      "-", "1. Query with fragment and @skip.", "-",
      "Fragments expanded; directives honored.", "Low")
    t(BOTH, "BE", "Type Switch", "Edge Case", "Switching clears content",
      "Raw JSON has data.", "1. To none.\n2. Back to JSON.", "-",
      "Cleared; new empty editor.", "Low")
    # web/desktop
    t(WEB, "BE", "File Upload", "Compatibility", "Browser file picker",
      "-", "1. Choose File.", "-", "Browser picker opens.", "Medium")
    t(WEB, "BE", "File Upload", "Edge Case", "Drag-drop from OS",
      "-", "1. Drag file onto row.", "-", "Captured via drop event.", "Medium")
    t(DESK, "BE", "File Upload", "Functional", "Native open dialog",
      "-", "1. Choose File.", "-", "Native dialog; remembered folder.", "Medium")
    t(DESK, "BE", "Drag-Drop", "Functional", "Drag from Finder/Explorer",
      "-", "1. Drag from OS.", "-", "Attached identically.", "Medium")


add_body_editor_tests()


# =====================================================================
# AUTHENTICATION (AU)
# =====================================================================
def add_auth_tests():
    t(BOTH, "AU", "Picker", "Functional", "All auth types listed",
      "Auth tab.", "1. Open dropdown.", "-",
      "None, Inherit, Basic, Bearer, API Key, OAuth1, OAuth2 grants, Digest, NTLM, Hawk, AWS SigV4, JWT Bearer, ASAP, EdgeGrid present.", "Medium")
    t(BOTH, "AU", "None", "Functional", "Auth none sends no Authorization",
      "-", "1. Auth=None. Send to httpbin/headers.", "-",
      "No Authorization in echo.", "Medium")
    t(BOTH, "AU", "Bearer", "Functional", "Bearer header sent",
      "httpbin/bearer.", "1. Token. Send.", "-",
      "Authorization: Bearer <t>; 200.", "High")
    t(BOTH, "AU", "Bearer", "Security", "Token masked in UI",
      "-", "1. Enter token.", "-", "Masked; reveal toggle.", "Medium")
    t(BOTH, "AU", "Bearer", "Edge Case", "Empty token does not send empty header",
      "-", "1. Token blank. Send.", "-",
      "No header or warning; not 'Bearer ' with trailing space.", "Low")
    t(BOTH, "AU", "Basic", "Functional", "Base64 credentials",
      "httpbin/basic-auth.", "1. user/pass. Send.", "-",
      "Authorization Basic dXNlcjpwYXNz; 200.", "High")
    t(BOTH, "AU", "Basic", "Edge Case", "Special chars in password",
      "-", "1. password='p@ss:w/ord'. Send.", "-",
      "Encoded; server verifies.", "Low")
    t(BOTH, "AU", "API Key", "Functional", "Header placement",
      "-", "1. X-Api-Key=secret. Send.", "-",
      "Header included.", "High")
    t(BOTH, "AU", "API Key", "Functional", "Query placement",
      "-", "1. api_key=abc. Send.", "-",
      "URL has param; Params tab does not show.", "Medium")
    t(BOTH, "AU", "API Key", "Functional", "Cookie placement",
      "-", "1. Cookie set. Send.", "-",
      "Cookie header includes key.", "Low")
    t(BOTH, "AU", "Digest", "Functional", "Challenge-response",
      "Digest endpoint.", "1. user/pass. Send.", "-",
      "401 -> retry -> 200.", "High")
    t(BOTH, "AU", "Digest", "Edge Case", "stale=true nonce rotation",
      "-", "1. Send.", "-", "Retries with new nonce.", "Low")
    t(BOTH, "AU", "Digest", "Edge Case", "qop=auth-int",
      "-", "1. POST with body.", "-",
      "Hashes include entity-body integrity.", "Low")
    t(BOTH, "AU", "NTLM", "Functional", "3-way handshake",
      "NTLM endpoint.", "1. user/pass/domain. Send.", "-",
      "Negotiate -> Challenge -> Authenticate; MIC; 200.", "High")
    t(BOTH, "AU", "NTLM", "Edge Case", "Without workstation",
      "-", "1. user/pass/domain only.", "-", "Default workstation used.", "Low")
    t(BOTH, "AU", "Hawk", "Functional", "MAC accepted",
      "Hawk endpoint.", "1. Set creds. Send.", "-", "200.", "Medium")
    t(BOTH, "AU", "Hawk", "Edge Case", "Payload hash for POST",
      "-", "1. POST with body.", "-",
      "Authorization includes hash=.", "Low")
    t(BOTH, "AU", "AWS SigV4", "Functional", "Sign GET",
      "AWS endpoint.", "1. Creds + region + service. Send.", "-",
      "AWS4-HMAC-SHA256 header; x-amz-date.", "High")
    t(BOTH, "AU", "AWS SigV4", "Functional", "Sign POST with body",
      "-", "1. Send POST.", "-",
      "x-amz-content-sha256 = SHA256(body).", "Medium")
    t(BOTH, "AU", "AWS SigV4", "Edge Case", "Session token included",
      "-", "1. Add sessionToken. Send.", "-",
      "x-amz-security-token header present.", "Low")
    t(BOTH, "AU", "JWT", "Functional", "HS256",
      "-", "1. Configure HS256. Send.", "-",
      "Bearer <jwt>; verifies.", "Medium")
    t(BOTH, "AU", "JWT", "Functional", "RS256 with PEM",
      "-", "1. Send.", "-", "Verifies with public key.", "Medium")
    t(BOTH, "AU", "JWT", "Functional", "PS256",
      "-", "1. Send.", "-", "PSS padding; verifies.", "Low")
    t(BOTH, "AU", "JWT", "Negative", "Expired exp still sent",
      "-", "1. Send.", "-", "Token unchanged; 401 from server.", "Low")
    t(BOTH, "AU", "JWT", "Edge Case", "Custom claims",
      "-", "1. Add role claim.", "-", "Decoded JWT shows claim.", "Low")
    t(BOTH, "AU", "Inherit", "Functional", "Folder Bearer inherited",
      "-", "1. Child Inherit. Send.", "-", "Folder Bearer applied.", "High")
    t(BOTH, "AU", "Inherit", "Functional", "Override beats folder",
      "-", "1. Child Basic. Send.", "-", "Basic used; folder ignored.", "High")
    t(BOTH, "AU", "Inherit", "Edge Case", "Nested folder walk",
      "-", "1. Coll Bearer; subfolder Inherit; request Inherit. Send.", "-",
      "Walks up to coll Bearer.", "Medium")
    t(BOTH, "AU", "ASAP", "Functional", "ASAP token",
      "-", "1. Send.", "-", "Bearer asap-jwt.", "Low")
    t(BOTH, "AU", "EdgeGrid", "Functional", "EdgeGrid signing",
      "-", "1. Send.", "-", "Signed nonce header.", "Low")
    t(BOTH, "AU", "Custom Header", "Functional", "Arbitrary pair",
      "-", "1. Set Authorization='Token abc'. Send.", "-",
      "Header sent verbatim.", "Medium")


add_auth_tests()


# =====================================================================
# OAUTH2 (O2)
# =====================================================================
def add_oauth2_tests():
    t(BOTH, "O2", "Client Credentials", "Functional", "Acquire token",
      "Mock IdP.", "1. Get token.", "-",
      "access_token + expiresAt + obtainedScope populated.", "High")
    t(BOTH, "O2", "Password", "Functional", "Password grant",
      "-", "1. Get token.", "-",
      "grant_type=password; access_token returned.", "Medium")
    t(BOTH, "O2", "Device Code", "Functional", "Device flow polls",
      "-", "1. Get token. Approve.", "-",
      "user_code shown; polling respected; token stored.", "Medium")
    t(BOTH, "O2", "PKCE", "Functional", "PKCE public client",
      "-", "1. Get token.", "-",
      "code_verifier 43-128 chars; S256 challenge; succeeds.", "High")
    t(BOTH, "O2", "Refresh", "Functional", "Manual refresh",
      "-", "1. Refresh.", "-",
      "grant_type=refresh_token; new access_token.", "High")
    t(BOTH, "O2", "Auto-Refresh", "Functional", "Send refreshes expired",
      "-", "1. Send with expired token.", "-",
      "Refreshed transparently; outgoing uses fresh.", "High")
    t(BOTH, "O2", "Clear", "Functional", "Clear token",
      "-", "1. Clear. Send.", "-",
      "Token removed; re-auth required.", "Medium")
    t(BOTH, "O2", "State", "Security", "State mismatch rejected",
      "-", "1. Tamper state.", "-",
      "Callback rejected; no token stored.", "High")
    t(BOTH, "O2", "Redirect URI", "Negative", "Mismatch error",
      "-", "1. Get token.", "-", "IdP error displayed.", "Medium")
    t(BOTH, "O2", "Private Key JWT", "Functional", "client_assertion",
      "-", "1. Get token.", "-",
      "Token request includes client_assertion JWT.", "Low")
    t(BOTH, "O2", "Scope", "Edge Case", "Narrower granted scope",
      "-", "1. Get token.", "-",
      "obtainedScope reflects narrower scope; UI shows.", "Low")
    t(BOTH, "O2", "Token Display", "Security", "Token masked",
      "-", "1. View token panel.", "-", "Masked; reveal toggle.", "Medium")
    t(BOTH, "O2", "Token Display", "UX/UI", "Expiry countdown",
      "-", "1. Hover token.", "-", "Countdown; warn near expiry.", "Low")
    t(BOTH, "O2", "Error", "Negative", "invalid_grant rendered",
      "-", "1. Get with wrong creds.", "-",
      "Error code+description from IdP shown.", "Medium")
    # Web
    t(WEB, "O2", "Popup", "Functional", "Auth code via popup",
      "-", "1. Get token. Approve.", "-",
      "BroadcastChannel relays code; token stored; popup auto-closes.", "High")
    t(WEB, "O2", "Popup", "Negative", "Popup blocked",
      "-", "1. Get token (blocker on).", "-",
      "Error 'Popup blocked'; instructions.", "High")
    t(WEB, "O2", "BroadcastChannel", "Functional", "Relay delivers code",
      "-", "1. Approve.", "-",
      "Main tab receives state-keyed message; no localStorage write.", "High")
    t(WEB, "O2", "postMessage", "Functional", "Fallback via postMessage",
      "-", "1. Approve.", "-",
      "Falls back; same-origin only.", "Medium")
    t(WEB, "O2", "Cross-Origin", "Security", "Wrong-origin opener rejected",
      "-", "1. Spoof opener.", "-", "Listener verifies origin.", "High")
    t(WEB, "O2", "Auth Code", "Functional", "End-to-end (web)",
      "-", "1. Get token via popup.", "-",
      "tokens stored; UI updates.", "High")
    t(WEB, "O2", "Implicit", "Functional", "Implicit grant fragment",
      "-", "1. Get token.", "-",
      "Token from URL fragment.", "Medium")
    t(WEB, "O2", "Multiple", "Edge Case", "Two simultaneous flows",
      "-", "1. Start A; start B before A.", "-",
      "Each routed by state.", "Medium")
    t(WEB, "O2", "Callback Page", "Security", "No token in DOM",
      "-", "1. Approve. Inspect.", "-",
      "Generic success; no token in DOM.", "Medium")
    # Desktop
    t(DESK, "O2", "Callback Server", "Functional", "Local server captures",
      "-", "1. Get token. Approve in browser.", "-",
      "Free port; redirect captured; token stored; server stops.", "High")
    t(DESK, "O2", "Port", "Edge Case", "Port conflict cycles",
      "-", "1. Get token.", "-", "Next free port chosen.", "High")
    t(DESK, "O2", "Browser Launch", "Functional", "OS default browser",
      "-", "1. Get token.", "-",
      "shell.openExternal opens default browser.", "High")
    t(DESK, "O2", "Server Lifecycle", "Functional", "Stop after capture",
      "-", "1. Get token.", "-", "Server closes; port freed.", "Medium")
    t(DESK, "O2", "Server Lifecycle", "Negative", "Cancelled flow tears down",
      "-", "1. Get token; Cancel.", "-",
      "Server stops within ~10s.", "Medium")
    t(DESK, "O2", "Bind", "Security", "127.0.0.1 only",
      "-", "1. Inspect sockets.", "-",
      "Not 0.0.0.0; unreachable from LAN.", "High")
    t(DESK, "O2", "Multiple", "Edge Case", "Concurrent flows different ports",
      "-", "1. Two quick flows.", "-",
      "Each own port; correct routing.", "Medium")
    t(DESK, "O2", "Fragment Relay", "Functional", "Implicit fragment",
      "-", "1. Get token.", "-",
      "Callback page POSTs fragment to server.", "Medium")


add_oauth2_tests()


# =====================================================================
# VARIABLES & ENVIRONMENTS (VR)
# =====================================================================
def add_variables_tests():
    t(BOTH, "VR", "Env", "Functional", "Create env",
      "-", "1. + new env Dev.", "-", "Env created; empty vars.", "High")
    t(BOTH, "VR", "Env", "Functional", "Activate env",
      "-", "1. Pick Prod.", "-", "Resolutions use Prod values.", "High")
    t(BOTH, "VR", "Env", "Functional", "Duplicate env",
      "Dev has 5 vars.", "1. Duplicate.", "-",
      "'Dev (copy)' with cloned vars.", "Medium")
    t(BOTH, "VR", "Env", "Functional", "Delete env with confirm",
      "-", "1. Delete -> Confirm.", "-", "Removed; fallback active.", "Medium")
    t(BOTH, "VR", "Env", "Functional", "Rename env updates refs",
      "Dev env.", "1. Rename Staging.", "-",
      "References updated.", "Medium")
    t(BOTH, "VR", "Var", "Functional", "Add plaintext var",
      "-", "1. baseUrl=https://api.test.", "-",
      "{{baseUrl}} resolves.", "High")
    t(BOTH, "VR", "Var", "Functional", "Add secret var masked",
      "-", "1. Row secret=topsecret.", "-",
      "Masked; encrypted; not pushed.", "High")
    t(BOTH, "VR", "Var", "Functional", "Toggle plaintext->secret",
      "Plaintext var.", "1. Toggle secret.", "-",
      "Encrypted at rest; UI marks secret.", "Medium")
    t(BOTH, "VR", "Var", "Functional", "Toggle secret->plaintext warns",
      "-", "1. Toggle off.", "-",
      "Warning about exposure; confirm required.", "Medium")
    t(BOTH, "VR", "Var", "Functional", "Autocomplete in URL",
      "-", "1. URL: '{{ba'.", "-", "Popup shows baseUrl.", "Medium")
    t(BOTH, "VR", "Var", "Edge Case", "Special chars in var name",
      "-", "1. Try '{{evil}}'.", "-", "Rejected or literal.", "Low")
    t(BOTH, "VR", "Var", "Edge Case", "Hyphen in name",
      "-", "1. Add my-var.", "-", "{{my-var}} resolves.", "Low")
    t(BOTH, "VR", "Var", "Edge Case", "Very long value (1MB)",
      "-", "1. Reference in body.", "-",
      "Interpolation works; UI does not freeze.", "Low")
    t(BOTH, "VR", "Scope", "Functional", "Request context overrides env",
      "-", "1. Script sets x=req; body uses {{x}}.", "-",
      "Resolves to 'req'.", "Medium")
    t(BOTH, "VR", "Scope", "Functional", "Workspace var fallback",
      "-", "1. Header uses {{app}}.", "-", "'studio'.", "Medium")
    t(BOTH, "VR", "Scope", "Edge Case", "Linked env per priority",
      "-", "1. Send.", "-", "Higher priority wins.", "Medium")
    t(BOTH, "VR", "Resolution", "Edge Case", "Circular ref no loop",
      "-", "1. Use {{a}} where a->b->a.", "-",
      "Terminates; literal or empty.", "Medium")
    t(BOTH, "VR", "Linked Env", "UX/UI", "Drag reorder priority",
      "-", "1. Drag link #2 above #1.", "-", "Order persists.", "Low")
    t(BOTH, "VR", "Linked Env", "Functional", "Consumer override",
      "-", "1. Override token=mine.", "-",
      "Resolves to mine; upstream unchanged.", "Medium")
    # Web
    t(WEB, "VR", "Encryption", "Security", "Passphrase AES-GCM",
      "-", "1. Inspect IndexedDB.", "-",
      "Secret values ciphertext; master JWK wrapped.", "High")
    t(WEB, "VR", "Encryption", "Negative", "Master JWK plaintext in IDB",
      "-", "1. Inspect.", "-",
      "Plaintext JWK present (web limitation); secrets remain wrapped; risk disclosed.", "Medium")
    t(WEB, "VR", "Passphrase", "Functional", "Change passphrase rewraps",
      "-", "1. Change passphrase.", "-",
      "Secrets re-encrypted; old no longer works.", "High")
    # Desktop
    t(DESK, "VR", "Encryption", "Security", "OS keychain wraps JWK",
      "-", "1. Inspect.", "-", "Ciphertext bound to machine.", "High")
    t(DESK, "VR", "Portability", "Edge Case", "IDB copy to other machine",
      "-", "1. Copy; open on B.", "-",
      "Secrets cannot decrypt on B.", "High")
    t(DESK, "VR", "Keychain", "Functional", "macOS keychain prompt",
      "-", "1. Save first secret.", "-",
      "OS prompt; subsequent silent.", "Medium")
    t(DESK, "VR", "Keychain", "Negative", "Keychain unavailable fallback",
      "-", "1. Save secret.", "-",
      "Passphrase fallback used; user informed.", "Medium")


add_variables_tests()


# =====================================================================
# RESPONSE PANEL (RP)
# =====================================================================
def add_response_tests():
    t(BOTH, "RP", "Status", "Functional", "200 OK badge",
      "-", "1. Send 200.", "-", "Green badge; time/size.", "High")
    t(BOTH, "RP", "Status", "Functional", "Non-2xx red badge",
      "-", "1. Send 404.", "-", "Red badge; body rendered.", "Medium")
    t(BOTH, "RP", "Status", "Functional", "5xx error badge",
      "-", "1. Send 500.", "-", "Error badge; body rendered.", "Medium")
    t(BOTH, "RP", "Status", "Functional", "301/302/307/308 redirects shown",
      "-", "1. Send to redirecting endpoint.", "-",
      "Final status; redirect chain optionally inspectable.", "Medium")
    t(BOTH, "RP", "Body Viewer", "Functional", "Pretty/Raw/Preview toggle",
      "JSON resp.", "1. Toggle modes.", "-",
      "Pretty prints; Raw minified; Preview highlighted.", "High")
    t(BOTH, "RP", "Body Viewer", "Edge Case", "Preview cap on large",
      "-", "1. Send >1MB.", "-",
      "Preview capped; download offered; viewer responsive.", "Medium")
    t(BOTH, "RP", "Body Viewer", "Edge Case", "Binary preview shows hex/download",
      "-", "1. Send returning binary.", "-",
      "Hex preview or download CTA; not garbled text.", "Medium")
    t(BOTH, "RP", "Headers", "Functional", "Headers tab",
      "-", "1. Open Headers.", "-",
      "Key-value table; copy works; case preserved.", "Medium")
    t(BOTH, "RP", "Cookies", "Functional", "Set-Cookie populates jar",
      "-", "1. Send with Set-Cookie.", "-",
      "Cookie listed under domain.", "High")
    t(BOTH, "RP", "Transformations", "Functional", "TOON/YAML/CSV savings",
      "JSON resp.", "1. View badges.", "-",
      "Each format shows percent vs minified JSON.", "Medium")
    t(BOTH, "RP", "Transformations", "Functional", "Switch to YAML preview",
      "-", "1. Click 'as YAML'.", "-", "YAML rendered; back to JSON works.", "Low")
    t(BOTH, "RP", "Transformations", "Edge Case", "Savings vs minified not pretty",
      "Pretty JSON shown.", "1. View badges.", "-",
      "Savings computed vs minified bytes (not pretty).", "Medium")
    t(BOTH, "RP", "Download", "Functional", "Download body",
      "-", "1. Click Download.", "-",
      "Save flow opens; bytes match.", "Medium")
    t(BOTH, "RP", "Snapshots", "Functional", "Timeline picker",
      "Multiple runs.", "1. Pick older.", "-",
      "Body viewer swaps; status/time/size update.", "Medium")
    t(BOTH, "RP", "Error", "Negative", "Network error renders error state",
      "-", "1. Send to non-routable.", "-",
      "Clear error; not blank.", "High")
    t(BOTH, "RP", "Render", "Security", "HTML preview sandboxed",
      "-", "1. Send response with <script>.", "-",
      "Script does NOT execute; rendered in sandbox or escaped.", "High")
    t(WEB, "RP", "CORS", "Negative", "Cross-origin without CORS",
      "-", "1. Send to no-CORS endpoint.", "-",
      "Browser blocks; clear actionable error.", "High")
    t(WEB, "RP", "Mixed Content", "Negative", "http from https app",
      "-", "1. Send http endpoint.", "-",
      "Browser blocks; clear message.", "High")
    t(DESK, "RP", "CORS Bypass", "Functional", "No CORS in desktop",
      "-", "1. Same endpoint that fails in web.", "-",
      "Succeeds; response rendered.", "Medium")


add_response_tests()


# =====================================================================
# PRE-REQUEST SCRIPTS & TESTS (SC)
# =====================================================================
def add_scripts_tests():
    t(BOTH, "SC", "Pre-request", "Functional", "pm.variables.set persists",
      "-", "1. Body uses {{token}}; Send.", "-", "'xyz' in body.", "High")
    t(BOTH, "SC", "Pre-request", "Negative", "Runtime error aborts send",
      "-", "1. Script throws.", "-",
      "Error surfaces; HTTP not sent.", "High")
    t(BOTH, "SC", "Console", "Functional", "console.log appears",
      "-", "1. Send.", "-", "Line in console with timestamp.", "Medium")
    t(BOTH, "SC", "Sandbox", "Security", "No access to window",
      "-", "1. Try window.foo=1.", "-",
      "Isolated; no global pollution.", "High")
    t(BOTH, "SC", "Tests", "Functional", "Status==200",
      "-", "1. Add assertion; Send 200.", "-", "Pass; 1/1.", "High")
    t(BOTH, "SC", "Tests", "Functional", "JSON path",
      "-", "1. body.id==42; Send.", "-", "Pass.", "High")
    t(BOTH, "SC", "Tests", "Negative", "Failing assertion red",
      "-", "1. status==200; Send 404.", "-", "Fail; 0/1.", "Medium")
    t(BOTH, "SC", "Tests", "Functional", "Regex matches",
      "-", "1. body.email matches.", "-", "Pass.", "Medium")
    t(BOTH, "SC", "Tests", "Edge Case", "Missing JSON path",
      "-", "1. body.b==2 against {a:1}.", "-",
      "Fails gracefully; actual undefined.", "Medium")
    t(BOTH, "SC", "Tests", "Edge Case", "Duration assertion",
      "-", "1. responseTime lt 1000.", "-",
      "Passes if within 1s; reports ms.", "Low")
    t(BOTH, "SC", "Tests", "Edge Case", "Header presence/value",
      "-", "1. Header contains 'json'.", "-", "Passes if matches.", "Medium")


add_scripts_tests()


# =====================================================================
# HISTORY (HS)
# =====================================================================
def add_history_tests():
    t(BOTH, "HS", "Log", "Functional", "Each send -> entry",
      "-", "1. Send 3.", "-", "3 entries newest first.", "High")
    t(BOTH, "HS", "Buckets", "UX/UI", "Date buckets",
      "-", "1. Open.", "-", "Today/Yesterday/Last 7 days groupings.", "Medium")
    t(BOTH, "HS", "Filter", "Functional", "Filter status range",
      "-", "1. Filter 4xx.", "-", "Only 4xx; counts update.", "Medium")
    t(BOTH, "HS", "Filter", "Functional", "Filter by method",
      "-", "1. POST.", "-", "Only POST.", "Low")
    t(BOTH, "HS", "Filter", "Functional", "Filter URL substring",
      "-", "1. Search 'users'.", "-", "Only matching.", "Low")
    t(BOTH, "HS", "Filter", "Functional", "Date-range filter",
      "-", "1. From-to picker.", "-", "Only entries in range.", "Medium")
    t(BOTH, "HS", "Replay", "Functional", "Replay restores state",
      "-", "1. Click Replay.", "-",
      "Editor restores state; can re-send.", "Medium")
    t(BOTH, "HS", "Delete", "Functional", "Delete one run",
      "-", "1. Delete -> Confirm.", "-", "Removed.", "Low")
    t(BOTH, "HS", "Clear", "Functional", "Clear all",
      "-", "1. Clear -> Confirm.", "-", "Empty state.", "Medium")
    t(BOTH, "HS", "Persistence", "Functional", "Survives reload",
      "-", "1. Reload.", "-", "Entries intact.", "Medium")
    t(BOTH, "HS", "Performance", "Performance", "1000 entries navigable",
      "-", "1. Open.", "-",
      "Smooth scroll; virtualization or pagination.", "Medium")


add_history_tests()


# =====================================================================
# COOKIES (CO)
# =====================================================================
def add_cookies_tests():
    t(BOTH, "CO", "Auto-populate", "Functional", "Set-Cookie stored",
      "-", "1. Send Set-Cookie.", "-",
      "Cookie listed under domain.", "High")
    t(BOTH, "CO", "Auto-send", "Functional", "Cookies sent on next",
      "-", "1. Send same domain.", "-", "Cookie header included.", "High")
    t(BOTH, "CO", "Expiry", "Edge Case", "Expired not sent",
      "-", "1. Send.", "-", "Not included; UI shows expired.", "Medium")
    t(BOTH, "CO", "Manual", "Functional", "Add manually",
      "-", "1. Add example.com tk=abc.", "-", "Persists; sent.", "Medium")
    t(BOTH, "CO", "Manual", "Functional", "Edit value",
      "-", "1. Change tk to xyz.", "-", "Next send uses xyz.", "Low")
    t(BOTH, "CO", "Manual", "Functional", "Toggle enable/disable",
      "-", "1. Disable cookie.", "-",
      "Not sent until re-enabled.", "Low")
    t(BOTH, "CO", "Clear", "Functional", "Clear domain cookies",
      "-", "1. Clear example.com.", "-",
      "Only that domain cleared.", "Medium")
    t(BOTH, "CO", "Path", "Edge Case", "Path matching",
      "-", "1. Cookie path=/api; send /api/v1.", "-",
      "Cookie included.", "Low")
    t(BOTH, "CO", "Secure", "Edge Case", "Secure cookies only over https",
      "-", "1. Secure cookie; send http.", "-",
      "Cookie NOT sent over http.", "Low")
    t(WEB, "CO", "Third-Party", "Compatibility", "Browser may block third-party",
      "-", "1. Cross-origin.", "-",
      "App does not assume; handles gracefully.", "Medium")


add_cookies_tests()


# =====================================================================
# MOCK SERVERS (MK)
# =====================================================================
def add_mock_tests():
    t(BOTH, "MK", "Definition", "Functional", "Create mock server",
      "-", "1. Create -> name 'Users API'.", "-", "Entry in list.", "High")
    t(BOTH, "MK", "Endpoint", "Functional", "Add GET /users/:id",
      "-", "1. Add endpoint; 200 JSON.", "-",
      "Saved; visible.", "High")
    t(BOTH, "MK", "Endpoint", "Edge Case", ":id and {id} both supported",
      "-", "1. Add both styles.", "-",
      "Both accepted; runtime matches.", "Low")
    t(BOTH, "MK", "Response", "Functional", "Multiple responses with selector",
      "-", "1. 200 for X=A; 400 for X=B.", "-",
      "Both saved; runtime picks by rule.", "Medium")
    t(BOTH, "MK", "Spec Import", "Functional", "OpenAPI generates endpoints",
      "-", "1. Import spec.", "-",
      "Endpoints + stubs created.", "Medium")
    t(BOTH, "MK", "Delete", "Functional", "Delete mock",
      "-", "1. Delete -> Confirm.", "-", "Removed.", "Medium")
    t(BOTH, "MK", "Duplicate", "Functional", "Duplicate mock",
      "-", "1. Duplicate.", "-",
      "Copy independent of original.", "Low")
    t(BOTH, "MK", "Rename", "Functional", "Rename mock",
      "-", "1. Rename.", "-",
      "Name updates; references updated.", "Low")
    t(WEB, "MK", "Runtime", "Negative", "Start disabled on web",
      "-", "1. Inspect Start button.", "-",
      "Disabled with tooltip 'Available in desktop app'.", "High")
    t(WEB, "MK", "Runtime", "Edge Case", "Export CLI command",
      "-", "1. 'Copy CLI command'.", "-",
      "Clipboard has 'apicircle mock <ws>'.", "Low")
    t(DESK, "MK", "Runtime", "Functional", "Start mock",
      "-", "1. Start.", "-",
      "Runs on free port; status Running.", "High")
    t(DESK, "MK", "Runtime", "Functional", "Endpoint returns defined response",
      "-", "1. GET /users/1.", "-", "200 with defined body.", "High")
    t(DESK, "MK", "Runtime", "Functional", "Stop mock",
      "-", "1. Stop.", "-", "Server stops; port freed.", "Medium")
    t(DESK, "MK", "Runtime", "Functional", "Multiple mocks concurrently",
      "-", "1. Start both.", "-",
      "Different ports; both reachable.", "Medium")
    t(DESK, "MK", "Runtime", "Edge Case", "Port conflict cycles",
      "-", "1. Default occupied.", "-", "Next free port.", "Medium")
    t(DESK, "MK", "Runtime", "Functional", "Quit stops all mocks",
      "-", "1. Quit.", "-", "Processes stopped; ports freed.", "High")
    t(DESK, "MK", "Runtime", "Edge Case", "Renderer reload preserves mocks",
      "-", "1. View -> Reload.", "-",
      "Mock keeps running; UI reattaches.", "Medium")
    t(DESK, "MK", "Logs", "Functional", "Request logs in panel",
      "-", "1. Hit mock.", "-", "Log row with timestamp/method/path/status.", "Medium")
    t(DESK, "MK", "Logs", "Edge Case", "Large log buffer caps memory",
      "-", "1. 10000 reqs.", "-", "Ring buffer; memory bounded.", "Low")


add_mock_tests()


# =====================================================================
# SETTINGS & THEMING (ST)
# =====================================================================
def add_settings_tests():
    t(BOTH, "ST", "Theme", "Functional", "Dark to light",
      "-", "1. Pick light theme.", "-",
      "UI re-themes; Monaco follows.", "High")
    t(BOTH, "ST", "Theme", "A11y", "High-contrast WCAG 2.1",
      "-", "1. Pick high-contrast.", "-", "Contrast >=7:1.", "Medium")
    t(BOTH, "ST", "Font", "Functional", "Change code font",
      "-", "1. Pick JetBrains Mono.", "-",
      "Monaco updates immediately.", "Low")
    t(BOTH, "ST", "Font Size", "Functional", "Increase UI text",
      "-", "1. Ctrl/Cmd+Shift+= x3.", "-",
      "Scales; persists.", "Medium")
    t(BOTH, "ST", "Font Size", "Functional", "Reset",
      "-", "1. Ctrl/Cmd+Shift+0.", "-", "100%.", "Low")
    t(BOTH, "ST", "Workspace Scoped", "Functional", "Theme per workspace",
      "-", "1. Set in A, B differently.", "-",
      "Each remembers theme.", "Medium")
    t(WEB, "ST", "Browser Zoom", "Compatibility", "Cumulative scale OK",
      "-", "1. Zoom 125% + size 110%.", "-", "Usable layout.", "Low")
    t(DESK, "ST", "Auto-update", "Functional", "Update channel shown",
      "-", "1. Settings -> Updates.", "-",
      "Channel + version + Check Now.", "Medium")


add_settings_tests()


# =====================================================================
# IMPORT / EXPORT (IE)
# =====================================================================
def add_import_export_tests():
    t(BOTH, "IE", "Postman", "Functional", "Postman v2.1 import",
      "-", "1. Import file.", "-",
      "Folders/requests imported; summary lists counts + warnings.", "High")
    t(BOTH, "IE", "Postman", "Edge Case", "Unsupported auth fallback",
      "-", "1. Import.", "-",
      "Auth downgraded with warning.", "Medium")
    t(BOTH, "IE", "cURL", "Functional", "Paste cURL",
      "-", "1. Paste curl.", "-",
      "Request created with all fields.", "High")
    t(BOTH, "IE", "cURL", "Edge Case", "Multi-line cURL with continuations",
      "-", "1. Paste.", "-", "Parsed correctly.", "Low")
    t(BOTH, "IE", "cURL", "Edge Case", "cURL with --data-urlencode",
      "-", "1. Paste.", "-", "Body type x-www-form-urlencoded.", "Low")
    t(BOTH, "IE", "cURL", "Edge Case", "cURL with -F multipart",
      "-", "1. Paste.", "-", "Body type form-data with fields.", "Low")
    t(BOTH, "IE", "Insomnia", "Functional", "Import Insomnia",
      "-", "1. Import.", "-", "Requests/envs imported.", "Medium")
    t(BOTH, "IE", "OpenAPI", "Functional", "Import OpenAPI 3.0 spec",
      "-", "1. Import YAML.", "-",
      "Endpoints + example bodies imported.", "Medium")
    t(BOTH, "IE", "HAR", "Functional", "Import HAR file",
      "-", "1. Import HAR.", "-",
      "Requests imported with method/URL/body.", "Low")
    t(BOTH, "IE", "Export", "Functional", "Export workspace JSON",
      "-", "1. Export.", "-",
      "Valid WorkspaceSynced JSON; secrets excluded.", "Medium")
    t(BOTH, "IE", "Copy cURL", "Functional", "Copy request as cURL",
      "-", "1. Copy as cURL.", "-",
      "Clipboard cURL works in terminal.", "Medium")
    t(BOTH, "IE", "Export", "Security", "Export omits secrets",
      "-", "1. Export.", "-",
      "Secret values empty or [SECRET].", "High")
    t(WEB, "IE", "Download", "Functional", "Browser download",
      "-", "1. Export.", "-", "Download bar shows file.", "Medium")
    t(WEB, "IE", "Upload", "Functional", "Browser file picker",
      "-", "1. Import.", "-", "Picker opens.", "Medium")
    t(DESK, "IE", "Native Dialog", "Functional", "Native open",
      "-", "1. Import.", "-", "OS dialog; filters; remembered folder.", "Medium")
    t(DESK, "IE", "Native Dialog", "Functional", "Native save",
      "-", "1. Export.", "-",
      "OS dialog; suggested name.", "Medium")


add_import_export_tests()


# =====================================================================
# GIT INTEGRATION (GT) - base
# =====================================================================
def add_git_tests():
    t(BOTH, "GT", "Push", "Functional", "Push to working branch",
      "-", "1. Push.", "-",
      "Commit metadata correct; sync snapshot updates.", "High")
    t(BOTH, "GT", "Push Conflict", "Negative", "Concurrent push second fails",
      "-", "1. Device A push; B push without pull.", "-",
      "B fails; prompted to pull.", "High")
    t(BOTH, "GT", "Three-way", "Functional", "Auto-merge non-conflicting",
      "-", "1. Pull with disjoint edits.", "-",
      "Both merged automatically.", "High")
    t(BOTH, "GT", "Three-way", "Functional", "Conflict surfaces resolution UI",
      "-", "1. Pull with same-field conflict.", "-",
      "Modal lists conflicts; mine/theirs/manual.", "High")
    t(BOTH, "GT", "Branch", "Functional", "Switch working branch",
      "-", "1. Switch.", "-",
      "Local synced replaced; local edits warned about.", "Medium")
    t(BOTH, "GT", "Branch", "Edge Case", "Switch with unsaved warns",
      "-", "1. Local edits + Switch branch.", "-",
      "Confirm dialog; stash/discard options.", "Medium")
    t(BOTH, "GT", "Retired", "Functional", "PR-merged cleanup prompt",
      "-", "1. Refresh.", "-",
      "Prompt: new branch or abandon.", "Medium")
    t(BOTH, "GT", "Commit Author", "Functional", "Author matches OAuth identity",
      "-", "1. Inspect commit on GitHub.", "-",
      "Author email/name match.", "Low")
    t(BOTH, "GT", "Network", "Negative", "Push offline",
      "-", "1. Network off; Push.", "-",
      "Toast; local edits preserved.", "High")
    t(BOTH, "GT", "PR Capability", "Functional", "Create PR button when capable",
      "-", "1. Push; look for button.", "-",
      "Opens GitHub compose page.", "Medium")
    t(BOTH, "GT", "Pull Race", "Edge Case", "Pull while typing",
      "-", "1. Click Pull mid-typing.", "-",
      "Pull pauses or warns; no data loss.", "Medium")
    t(BOTH, "GT", "Commit Msg", "UX/UI", "Auto-generated commit message",
      "-", "1. Push edits.", "-",
      "Message summarizes changes ('Modified 3 requests, added 1 env').", "Low")
    t(BOTH, "GT", "Rebase", "Edge Case", "Force-push detection on remote",
      "-", "1. Someone force-pushed; Pull.", "-",
      "App detects history rewrite; warns; offers reset.", "Medium")


add_git_tests()


# =====================================================================
# GRAPHQL (GQ)
# =====================================================================
def add_graphql_tests():
    t(BOTH, "GQ", "Introspect", "Functional", "Fetch schema",
      "-", "1. Body -> GraphQL; Fetch Schema.", "-",
      "Schema saved to globalAssets.", "High")
    t(BOTH, "GQ", "Completions", "Functional", "Field autocomplete",
      "-", "1. Type 'query { us'.", "-",
      "Suggestions include schema fields.", "Medium")
    t(BOTH, "GQ", "Variables", "Functional", "Variables sent",
      "-", "1. Variables JSON; Send.", "-",
      "Body has query+variables; resolved.", "Medium")
    t(BOTH, "GQ", "Schema Reuse", "Functional", "Persisted across reload",
      "-", "1. Reload workspace.", "-",
      "Completions work without re-fetch.", "Low")
    t(BOTH, "GQ", "Introspect", "Negative", "Introspection disabled",
      "-", "1. Fetch Schema.", "-",
      "Clear error; manual upload option.", "Medium")
    t(BOTH, "GQ", "Multi-Operation", "Edge Case", "Operation name picker",
      "-", "1. Doc with 2 ops.", "-",
      "Picker shown; correct op sent.", "Low")
    t(BOTH, "GQ", "Errors", "Negative", "GraphQL errors rendered",
      "-", "1. Send malformed query.", "-",
      "errors[] surfaced clearly in response.", "Medium")


add_graphql_tests()


# =====================================================================
# ASSERTIONS & PLANS (AS)
# =====================================================================
def add_plans_tests():
    t(BOTH, "AS", "Plan Create", "Functional", "Create plan with 3 steps",
      "-", "1. New plan; add 3 steps.", "-",
      "Saved; order; each step shows source request.", "High")
    t(BOTH, "AS", "Plan Run", "Functional", "Run sequentially",
      "-", "1. Run.", "-",
      "Each step runs in order; aggregate visible.", "High")
    t(BOTH, "AS", "Plan Run", "Functional", "Disabled step skipped",
      "-", "1. Run.", "-",
      "Skipped count separated.", "Medium")
    t(BOTH, "AS", "Plan Run", "Functional", "Stop on failure",
      "-", "1. Run.", "-",
      "Subsequent steps 'not run'.", "Medium")
    t(BOTH, "AS", "Plan Report", "Functional", "Per-step results",
      "-", "1. Expand step.", "-",
      "Request, response, assertions.", "Medium")
    t(BOTH, "AS", "Plan Reorder", "Functional", "Drag reorder",
      "-", "1. Drag step 4 to 1.", "-", "Persists.", "Medium")
    t(BOTH, "AS", "Plan Env", "Functional", "Plan-level env priority",
      "-", "1. Run.", "-", "Honors plan priority.", "Medium")
    t(BOTH, "AS", "Plan Loop", "Edge Case", "Same step twice",
      "-", "1. R as steps 1 and 3.", "-",
      "Both tracked separately.", "Low")
    t(BOTH, "AS", "Plan Data", "Edge Case", "Pass data step->step",
      "-", "1. Step 1 extracts token; step 2 uses {{token}}.", "-",
      "Step 2 uses extracted value.", "Medium")


add_plans_tests()


# =====================================================================
# DOCUMENTATION (DC)
# =====================================================================
def add_docs_tests():
    t(BOTH, "DC", "Help", "Functional", "Search topic",
      "-", "1. Search 'oauth'.", "-",
      "Matching topics surface; Markdown renders.", "Medium")
    t(BOTH, "DC", "Help", "Security", "Markdown XSS safe",
      "-", "1. Open crafted topic.", "-",
      "Script tags escaped; no execution.", "High")
    t(BOTH, "DC", "Request Docs", "Functional", "Markdown renders",
      "-", "1. Open Docs tab.", "-",
      "Heading and bullet render.", "Medium")
    t(BOTH, "DC", "External", "UX/UI", "External links open externally",
      "-", "1. Click external link.", "-",
      "Web: new tab. Desktop: OS browser.", "Medium")


add_docs_tests()


# =====================================================================
# SEARCH & MARKETPLACE (SE)
# =====================================================================
def add_search_tests():
    t(BOTH, "SE", "Marketplace", "Functional", "Search public workspaces",
      "-", "1. Search 'stripe'.", "-", "Matching listed.", "Low")
    t(BOTH, "SE", "Marketplace", "Functional", "Link public workspace",
      "-", "1. Add to workspace.", "-",
      "Linked WS appears; collections referenceable.", "Medium")
    t(BOTH, "SE", "Marketplace", "Negative", "Empty results",
      "-", "1. Search 'xyzzy'.", "-", "Empty state message.", "Low")


add_search_tests()


# =====================================================================
# KEYBOARD SHORTCUTS (KB)
# =====================================================================
def add_keyboard_tests():
    t(BOTH, "KB", "Send", "Functional", "Ctrl/Cmd+Enter sends",
      "-", "1. Anywhere; shortcut.", "-", "Request sent.", "High")
    t(BOTH, "KB", "Panels", "Functional", "Ctrl/Cmd+1..9 switches",
      "-", "1. 1, 3, 7.", "-", "Switches; suppressed while typing in inputs.", "Medium")
    t(BOTH, "KB", "Vault", "Functional", "Ctrl/Cmd+K opens Vault",
      "-", "1. Shortcut.", "-", "Right dock Vault.", "Low")
    t(BOTH, "KB", "Refresh", "Functional", "Ctrl/Cmd+Shift+R refresh",
      "-", "1. Shortcut.", "-", "Refresh; not page reload.", "Medium")
    t(BOTH, "KB", "Font Size", "Functional", "Ctrl/Cmd+Shift+= increases",
      "-", "1. x3.", "-", "Scales up.", "Low")
    t(BOTH, "KB", "New Request", "Functional", "Ctrl/Cmd+N",
      "-", "1. Shortcut in Editor.", "-", "New request.", "Medium")
    t(WEB, "KB", "Conflict", "Compatibility", "Ctrl+R reloads browser",
      "-", "1. Shortcut.", "-",
      "Browser reload; use Shift+R for workspace refresh.", "Medium")
    t(WEB, "KB", "Conflict", "Compatibility", "Ctrl+W closes tab",
      "-", "1. Shortcut.", "-", "Tab closes; beforeunload may warn.", "Medium")
    t(WEB, "KB", "Conflict", "Compatibility", "Ctrl+F browser Find",
      "-", "1. Shortcut.", "-", "Browser Find bar.", "Low")
    t(WEB, "KB", "Conflict", "Compatibility", "Ctrl+P browser Print",
      "-", "1. Shortcut.", "-", "Print dialog.", "Low")
    t(DESK, "KB", "Native Menu", "Functional", "Menu accelerators match shortcuts",
      "-", "1. Inspect menus.", "-",
      "Accelerators visible; trigger same actions.", "Medium")
    t(DESK, "KB", "Reload", "Functional", "View Reload preserves state",
      "-", "1. Reload.", "-",
      "Renderer reloads; IndexedDB preserved; mocks survive.", "Medium")
    t(DESK, "KB", "DevTools", "Functional", "Toggle DevTools",
      "-", "1. Cmd/Ctrl+Alt+I.", "-", "Toggles.", "Low")
    t(DESK, "KB", "Quit", "Functional", "Cmd+Q / Alt+F4",
      "-", "1. Quit shortcut.", "-", "App quits cleanly.", "Medium")


add_keyboard_tests()


# =====================================================================
# ACCESSIBILITY (AL)
# =====================================================================
def add_a11y_tests():
    t(BOTH, "AL", "Tab Order", "A11y", "Logical tab order",
      "-", "1. Tab through.", "-", "No focus traps.", "High")
    t(BOTH, "AL", "Focus Ring", "A11y", "Visible focus",
      "-", "1. Tab each.", "-",
      "Purple focus ring; >=3:1 contrast.", "High")
    t(BOTH, "AL", "Screen Reader", "A11y", "Buttons announce purpose",
      "-", "1. SR on; audit buttons.", "-",
      "Clear announcements; state changes announced.", "Medium")
    t(BOTH, "AL", "Color Independence", "A11y", "Status not color-only",
      "-", "1. Inspect Tests.", "-",
      "Icon + color; color-blind friendly.", "Medium")
    t(BOTH, "AL", "Reduced Motion", "A11y", "Respects prefers-reduced-motion",
      "-", "1. Trigger animations.", "-", "Reduced/removed.", "Low")
    t(BOTH, "AL", "Keyboard Only", "A11y", "Full app keyboard-navigable",
      "-", "1. Create-Add-Send no mouse.", "-",
      "All flows reachable.", "High")
    t(BOTH, "AL", "ARIA", "A11y", "Roles match semantics",
      "-", "1. a11y devtools.", "-", "Roles correct.", "Medium")
    t(BOTH, "AL", "Contrast", "A11y", "WCAG AA contrast",
      "-", "1. Contrast checker.", "-",
      "All text passes AA.", "Medium")


add_a11y_tests()


# =====================================================================
# NETWORK CONDITIONS (NW)
# =====================================================================
def add_network_tests():
    t(BOTH, "NW", "Timeout", "Negative", "Default timeout",
      "-", "1. Send to non-responding.", "-",
      "Times out; error 'Request timeout'.", "High")
    t(BOTH, "NW", "Slow", "Edge Case", "10s response",
      "-", "1. Send.", "-",
      "Spinner; renders when received.", "Medium")
    t(BOTH, "NW", "DNS", "Negative", "Unresolvable host",
      "-", "1. Send to invalid host.", "-",
      "DNS resolution failed; guidance.", "Medium")
    t(BOTH, "NW", "Connection", "Negative", "Connection refused",
      "-", "1. Send to closed port.", "-", "Clear error.", "Medium")
    t(BOTH, "NW", "TLS", "Negative", "Self-signed cert",
      "-", "1. Send.", "-",
      "Error with cert details; allow-once option.", "Medium")
    t(BOTH, "NW", "Redirect", "Functional", "Follow 302",
      "-", "1. Send.", "-",
      "Final 200 shown; chain optional.", "Medium")
    t(BOTH, "NW", "Redirect", "Edge Case", "Redirect loop terminated",
      "-", "1. Send.", "-",
      "Detected after N redirects; error.", "Low")
    t(BOTH, "NW", "Streaming", "Functional", "Chunked response",
      "-", "1. Send.", "-",
      "Body progressive or after receive.", "Low")
    t(WEB, "NW", "CORS Preflight", "Functional", "OPTIONS preflight",
      "-", "1. POST with custom header.", "-",
      "Preflight sent; real request if allowed.", "Medium")
    t(WEB, "NW", "Offline", "Negative", "Browser offline mode",
      "-", "1. Offline; Send.", "-", "Clear messaging.", "Medium")
    t(DESK, "NW", "Wi-Fi", "Edge Case", "Switch network mid-flight",
      "-", "1. Switch during send.", "-",
      "Fails or completes; no crash.", "Medium")
    t(DESK, "NW", "Sleep", "Edge Case", "OS sleep during in-flight",
      "-", "1. Sleep; wake.", "-",
      "Either timeout or completion; UI accurate.", "Low")


add_network_tests()


# =====================================================================
# PERFORMANCE (PE)
# =====================================================================
def add_performance_tests():
    t(BOTH, "PE", "Large Workspace", "Performance", "500 requests opens <3s",
      "-", "1. Open.", "-",
      "Hydrate + first paint <3s; tree smooth.", "Medium")
    t(BOTH, "PE", "Large Response", "Performance", "10MB response no freeze",
      "-", "1. Send.", "-",
      "Preview cap; download; responsive UI.", "Medium")
    t(BOTH, "PE", "Many Vars", "Performance", "1000 vars load fast",
      "-", "1. Open Variables.", "-",
      "<1s; filtering responsive.", "Low")
    t(BOTH, "PE", "Many Attachments", "Performance", "50 attachments lazy",
      "-", "1. Open workspace.", "-",
      "Attachments load on demand.", "Low")
    t(BOTH, "PE", "Debounce", "Performance", "Rapid keystrokes coalesced",
      "-", "1. 100 keystrokes fast.", "-",
      "Few writes; final state intact.", "Medium")
    t(BOTH, "PE", "Workspace Switch", "Performance", "<1s switch",
      "-", "1. A -> B.", "-", "No freeze.", "Medium")
    t(BOTH, "PE", "Tree Render", "Performance", "1000-request tree smooth",
      "-", "1. Scroll tree.", "-",
      "Virtualized or paginated; smooth.", "Medium")


add_performance_tests()


# =====================================================================
# SECURITY (SY)
# =====================================================================
def add_security_tests():
    t(BOTH, "SY", "XSS", "Security", "HTML preview sandboxed",
      "-", "1. Response with <script>.", "-",
      "No execution; sandbox iframe.", "High")
    t(BOTH, "SY", "XSS", "Security", "Variable values not HTML-interpreted",
      "-", "1. Var with HTML; render label.", "-",
      "Text only; no execution.", "High")
    t(BOTH, "SY", "Secrets", "Security", "Secrets not in plaintext history",
      "-", "1. Send with Bearer {{token}}.", "-",
      "Authorization redacted in history.", "High")
    t(BOTH, "SY", "Secrets", "Security", "Console redacts secrets",
      "-", "1. Pre-script logs token.", "-",
      "Value redacted in console output.", "Medium")
    t(BOTH, "SY", "Imports", "Security", "Path traversal sanitized",
      "-", "1. Import malicious collection.", "-",
      "Names literal; no FS access.", "High")
    t(BOTH, "SY", "URL", "Security", "javascript: blocked",
      "-", "1. Send javascript: URL.", "-",
      "Blocked with rationale.", "High")
    t(BOTH, "SY", "Headers", "Security", "Auth header masked in cURL preview",
      "-", "1. Open Copy as cURL.", "-",
      "UI preview masked; clipboard has full (with warning).", "Medium")
    t(WEB, "SY", "CSP", "Security", "CSP enforced",
      "-", "1. Inspect headers.", "-",
      "connect-src 'self' + protocols; no unsafe-inline/eval.", "Medium")
    t(WEB, "SY", "Iframe", "Security", "App refuses framing",
      "-", "1. Embed in third-party.", "-",
      "Blocked via X-Frame-Options/CSP.", "Medium")
    t(DESK, "SY", "IPC", "Security", "Untrusted IPC rejected",
      "-", "1. Call from sub-frame.", "-",
      "assertTrustedSender rejects.", "High")
    t(DESK, "SY", "Bind", "Security", "OAuth server 127.0.0.1 only",
      "-", "1. Inspect.", "-", "Not on 0.0.0.0.", "High")
    t(DESK, "SY", "Code Signing", "Security", "First-launch prompt",
      "-", "1. Fresh download.", "-",
      "Gatekeeper/SmartScreen prompt; launches cleanly.", "Medium")


add_security_tests()


# =====================================================================
# CROSS-CUTTING UX (CC)
# =====================================================================
def add_cross_cutting_tests():
    t(BOTH, "CC", "Toasts", "Functional", "Success toast on push",
      "-", "1. Push.", "-", "Toast ~5s; dismissible.", "Medium")
    t(BOTH, "CC", "Toasts", "Functional", "Error toast on network failure",
      "-", "1. Send offline.", "-",
      "Actionable hint.", "High")
    t(BOTH, "CC", "Toasts", "UX/UI", "Stacked dismiss independently",
      "-", "1. 3 toasts quickly.", "-", "Independent timers.", "Low")
    t(BOTH, "CC", "Confirm", "UX/UI", "Destructive uses red",
      "-", "1. Delete.", "-",
      "Red button; Cancel default.", "High")
    t(BOTH, "CC", "Modal", "Functional", "Esc closes non-critical",
      "-", "1. Import modal -> Esc.", "-",
      "Closes; focus returns.", "Medium")
    t(BOTH, "CC", "Modal", "UX/UI", "Backdrop click does not close destructive",
      "-", "1. Delete modal -> click outside.", "-",
      "Modal remains; explicit cancel needed.", "Medium")
    t(BOTH, "CC", "DnD", "Functional", "Reorder env priority",
      "-", "1. Drag row.", "-", "Persists.", "Medium")
    t(BOTH, "CC", "Persistence", "Performance", "250ms debounce",
      "-", "1. 10 keystrokes <250ms.", "-",
      "1 IDB write.", "Medium")
    t(BOTH, "CC", "Error Recovery", "Negative", "Corrupted workspace recovery",
      "-", "1. Corrupt JSON; reopen.", "-",
      "Recovery screen; reset/restore options.", "High")
    t(BOTH, "CC", "Empty States", "UX/UI", "Empty workspace CTA",
      "-", "1. New empty.", "-",
      "'Create your first request' CTA.", "Medium")
    t(BOTH, "CC", "Empty States", "UX/UI", "Empty history",
      "-", "1. Fresh workspace.", "-",
      "Helpful message.", "Low")
    t(BOTH, "CC", "Empty States", "UX/UI", "Empty mocks",
      "-", "1. Fresh.", "-", "Create CTA.", "Low")


add_cross_cutting_tests()


# =====================================================================
# WEB-SPECIFIC (WB)
# =====================================================================
def add_web_specific_tests():
    t(WEB, "WB", "Multi-Tab", "Edge Case", "Two tabs same workspace",
      "-", "1. Open A; open B same URL; edit in A; refresh B.", "-",
      "B reflects A's persisted edits; last-write wins.", "High")
    t(WEB, "WB", "Multi-Tab", "Edge Case", "Two tabs editing same request",
      "-", "1. Both type in body; wait debounce.", "-",
      "Deterministic last-write state.", "Medium")
    t(WEB, "WB", "Tab Close", "Edge Case", "Close tab with unsaved",
      "-", "1. Close.", "-",
      "Edits persisted; restored on reopen.", "Medium")
    t(WEB, "WB", "Refresh", "Functional", "F5 preserves body input",
      "-", "1. Body has content; F5.", "-",
      "Content survives via debounce.", "Medium")
    t(WEB, "WB", "Refresh", "Edge Case", "Hard reload preserves IDB",
      "-", "1. Ctrl+Shift+R.", "-", "IDB intact.", "Medium")
    t(WEB, "WB", "Browser Compat", "Compatibility", "Smoke test Chrome",
      "-", "1. Create-Send-Push.", "-", "No errors.", "High")
    t(WEB, "WB", "Browser Compat", "Compatibility", "Smoke Firefox",
      "-", "1. Smoke.", "-", "Works.", "High")
    t(WEB, "WB", "Browser Compat", "Compatibility", "Smoke Safari",
      "-", "1. Smoke + popup OAuth.", "-",
      "Works including BroadcastChannel.", "High")
    t(WEB, "WB", "Browser Compat", "Compatibility", "Smoke Edge",
      "-", "1. Smoke.", "-", "Works.", "High")
    t(WEB, "WB", "Privacy Mode", "Compatibility", "Incognito",
      "-", "1. Use; close; reopen.", "-",
      "Session-only persistence; informs user.", "Medium")
    t(WEB, "WB", "Quota", "Edge Case", "QuotaExceededError",
      "-", "1. Fill IDB.", "-",
      "Toast; offers cleanup.", "High")
    t(WEB, "WB", "Service Worker", "Functional", "SW doesn't break OAuth callback",
      "-", "1. OAuth flow.", "-",
      "Callback passes through SW.", "Medium")
    t(WEB, "WB", "Mixed Content", "Security", "http://localhost from https",
      "-", "1. Send.", "-",
      "localhost allowed as secure context; else clear error.", "Medium")
    t(WEB, "WB", "Browser Back", "Edge Case", "Back button",
      "-", "1. Press Back.", "-",
      "In-app or warns about unsaved.", "Medium")
    t(WEB, "WB", "Bookmark", "Functional", "Bookmark URL reopens view",
      "-", "1. Bookmark; reopen.", "-",
      "Same view; fallback graceful.", "Low")
    t(WEB, "WB", "DevTools", "Compatibility", "DevTools doesn't break",
      "-", "1. F12.", "-", "No exceptions.", "Low")
    t(WEB, "WB", "Clipboard", "Functional", "Copy cURL via Clipboard API",
      "-", "1. Click Copy.", "-", "writeText succeeds.", "Medium")
    t(WEB, "WB", "Clipboard", "Negative", "Clipboard denied non-secure",
      "-", "1. Click Copy.", "-",
      "Fallback to manual copy.", "Low")
    t(WEB, "WB", "Popup", "Negative", "OAuth popup blocked",
      "-", "1. Get token (blocker).", "-",
      "Error 'Popup blocked'.", "High")
    t(WEB, "WB", "Third-Party Cookies", "Compatibility", "Strict blocking",
      "-", "1. Send cross-origin.", "-",
      "Handles gracefully.", "Medium")
    t(WEB, "WB", "Storage Events", "Edge Case", "No auth leak via localStorage",
      "-", "1. Storage write; observe other tabs.", "-",
      "OAuth callback avoids localStorage.", "High")
    t(WEB, "WB", "Inactive Tab", "Performance", "Background throttling",
      "-", "1. Switch away during request.", "-",
      "Completes on focus or via background timers.", "Low")
    t(WEB, "WB", "Permissions", "Compatibility", "No unwanted OS prompts",
      "-", "1. First load.", "-",
      "No notification/geo/mic prompts.", "Low")
    t(WEB, "WB", "PWA", "Compatibility", "Install as PWA",
      "-", "1. Install app.", "-", "Standalone window works.", "Low")
    t(WEB, "WB", "URL Scheme", "Security", "Dangerous schemes blocked",
      "-", "1. URL data:text/html.", "-", "Blocked or fetch refuses.", "Medium")
    t(WEB, "WB", "Vite Proxy", "Functional", "Dev proxy works",
      "-", "1. Use proxied path.", "-", "Cookies preserved.", "Low")
    t(WEB, "WB", "Visibility", "Edge Case", "Hidden during OAuth callback",
      "-", "1. Hide; approve.", "-",
      "BroadcastChannel delivers regardless.", "Medium")


add_web_specific_tests()


# =====================================================================
# DESKTOP-SPECIFIC (DS)
# =====================================================================
def add_desktop_specific_tests():
    t(DESK, "DS", "Auto-Update", "Functional", "Check on startup",
      "-", "1. Launch.", "-",
      "Background ~5s check; IPC event if available.", "Medium")
    t(DESK, "DS", "Auto-Update", "Functional", "Update banner",
      "-", "1. Launch.", "-",
      "Banner shown; Apply triggers quitAndInstall.", "High")
    t(DESK, "DS", "Auto-Update", "Functional", "Check Now button",
      "-", "1. Click.", "-",
      "Immediate check; result toast.", "Medium")
    t(DESK, "DS", "Auto-Update", "Negative", "Check offline",
      "-", "1. Offline; Check.", "-",
      "Error; app keeps running.", "Medium")
    t(DESK, "DS", "Auto-Update", "Security", "Signature failure aborts",
      "-", "1. Apply tampered update.", "-",
      "Verification fails; abort; current version intact.", "High")
    t(DESK, "DS", "Native Menu", "Functional", "File menu items work",
      "-", "1. Use each.", "-",
      "Trigger same flows.", "Medium")
    t(DESK, "DS", "Native Menu", "Functional", "Edit menu Cut/Copy/Paste/Undo/Redo",
      "-", "1. Use each in input.", "-",
      "Native behavior.", "Medium")
    t(DESK, "DS", "Native Menu", "Functional", "View Reload/DevTools/Zoom",
      "-", "1. Use each.", "-",
      "Reload preserves state; DevTools toggles; Zoom OK.", "Medium")
    t(DESK, "DS", "Native Menu", "Functional", "Window Minimize/Zoom/Close",
      "-", "1. Use each.", "-",
      "Standard behaviors.", "Low")
    t(DESK, "DS", "Native Menu", "Functional", "Help About/Docs/Report",
      "-", "1. Use each.", "-",
      "About shows version; docs external; report opens form.", "Low")
    t(DESK, "DS", "Window State", "Functional", "Bounds persist",
      "-", "1. Resize 1200x800.\n2. Restart.", "-",
      "Restored.", "Medium")
    t(DESK, "DS", "Window State", "Edge Case", "Monitor disconnect clamps",
      "-", "1. Secondary monitor disconnected.\n2. Relaunch.", "-",
      "Window on primary monitor.", "Medium")
    t(DESK, "DS", "Window State", "Functional", "Fullscreen persists",
      "-", "1. Fullscreen; quit; relaunch.", "-",
      "Reopens fullscreen or last bounds.", "Low")
    t(DESK, "DS", "App Quit", "Functional", "Cmd+Q",
      "-", "1. Press.", "-",
      "Windows close; main exits; mocks stopped.", "Medium")
    t(DESK, "DS", "App Quit", "Functional", "Alt+F4",
      "-", "1. Press.", "-", "Quits cleanly.", "Medium")
    t(DESK, "DS", "macOS Dock", "Functional", "Dock click reopens",
      "-", "1. Close window; click dock.", "-",
      "Window reopens.", "Low")
    t(DESK, "DS", "macOS Menu Bar", "Functional", "Persists when window closed",
      "-", "1. Close window.", "-",
      "Menu bar remains; Cmd+N new window.", "Low")
    t(DESK, "DS", "Single Instance", "Functional", "Second launch focuses",
      "-", "1. Launch twice.", "-",
      "Existing focused.", "Medium")
    t(DESK, "DS", "Code Signing", "Compatibility", "macOS Gatekeeper",
      "-", "1. Open.", "-",
      "Prompt; launches.", "Medium")
    t(DESK, "DS", "Code Signing", "Compatibility", "Windows SmartScreen",
      "-", "1. Run installer.", "-",
      "Signed cert resolves; install proceeds.", "Medium")
    t(DESK, "DS", "Linux", "Compatibility", "AppImage/deb runs",
      "-", "1. Run.", "-",
      "Launches; desktop entry/icon.", "Low")
    t(DESK, "DS", "IPC Security", "Security", "Sub-frame IPC rejected",
      "-", "1. Call from non-file://.", "-",
      "Rejected; no privileged action.", "High")
    t(DESK, "DS", "Native Secret", "Functional", "First write prompts keychain",
      "-", "1. Save secret.", "-",
      "Prompt; subsequent silent.", "Medium")
    t(DESK, "DS", "Native Secret", "Negative", "Keychain unavailable fallback",
      "-", "1. Save secret.", "-",
      "Passphrase fallback; user informed.", "Medium")
    t(DESK, "DS", "Mock Manager", "Functional", "Reload doesn't kill mocks",
      "-", "1. Reload.", "-",
      "Mocks survive; UI reattaches.", "Medium")
    t(DESK, "DS", "MCP Bridge", "Functional", "Config snippet copy",
      "-", "1. Click Copy snippet.", "-",
      "Clipboard has JSON snippet.", "Medium")
    t(DESK, "DS", "MCP Bridge", "Functional", "Per-OS path shown",
      "-", "1. Open MCP settings.", "-",
      "Path matches OS convention.", "Low")
    t(DESK, "DS", "First Run", "UX/UI", "First-run wizard",
      "-", "1. Fresh install.", "-",
      "Welcome flow or empty state.", "Low")
    t(DESK, "DS", "Crash", "Negative", "Force kill recover",
      "-", "1. Kill; relaunch.", "-",
      "Last workspace restored.", "High")
    t(DESK, "DS", "Power", "Edge Case", "Sleep during mock run",
      "-", "1. Sleep 10min; wake.", "-",
      "Survives or restarts cleanly.", "Low")
    t(DESK, "DS", "Network", "Edge Case", "Wi-Fi switch during push",
      "-", "1. Switch mid-push.", "-",
      "Fails or retries; edits preserved.", "Medium")
    t(DESK, "DS", "Storage", "Functional", "User data dir shown",
      "-", "1. About.", "-",
      "Path; 'Open folder' link.", "Low")
    t(DESK, "DS", "Window", "Functional", "Windows menu bar auto-hide",
      "-", "1. Press Alt.", "-",
      "Menu appears; hides automatically.", "Low")


add_desktop_specific_tests()


# =====================================================================
# CLI (CL)
# =====================================================================
def add_cli_tests():
    t(DESK, "CL", "Help", "Functional", "apicircle --help",
      "-", "1. Run.", "-", "Usage; exit 0.", "Medium")
    t(DESK, "CL", "Mock", "Functional", "apicircle mock starts",
      "-", "1. Run with workspace.", "-",
      "Server starts; Ctrl+C stops.", "High")
    t(DESK, "CL", "Mock", "Edge Case", "Port in use",
      "-", "1. Run twice.", "-",
      "Second fails clearly; non-zero exit.", "Medium")
    t(DESK, "CL", "Import", "Functional", "apicircle import OpenAPI",
      "-", "1. Run with spec.", "-",
      "Workspace updated; summary; non-zero on error.", "Medium")
    t(DESK, "CL", "MCP", "Functional", "apicircle mcp stdio",
      "-", "1. Run.", "-",
      "Reads stdio; responds to handshake.", "Low")
    t(DESK, "CL", "Secrets", "Functional", "Env var decrypts",
      "-", "1. Run with VAULT_SECRET_KEY.", "-",
      "Secrets decrypted; not logged.", "High")
    t(DESK, "CL", "Validation", "Negative", "Missing workspace",
      "-", "1. Run with bad path.", "-",
      "Clear error; exit 1.", "Low")
    t(DESK, "CL", "Logs", "Functional", "Structured mock logs",
      "-", "1. Hit mock.", "-",
      "Each log: timestamp/method/path/status/duration.", "Low")


add_cli_tests()


# =====================================================================
# COMBINATORIAL GENERATORS
# =====================================================================

HTTP_METHODS = ["GET", "POST", "PUT", "PATCH", "DELETE", "HEAD", "OPTIONS"]
BODY_TYPES = ["none", "form-data", "x-www-form-urlencoded", "raw-json",
              "raw-xml", "raw-html", "raw-text", "binary", "graphql"]

# Methods that, by HTTP semantics, do NOT typically carry a request body.
BODYLESS_METHODS = {"GET", "HEAD", "OPTIONS"}


def method_body_expected(method, body):
    """Return (expected, priority, ttype) for the (method, body) combination."""
    if body == "none":
        if method in BODYLESS_METHODS:
            return ("Request sent with no body; Content-Type not set; server returns expected status.",
                    "High", "Functional")
        else:
            return ("Request sent with no body; Content-Type not set; server may return 411 Length-Required or 200 depending on endpoint.",
                    "Medium", "Functional")
    # Body present
    if method in BODYLESS_METHODS:
        return (f"Client allows configuring a {body} body for {method}; on send, RFC-discouraged combination is transmitted as-is. Verify httpbin/echo endpoint receives the body OR app warns about violation but still sends.",
                "Medium", "Edge Case")
    if body == "form-data":
        return (f"{method} sent as multipart/form-data; boundary set; server echo confirms parsed fields/files.", "High", "Functional")
    if body == "x-www-form-urlencoded":
        return (f"{method} sent with application/x-www-form-urlencoded body; server parses fields.", "High", "Functional")
    if body == "raw-json":
        return (f"{method} sent with application/json body; server parses JSON.", "High", "Functional")
    if body == "raw-xml":
        return (f"{method} sent with application/xml body; server receives XML verbatim.", "Medium", "Functional")
    if body == "raw-html":
        return (f"{method} sent with text/html body; server receives HTML.", "Low", "Functional")
    if body == "raw-text":
        return (f"{method} sent with text/plain body.", "Low", "Functional")
    if body == "binary":
        return (f"{method} sent with binary body and inferred Content-Type from picked file.", "Medium", "Functional")
    if body == "graphql":
        # GraphQL is typically POST; other methods are unusual
        if method == "POST":
            return ("POST sent with application/json body wrapping {query, variables}; server returns GraphQL response.", "High", "Functional")
        return (f"{method} with GraphQL body is unusual. Verify app either prevents this combination, warns, or sends as configured. No silent corruption.",
                "Low", "Edge Case")
    return ("Documented behavior.", "Low", "Functional")


def gen_method_body_matrix():
    for m in HTTP_METHODS:
        for b in BODY_TYPES:
            exp, prio, ttype = method_body_expected(m, b)
            sub = f"{m} x {b}"
            title = f"{m} with body type '{b}'"
            steps = (f"1. New request; method {m}.\n"
                     f"2. Body type {b}; provide minimal valid content.\n"
                     f"3. Send to https://httpbin.org/anything (or equivalent echo).")
            pre = "Editor open."
            data = f"method={m}; body={b}"
            t(BOTH, "MM", sub, ttype, title, pre, steps, data, exp, prio)


gen_method_body_matrix()


# ---- HTTP METHOD EDGE CASES (ME) ----
def gen_method_edge_cases():
    # Each method individually, common edge cases beyond the matrix
    cases = [
        ("GET", "Functional", "Send simple GET to echo endpoint",
         "URL https://httpbin.org/get; method GET.",
         "1. Send.", "200; response includes args/headers.",
         "High"),
        ("GET", "Edge Case", "GET with attempted JSON body sent anyway",
         "Method GET; raw JSON body present.", "1. Send.",
         "Many clients strip body for GET; verify app behavior is documented and consistent. If sent, server may ignore. Important: app should not crash.",
         "Medium"),
        ("POST", "Functional", "POST with empty body",
         "Method POST; body none.", "1. Send to httpbin/post.",
         "Content-Length=0 sent; server returns 200; empty form/data.",
         "Medium"),
        ("POST", "Edge Case", "POST with empty JSON object",
         "Method POST; raw JSON '{}'.", "1. Send.",
         "Body sent as '{}'; Content-Type application/json.", "Low"),
        ("PUT", "Functional", "PUT replaces resource",
         "Endpoint /items/1 expects PUT.", "1. Send PUT with JSON body.",
         "Resource updated.", "Medium"),
        ("PATCH", "Functional", "PATCH partial update",
         "Endpoint expects PATCH.", "1. Send PATCH JSON with single field.",
         "Field updated; other fields untouched.", "Medium"),
        ("DELETE", "Functional", "DELETE without body",
         "Method DELETE.", "1. Send.", "Resource deleted; 204 or 200.", "Medium"),
        ("DELETE", "Edge Case", "DELETE with body (e.g., Elasticsearch)",
         "DELETE + JSON body.", "1. Send to endpoint accepting body.",
         "Body sent; server processes (e.g., bulk delete by query).", "Medium"),
        ("HEAD", "Functional", "HEAD returns headers only",
         "Method HEAD.", "1. Send to GET endpoint as HEAD.",
         "Status + headers shown; body empty.", "Medium"),
        ("HEAD", "Edge Case", "HEAD with body configured",
         "HEAD + body.", "1. Send.",
         "App should either strip body or warn; no crash.", "Low"),
        ("OPTIONS", "Functional", "OPTIONS lists allowed methods",
         "Method OPTIONS.", "1. Send to CORS-aware endpoint.",
         "Response shows Allow header.", "Medium"),
        ("OPTIONS", "Edge Case", "OPTIONS preflight headers shown",
         "OPTIONS with Origin/Access-Control-Request-* headers.", "1. Send.",
         "Response Access-Control-Allow-* headers visible.", "Medium"),
    ]
    for method, ttype, title, pre, steps, expected, prio in cases:
        t(BOTH, "ME", method, ttype, title, pre, steps, f"method={method}", expected, prio)
    # Method casing
    t(BOTH, "ME", "Casing", "Edge Case", "Lowercase method 'get' normalized",
      "Method 'get' (lower).", "1. Send.", "method=get",
      "Either normalized to GET or sent as-is; server typically case-insensitive.", "Low")
    # Custom method (if supported)
    t(BOTH, "ME", "Custom", "Edge Case", "Custom HTTP method 'PURGE'",
      "Method PURGE (Varnish-style).", "1. Send.", "method=PURGE",
      "If supported, sent verbatim; otherwise picker disallows.", "Low")


gen_method_edge_cases()


# ---- AUTH x METHOD MATRIX (AM) ----
AUTH_TYPES = [
    ("Bearer", "High"),
    ("Basic", "High"),
    ("ApiKeyHeader", "Medium"),
    ("ApiKeyQuery", "Medium"),
    ("Digest", "Medium"),
    ("NTLM", "Medium"),
    ("Hawk", "Low"),
    ("AwsSigV4", "High"),
    ("JwtHS256", "Medium"),
    ("JwtRS256", "Medium"),
    ("OAuth2ClientCreds", "High"),
    ("OAuth2AuthCode", "High"),
    ("OAuth2PKCE", "High"),
    ("OAuth2Refresh", "Medium"),
    ("Asap", "Low"),
    ("EdgeGrid", "Low"),
    ("CustomHeader", "Medium"),
]

# Representative methods for auth combinatorics (full Cartesian would explode)
AUTH_METHODS = ["GET", "POST", "PUT", "DELETE"]


def gen_auth_method_matrix():
    for auth, base_prio in AUTH_TYPES:
        for m in AUTH_METHODS:
            sub = f"{auth} x {m}"
            title = f"{auth} authentication on {m}"
            steps = (f"1. New {m} request to a representative endpoint.\n"
                     f"2. Auth type {auth}; configure minimum credentials.\n"
                     f"3. Send and inspect outgoing request.")
            pre = "Editor open."
            if auth == "AwsSigV4" and m in ("POST", "PUT"):
                expected = ("AWS SigV4 signs the body: x-amz-content-sha256 reflects "
                            "SHA256(payload); signature verifies; server returns expected status.")
                prio = "High"
                ttype = "Functional"
            elif auth.startswith("OAuth2"):
                expected = (f"OAuth2 access token obtained (if not present); "
                            f"Authorization: Bearer <token> attached; {m} succeeds.")
                prio = base_prio
                ttype = "Functional"
            elif auth == "Digest":
                expected = (f"401 challenge -> retry with response hash -> {m} succeeds.")
                prio = "Medium"
                ttype = "Functional"
            elif auth == "NTLM":
                expected = (f"NTLM 3-way handshake completes; {m} request authenticated; "
                            f"if the request body would be sent before authentication is complete, "
                            f"client handles re-send correctly.")
                prio = "Medium"
                ttype = "Functional" if m in ("GET", "POST") else "Edge Case"
            elif auth in ("Bearer", "Basic", "ApiKeyHeader", "CustomHeader"):
                expected = (f"Authorization or custom header attached; {m} request includes it; "
                            f"server returns 200 for valid creds.")
                prio = base_prio
                ttype = "Functional"
            elif auth == "ApiKeyQuery":
                expected = (f"API key appended to query string on {m}; URL retains other params.")
                prio = base_prio
                ttype = "Functional"
            elif auth == "Hawk":
                expected = (f"Hawk MAC computed including method, URL, and body (POST/PUT) "
                            f"or empty payload (GET/DELETE); server verifies.")
                prio = "Low"
                ttype = "Functional"
            elif auth.startswith("Jwt"):
                expected = (f"JWT signed and attached as Bearer; {m} succeeds.")
                prio = "Medium"
                ttype = "Functional"
            elif auth in ("Asap", "EdgeGrid"):
                expected = (f"Auth scheme applied to {m}; server accepts.")
                prio = "Low"
                ttype = "Functional"
            else:
                expected = (f"Auth applied to {m} correctly.")
                prio = base_prio
                ttype = "Functional"
            t(BOTH, "AM", sub, ttype, title, pre, steps, f"auth={auth}; method={m}",
              expected, prio)


gen_auth_method_matrix()


# ---- BODY CONTENT VARIATIONS (BC) ----
def gen_body_content_variations():
    # JSON variations
    json_cases = [
        ("Empty object {}", "POST raw JSON.", '1. Body: {}', "Sent as '{}'.", "Medium", "Functional"),
        ("Empty array []", "POST raw JSON.", '1. Body: []', "Sent as '[]'.", "Medium", "Functional"),
        ("Single value (string)", "-", '1. Body: \"hello\"',
         "Sent as JSON string; Content-Type application/json.", "Low", "Edge Case"),
        ("Single value (number)", "-", '1. Body: 42',
         "Sent as JSON number.", "Low", "Edge Case"),
        ("Single value (boolean)", "-", '1. Body: true', "Sent verbatim.", "Low", "Edge Case"),
        ("Single value (null)", "-", '1. Body: null', "Sent verbatim.", "Low", "Edge Case"),
        ("Deeply nested object (10 levels)", "-",
         "1. Build nested JSON 10 levels deep.",
         "Renders without performance issues; sent intact.", "Medium", "Edge Case"),
        ("Large flat object (1000 keys)", "-",
         "1. Generate 1000-key JSON.",
         "Editor remains responsive; sent intact.", "Medium", "Performance"),
        ("Large array (10000 entries)", "-",
         "1. Generate JSON array of 10000 items.",
         "Editor responsive; preview cap may apply.", "Medium", "Performance"),
        ("Unicode key", "-", '1. {\"测试\":\"v\"}',
         "Keys/values preserve unicode bytes; server echo confirms.", "Low", "Edge Case"),
        ("Emoji in value", "-", '1. {\"a\":\"🚀\"}',
         "Emoji bytes preserved.", "Low", "Edge Case"),
        ("Escape sequences", "-", r'1. {"a":"line1\nline2\tx"}',
         "Escape sequences preserved.", "Medium", "Edge Case"),
        ("Trailing comma (invalid JSON)", "-", '1. {\"a\":1,}',
         "Squiggle; send still attempts (server likely rejects).", "Medium", "Negative"),
        ("Comments (invalid JSON)", "-", '1. {\"a\":1 /* c */}',
         "Squiggle; non-standard JSON; send is user responsibility.", "Low", "Negative"),
        ("BOM at start", "-", "1. Paste BOM + JSON.",
         "Either auto-stripped or sent; consistent behavior.", "Low", "Edge Case"),
        ("JSON with very large number (>2^53)", "-",
         '1. Body {\"id\":9007199254740993}.',
         "Sent verbatim; precision warning shown if applicable.", "Low", "Edge Case"),
        ("JSON with NaN/Infinity (non-standard)", "-",
         '1. Body {\"x\":NaN}', "Invalid JSON; squiggle.", "Low", "Negative"),
        ("Variable interpolation in JSON value", "Env name=alice.",
         '1. Body {\"name\":\"{{name}}\"}.', "Resolves to alice.", "Medium", "Functional"),
        ("Variable interpolation in JSON key", "Env k=greeting.",
         '1. Body {\"{{k}}\":\"hi\"}', "Resolves to {greeting:hi}.", "Low", "Edge Case"),
    ]
    for title, pre, steps, expected, prio, ttype in json_cases:
        t(BOTH, "BC", "JSON", ttype, title, pre, steps, "raw JSON", expected, prio)

    # XML variations
    xml_cases = [
        ("Well-formed minimal", "-", "1. <a/>.", "Sent verbatim.", "Low", "Functional"),
        ("With XML declaration", "-", "1. <?xml version='1.0'?><root/>.",
         "Sent verbatim; encoding respected.", "Medium", "Functional"),
        ("With namespaces", "-", "1. Namespaced XML.", "Sent.", "Low", "Functional"),
        ("With CDATA", "-", "1. <r><![CDATA[<x>]]></r>.", "CDATA preserved.", "Low", "Edge Case"),
        ("Malformed XML", "-", "1. <a><b></a>.", "Squiggle; send still attempts.", "Medium", "Negative"),
        ("Large XML (1MB)", "-", "1. Big XML.", "Editor responsive.", "Low", "Performance"),
        ("Special chars in attribute", "-", '1. <a x=\"a&amp;b\"/>.', "Sent verbatim.", "Low", "Edge Case"),
        ("Variable in element value", "Env x=v.", "1. <a>{{x}}</a>.", "Resolves to v.", "Medium", "Functional"),
    ]
    for title, pre, steps, expected, prio, ttype in xml_cases:
        t(BOTH, "BC", "XML", ttype, title, pre, steps, "raw XML", expected, prio)

    # Form-data variations
    fd_cases = [
        ("Single text field", "-", "1. a=1.", "Sent.", "High", "Functional"),
        ("Multiple text fields", "-", "1. a=1, b=2, c=3.", "All sent.", "Medium", "Functional"),
        ("Text + file mixed", "-", "1. a=1; b=file.png.",
         "Both transmitted; multipart correct.", "High", "Functional"),
        ("Same key twice", "-", "1. a=1, a=2.", "Both transmitted.", "Low", "Edge Case"),
        ("Disabled row ignored", "-", "1. a=1 disabled; b=2.", "Only b sent.", "Medium", "Functional"),
        ("Empty value", "-", "1. a=''.", "Sent as empty.", "Low", "Edge Case"),
        ("Unicode key/value", "-", "1. 测试=值.", "Encoded; server confirms.", "Low", "Edge Case"),
        ("File with unicode filename", "-", "1. Pick '文件.txt'.",
         "Filename preserved or encoded per RFC.", "Low", "Edge Case"),
        ("File with no extension", "-", "1. Pick file with no extension.",
         "Content-Type fallback (octet-stream).", "Low", "Edge Case"),
        ("Variable in field value", "Env n=alice.", "1. name={{n}}.", "Resolves to alice.", "Medium", "Functional"),
    ]
    for title, pre, steps, expected, prio, ttype in fd_cases:
        t(BOTH, "BC", "FormData", ttype, title, pre, steps, "form-data", expected, prio)

    # Binary variations
    bin_cases = [
        ("Small text file (1KB)", "-", "1. Send.", "Bytes match.", "High", "Functional"),
        ("Image (PNG 100KB)", "-", "1. Send.", "Content-Type image/png.", "Medium", "Functional"),
        ("PDF (1MB)", "-", "1. Send.", "Content-Type application/pdf.", "Medium", "Functional"),
        ("Empty file (0 bytes)", "-", "1. Pick empty. Send.",
         "Either allows with warning or blocks; consistent.", "Low", "Edge Case"),
        ("Locked file (Windows in-use)", "-", "1. Pick a file open in another process.",
         "OS error surfaced clearly.", "Low", "Negative"),
    ]
    for title, pre, steps, expected, prio, ttype in bin_cases:
        t(BOTH, "BC", "Binary", ttype, title, pre, steps, "binary", expected, prio)

    # URL-encoded variations
    ue_cases = [
        ("Single pair", "-", "1. a=1.", "Body 'a=1'.", "High", "Functional"),
        ("Reserved chars value", "-", "1. q='a&b=c'.", "Encoded.", "Medium", "Edge Case"),
        ("Empty value", "-", "1. a=.", "Body 'a='.", "Low", "Edge Case"),
        ("Variable in value", "Env x=alpha.", "1. a={{x}}.", "Body 'a=alpha'.", "Medium", "Functional"),
    ]
    for title, pre, steps, expected, prio, ttype in ue_cases:
        t(BOTH, "BC", "Urlencoded", ttype, title, pre, steps, "urlencoded", expected, prio)

    # GraphQL variations
    gql_cases = [
        ("Simple query", "-", "1. Send.", "200 with data.", "High", "Functional"),
        ("Query with variables", "-", "1. Send.", "data resolves.", "High", "Functional"),
        ("Mutation", "-", "1. Send.", "Mutation runs.", "Medium", "Functional"),
        ("Fragment composition", "-", "1. Send.", "Fragment expanded.", "Medium", "Functional"),
        ("Directives @skip/@include", "-", "1. Send.", "Directives honored.", "Low", "Edge Case"),
        ("Operation name picker", "Multi-op doc.", "1. Pick op; Send.",
         "Selected op sent.", "Medium", "Functional"),
        ("Variables with nested object", "-", "1. {\"input\":{\"name\":\"a\"}}.",
         "Nested vars sent.", "Medium", "Functional"),
        ("Invalid syntax", "-", "1. Send malformed.", "GraphQL errors[] rendered.", "Medium", "Negative"),
    ]
    for title, pre, steps, expected, prio, ttype in gql_cases:
        t(BOTH, "BC", "GraphQL", ttype, title, pre, steps, "graphql", expected, prio)


gen_body_content_variations()


# ---- VARIABLE INTERPOLATION MATRIX (VI) ----
INTERPOLATION_SURFACES = [
    "URL path", "URL query value", "Header value", "Header key (rare)",
    "JSON body value", "JSON body key", "Form-data value", "Form-data key",
    "Auth Basic username", "Auth Basic password", "Auth Bearer token",
    "Auth API Key value", "Pre-request script body", "Test assertion expected",
    "Cookie value",
]
INTERPOLATION_SCOPES = [
    "Workspace var", "Env var (active)", "Env var (linked higher priority)",
    "Env var (linked lower priority)", "Secret var (plaintext after decrypt)",
    "Request context var (pm.variables.set)", "Linked workspace override",
]


def gen_variable_interpolation_matrix():
    for surface in INTERPOLATION_SURFACES:
        for scope in INTERPOLATION_SCOPES:
            sub = f"{surface} <- {scope}"
            ttype = "Functional"
            prio = "Medium"
            # Boost interesting combos
            if "Secret" in scope and "URL" in surface:
                ttype = "Security"; prio = "High"
            elif "Linked" in scope:
                prio = "Medium"
            elif "script" in surface.lower():
                prio = "Medium"
            title = f"Interpolate {scope} in {surface}"
            steps = (f"1. Define a value in {scope}.\n"
                     f"2. Reference it via {{{{var}}}} in {surface}.\n"
                     f"3. Send.")
            pre = "Editor open; workspace has the var configured per scope."
            expected = (f"Variable resolves correctly into {surface}; sent request reflects the resolved value; "
                        f"resolution precedence honored.")
            t(BOTH, "VI", sub, ttype, title, pre, steps, f"{scope}->{surface}", expected, prio)

    # Edge cases
    edges = [
        ("Undefined var", "Functional", "Undefined var resolves empty",
         "1. URL uses {{missing}}.\n2. Send.",
         "Becomes empty string; request attempted; error if URL invalid.", "Medium"),
        ("Whitespace inside braces", "Edge Case", "{{ var }} with spaces",
         "1. URL uses {{ baseUrl }}.\n2. Send.",
         "Trimmed and resolved or rejected; consistent.", "Low"),
        ("Adjacent variables", "Functional", "{{a}}{{b}} concatenation",
         "1. Header 'A: {{a}}{{b}}'.",
         "Resolves to a+b concatenated.", "Medium"),
        ("Triple braces literal", "Edge Case", "{{{a}}} parsed correctly",
         "1. Body '{{{a}}}'.", "Either '{<value>}' or treated literal; consistent.", "Low"),
        ("Var inside string vs JSON key", "Edge Case", "Quoted vs unquoted contexts",
         "1. JSON {\"x\":\"{{v}}\"} vs {\"{{v}}\":1}.",
         "Both resolve; quoting preserved.", "Medium"),
    ]
    for sub, ttype, title, steps, expected, prio in edges:
        t(BOTH, "VI", sub, ttype, title, "-", steps, "interpolation", expected, prio)


gen_variable_interpolation_matrix()


# ---- HEADERS DEEP MATRIX (HD) ----
def gen_headers_deep():
    standard_headers = [
        "Accept", "Accept-Encoding", "Accept-Language", "Authorization",
        "Cache-Control", "Content-Type", "Content-Length", "Cookie",
        "Origin", "Referer", "User-Agent", "X-Forwarded-For", "If-Match",
        "If-None-Match", "If-Modified-Since", "Range", "Connection",
    ]
    for h in standard_headers:
        t(BOTH, "HD", "Standard", "Functional",
          f"Send standard header '{h}'",
          "Editor open.",
          f"1. Headers add '{h}' with appropriate value.\n2. Send to httpbin/headers.",
          f"header={h}",
          f"Header transmitted; server echo confirms; case preserved or normalized per HTTP spec.",
          "Medium" if h in ("Authorization", "Content-Type") else "Low")

    # Edge cases
    edge_cases = [
        ("Case insensitivity", "Edge Case",
         "Header key case preserved/normalized",
         "1. Add 'x-trace-id' and 'X-TRACE-ID' as separate rows.\n2. Send.",
         "App handles per HTTP semantics; either merged (RFC 7230 sec 3.2) or both sent.", "Medium"),
        ("Duplicate header keys", "Edge Case",
         "Same header twice",
         "1. Add Accept: text/html and Accept: application/json.\n2. Send.",
         "Both sent (combined with comma) or both rows transmitted; server-side semantics preserved.", "Medium"),
        ("Empty value", "Edge Case",
         "Header with empty value",
         "1. Add X-Empty=''.\n2. Send.",
         "Either not sent or sent with empty value; consistent behavior.", "Low"),
        ("Whitespace value", "Edge Case",
         "Header value trimmed",
         "1. Add X-Foo='  bar  '.\n2. Send.",
         "Trimmed to 'bar' or sent with spaces; consistent.", "Low"),
        ("Newline in value (CRLF injection)", "Security",
         "Reject newline in header value",
         "1. Try value 'a\\r\\nX-Injected: 1'.\n2. Send.",
         "Newline stripped/escaped; injection prevented; toast warning may appear.", "High"),
        ("Null byte in header", "Security",
         "Reject null byte",
         "1. Value contains null byte.",
         "Rejected with clear error.", "High"),
        ("Very long header value (8KB)", "Edge Case",
         "Long value sent",
         "1. Add header with 8KB value.\n2. Send.",
         "Either sent or rejected per server limit; clear behavior.", "Low"),
        ("Header overrides auth-managed", "Edge Case",
         "User Authorization beats Auth tab",
         "1. Auth tab Bearer abc.\n2. Headers Authorization='Bearer xyz'.",
         "User-specified header wins; auth tab override flagged in UI.", "Medium"),
        ("Header disabled", "Functional",
         "Toggle disable a header",
         "1. Disable an enabled row.\n2. Send.",
         "Header not transmitted; row greyed.", "Medium"),
        ("Header autocomplete suggests", "UX/UI",
         "Autocomplete for standard names",
         "1. Type 'cont' in key.",
         "Content-Type/Content-Length/Content-Encoding suggested.", "Low"),
        ("Content-Length auto-computed for body", "Functional",
         "Body present sets Content-Length",
         "1. POST raw JSON; do not set Content-Length manually.",
         "Outgoing request has Content-Length = body bytes.", "Medium"),
        ("Content-Type from body type", "Functional",
         "Body type sets Content-Type",
         "1. Body form-data; do not set Content-Type.",
         "multipart/form-data; boundary=... set automatically.", "High"),
        ("User-set Content-Type overrides", "Functional",
         "Manual Content-Type wins",
         "1. Body form-data + manual Content-Type: text/plain.",
         "User header wins (rare and dangerous); UI warns about mismatch.", "Medium"),
        ("Cookie header from jar vs manual", "Edge Case",
         "Manual Cookie beats jar",
         "1. Jar has session=abc.\n2. Manually set Cookie=session=xyz.",
         "Manual wins; jar value not appended.", "Medium"),
        ("Header value with non-ASCII", "Edge Case",
         "Non-ASCII header value",
         "1. Add X-User='测试'.\n2. Send.",
         "Encoded per RFC 5987 or sent literally (server-dependent).", "Low"),
        ("Header name with non-ASCII", "Negative",
         "Reject non-ASCII header name",
         "1. Try '测试: 1'.",
         "Rejected; HTTP requires token-set name.", "Low"),
        ("Header name with space", "Negative",
         "Reject space in header name",
         "1. Try 'X Trace: 1'.",
         "Rejected; not a valid token.", "Low"),
        ("Header value with semicolon", "Edge Case",
         "Multi-part value preserved",
         "1. Cache-Control='no-cache, no-store, max-age=0'.",
         "Sent verbatim.", "Low"),
        ("Header order preserved", "Edge Case",
         "Editor order = wire order",
         "1. Add headers A, B, C.\n2. Send.",
         "Wire order matches editor (per HTTP 7230 5.3).", "Low"),
        ("Variable in header value", "Functional",
         "Var interpolation",
         "1. Header 'X-API: {{key}}'.",
         "Resolved value sent.", "Medium"),
        ("Variable in header key", "Edge Case",
         "Var as header name",
         "1. Header '{{hk}}: v'.",
         "Resolves; rejected if resulting name invalid.", "Low"),
    ]
    for sub, ttype, title, steps, expected, prio in edge_cases:
        t(BOTH, "HD", sub, ttype, title, "-", steps, "header edge case", expected, prio)


gen_headers_deep()


# ---- CHANGES-TO-PUSH VIEW (CP) ----
# 17 entity buckets from threeWayDiff.ts. Operations: Added, Modified, Removed.
ENTITY_BUCKETS = [
    ("Request", "request",
     ["name", "method", "url", "headers", "params", "body", "auth", "preRequestScript",
      "tests", "docs", "settings", "bodySchemaId"]),
    ("Folder", "folder", ["name", "auth", "settings"]),
    ("Folder tree", "tree", ["reorder", "restructure"]),
    ("Collection", "collection", ["name", "settings"]),
    ("Environment", "environment", ["name (key)"]),
    ("Active environment", "environmentsActive", ["change"]),
    ("Environment priority", "environmentsPriority", ["reorder"]),
    ("Environment variable", "environmentVar",
     ["key rename", "value change", "type toggle (plain<->secret)"]),
    ("Execution plan", "executionPlan",
     ["name", "step add", "step remove", "step reorder", "step enable/disable"]),
    ("Mock server", "mockServer", ["name", "endpoints", "responses"]),
    ("Mock endpoint", "mockEndpoint", ["path", "method", "matching rule"]),
    ("Mock response", "mockResponse", ["status", "headers", "body"]),
    ("Linked workspace", "linkedWorkspace", ["link", "unlink", "metadata"]),
    ("Linked request override", "linkedRequestOverride",
     ["add override", "modify override", "delete override"]),
    ("Linked env var override", "linkedEnvOverride",
     ["add override", "modify override", "delete override"]),
    ("Linked release ledger", "releasePerLink", ["refresh ledger"]),
    ("Global JSON Schema", "globalSchema", ["name", "content"]),
    ("Global GraphQL doc", "globalGraphql", ["name", "content"]),
    ("Secret key/slot", "secretKey", ["add", "delete", "rename label", "salt change"]),
    ("Workspace passphrase", "secretCrypto", ["set", "update"]),
    ("Release ledger (self)", "releaseSelf", ["publish", "update version"]),
]

OPS = ["Added", "Modified", "Removed"]


def gen_changes_to_push_matrix():
    for label, bucket, fields in ENTITY_BUCKETS:
        for op in OPS:
            # Modified gets sub-tests for each modifiable field
            if op == "Modified" and fields and bucket != "tree":
                for f in fields:
                    sub = f"{label} - Modified.{f}"
                    title = f"Change-to-Push: Modify {label} field '{f}'"
                    steps = (f"1. Open the {label}.\n"
                             f"2. Change '{f}' to a different value.\n"
                             f"3. Open Workspace panel -> Unpushed changes strip / modal.")
                    expected = (f"Strip count increments (~M modified); modal lists row with badge 'modified', "
                                f"label = {label} display name; expanding row shows before/after diff focused on '{f}'. "
                                f"Push then clears the row from the list.")
                    t(BOTH, "CP", label, "Functional", title,
                      "Linked workspace; clean state.", steps,
                      f"bucket={bucket}; op=modify {f}",
                      expected, "High" if bucket in ("request", "environment", "environmentVar",
                                                       "executionPlan", "mockServer") else "Medium")
            else:
                sub = f"{label} - {op}"
                title = f"Change-to-Push: {op} a {label}"
                if op == "Added":
                    base_steps = f"1. Create a new {label}."
                elif op == "Removed":
                    base_steps = f"1. Delete an existing {label}."
                else:  # Modified for tree / no-fields buckets
                    base_steps = f"1. Reorder / restructure {label}."
                steps = (f"{base_steps}\n"
                         f"2. Open Workspace panel.\n"
                         f"3. Inspect Unpushed changes strip and modal.")
                expected = (f"Strip count reflects the change (+N added / ~M modified / -K removed); "
                            f"modal lists a row with badge '{op.lower()}' and the correct label. "
                            f"Pushing clears it.")
                t(BOTH, "CP", label, "Functional", title,
                  "Linked workspace; clean state.", steps,
                  f"bucket={bucket}; op={op}",
                  expected, "High" if bucket in ("request", "environmentVar", "mockServer",
                                                   "executionPlan") else "Medium")

    # Edge cases
    edges = [
        ("Multiple changes batch view", "UX/UI",
         "10 mixed edits aggregated in modal",
         "1. Make 4 adds + 3 modifies + 3 removes.\n2. Open modal.",
         "Counts add up; rows sorted by bucket order, then label; no duplicate rows.", "High"),
        ("Round-trip: add then delete = no change", "Edge Case",
         "Net-zero ops produce nothing",
         "1. Add a folder.\n2. Delete it.",
         "Strip shows 'No unpushed changes'; modal empty.", "Medium"),
        ("Round-trip: rename then rename back = no change", "Edge Case",
         "Renaming back to original",
         "1. Rename A to B.\n2. Rename back to A.",
         "Strip empty.", "Low"),
        ("Modify then revert during session", "Edge Case",
         "Manual revert to base value",
         "1. Edit a request URL.\n2. Re-enter the original URL.",
         "Row drops from changes list.", "Medium"),
        ("Pull merges remote without losing local strip", "Functional",
         "Pull does not erase local-only list",
         "1. Local edits present.\n2. Pull with non-conflicting remote.",
         "Local strip still shows local edits; merged remote applied; on push, only local diffs go out.", "High"),
        ("After push, strip resets to empty",
         "Functional", "Push clears strip",
         "1. Several local edits.\n2. Push.",
         "Strip shows 'No unpushed changes' after success.", "High"),
        ("Refresh detects merged-PR retire shows nothing",
         "Edge Case", "PR merged then Refresh",
         "1. PR merged on remote.\n2. Refresh.",
         "Local strip cleared (last-pulled snapshot updated to main); user prompted to switch branches.", "High"),
        ("Strip count matches modal row count",
         "Regression", "Counts equal modal rows",
         "1. Variety of edits.\n2. Compare strip to modal.",
         "Counts identical to row counts.", "Medium"),
        ("Sort order by bucket then label",
         "UX/UI", "Modal sorted predictably",
         "1. Diverse edits.\n2. Open modal.",
         "Rows ordered: tree, request, folder, env, ... then alphabetic by label.", "Low"),
        ("Diff view for binary body change",
         "Edge Case", "Binary body diff",
         "1. Edit request body from JSON to binary.\n2. Expand row.",
         "Diff shows body kind change; binary content not rendered in full (size shown).", "Low"),
        ("Secret value change shows masked diff",
         "Security", "Masked diff for secrets",
         "1. Change a secret var value.\n2. Expand row.",
         "Diff shows '<secret changed>' not plaintext.", "High"),
        ("Linked override deletion in modal",
         "Functional", "Override removal",
         "1. Delete a linked var override.\n2. Inspect modal.",
         "Row badge 'removed' with label 'linked env var override (<key>)'.", "Medium"),
    ]
    for sub, ttype, title, steps, expected, prio in edges:
        t(BOTH, "CP", sub, ttype, title, "-", steps, "change-to-push edge", expected, prio)


gen_changes_to_push_matrix()


# ---- GIT CONFLICT MATRIX (GC) ----
CONFLICT_BUCKETS = [
    "Request", "Folder", "Folder tree", "Environment", "Environment variable",
    "Active environment", "Environment priority", "Execution plan",
    "Mock server", "Mock endpoint", "Mock response", "Linked workspace",
    "Linked request override", "Linked env var override",
    "Global JSON Schema", "Global GraphQL doc", "Secret key", "Workspace passphrase",
]


def gen_git_conflict_matrix():
    shapes = [
        ("Add vs Add same id", "Add same entity on two devices",
         "1. Device A creates entity X with id I.\n2. Device B creates same id I (rare; typically only if id collision).\n3. B pulls A's commit.",
         "Detected as conflict; user chooses theirs/mine/manual.",
         "Medium", "Negative"),
        ("Edit vs Edit same field", "Both edit same field differently",
         "1. A edits field F to 'a'.\n2. B edits same F to 'b'.\n3. B pulls.",
         "Conflict modal lists field F with both values; user picks resolution.",
         "High", "Negative"),
        ("Edit vs Edit different fields", "Disjoint edits auto-merge",
         "1. A edits field F1; B edits F2.\n2. B pulls.",
         "Auto-merged; no conflict modal; final state has both edits.",
         "High", "Functional"),
        ("Edit vs Delete", "One edits, other deletes",
         "1. A edits entity X.\n2. B deletes X.\n3. B pulls.",
         "Conflict: keep edit (resurrect) vs accept delete; explicit choice.",
         "High", "Negative"),
        ("Delete vs Edit", "Mirror of above",
         "1. A deletes X.\n2. B edits X.\n3. B pulls.",
         "Same: keep edit or accept delete.", "High", "Negative"),
        ("Delete vs Delete", "Both delete - no conflict",
         "1. Both delete X.\n2. B pulls.",
         "Auto-merged; X gone in both.", "Medium", "Functional"),
        ("Reorder vs Reorder", "Order conflicts",
         "1. A reorders A,B,C to C,B,A.\n2. B reorders to B,A,C.\n3. B pulls.",
         "Conflict on order; user picks one ordering or merges manually.", "Medium", "Negative"),
        ("Rename vs Rename", "Both rename same entity",
         "1. A renames X to Y; B renames X to Z.\n2. B pulls.",
         "Conflict; user picks new name.", "Medium", "Negative"),
        ("Add vs Reparent", "Tree conflict",
         "1. A adds new request under folder F.\n2. B moves F under a different parent.\n3. B pulls.",
         "Tree merge: new request follows F to its new parent.", "Medium", "Edge Case"),
        ("Bulk edits", "10 changes each side",
         "1. A makes 10 edits; B makes 10 disjoint edits.\n2. B pulls.",
         "All 20 merged; no false positives in conflict modal.", "Medium", "Performance"),
    ]
    for bucket in CONFLICT_BUCKETS:
        for shape_name, shape_title, shape_steps, shape_expected, prio, ttype in shapes:
            # Some shapes only meaningful for certain buckets
            if bucket in ("Active environment", "Environment priority", "Workspace passphrase"):
                if shape_name in ("Add vs Add same id", "Add vs Reparent", "Reorder vs Reorder"):
                    continue
            if bucket == "Workspace passphrase" and shape_name in (
                "Delete vs Delete", "Edit vs Delete", "Delete vs Edit", "Rename vs Rename"):
                continue
            sub = f"{bucket} - {shape_name}"
            title = f"Conflict on {bucket}: {shape_title}"
            steps = shape_steps
            expected = shape_expected
            t(BOTH, "GC", sub, ttype, title,
              "Linked workspace; two devices/sessions.", steps,
              f"bucket={bucket}; shape={shape_name}",
              expected, prio)

    # Resolution strategies (apply to any conflict)
    for strat in ["keep mine", "accept theirs", "manual edit"]:
        t(BOTH, "GC", f"Resolution - {strat}", "Functional",
          f"Resolve conflict by '{strat}'",
          "Conflict modal open.",
          f"1. Pick '{strat}' for a conflict.\n2. Apply.\n3. Push.",
          f"resolution={strat}",
          f"Resolution applied: 'keep mine' retains local; 'accept theirs' uses remote; 'manual edit' uses user-typed value. New commit pushed.",
          "High")

    # Re-conflict cycle
    t(BOTH, "GC", "Re-conflict", "Edge Case",
      "Resolution then push creates new conflict on B",
      "Two devices.",
      "1. A pushes resolved state.\n2. B (which still has local edits) pulls.",
      "Conflict re-evaluated against the resolved state; new conflict modal if needed.",
      "Medium")

    t(BOTH, "GC", "Conflict during retire", "Edge Case",
      "Working branch retired mid-conflict",
      "-", "1. PR for working branch merges to main while user has conflict modal open.",
      "-", "Banner informs user; resolution may need to be retargeted; no silent corruption.",
      "Medium")


gen_git_conflict_matrix()


# ---- MOCK RESPONSE MATRIX (MR) ----
def gen_mock_response_matrix():
    statuses = [
        ("200", "OK with JSON body", "Medium"),
        ("201", "Created with Location header", "Medium"),
        ("204", "No Content - empty body", "Medium"),
        ("301", "Permanent redirect to Location", "Low"),
        ("302", "Temporary redirect", "Low"),
        ("304", "Not Modified - no body", "Low"),
        ("400", "Bad Request with error body", "Medium"),
        ("401", "Unauthorized with WWW-Authenticate", "Medium"),
        ("403", "Forbidden", "Low"),
        ("404", "Not Found", "Medium"),
        ("422", "Unprocessable Entity with field errors", "Medium"),
        ("429", "Too Many Requests with Retry-After", "Low"),
        ("500", "Internal Server Error", "Medium"),
        ("502", "Bad Gateway", "Low"),
        ("503", "Service Unavailable", "Low"),
    ]
    for code, descr, prio in statuses:
        sub = f"Status {code}"
        title = f"Mock returns {code}: {descr}"
        steps = (f"1. Define endpoint with status {code} and the body/headers described.\n"
                 f"2. Send a matching request.")
        expected = (f"Client receives status {code}; response panel shows correct status badge; "
                    f"body and headers match the definition.")
        # Desktop runtime
        t(DESK, "MR", sub, "Functional", title,
          "Mock running.", steps, f"status={code}", expected, prio)
        # Web definition (just verify the definition saves; runtime is desktop)
        t(WEB, "MR", sub, "Functional", f"Define mock response {code}",
          "Web build.",
          f"1. Define endpoint with status {code} and body.\n"
          f"2. Save.",
          f"status={code}",
          f"Definition saved correctly; (runtime only in desktop).", prio)

    content_types = [
        ("application/json", "JSON body"),
        ("application/xml", "XML body"),
        ("text/html", "HTML body"),
        ("text/plain", "Plain text"),
        ("application/octet-stream", "Binary body"),
        ("application/x-www-form-urlencoded", "Form-urlencoded"),
        ("multipart/form-data", "Multipart"),
        ("text/event-stream", "Server-Sent Events"),
    ]
    for ct, descr in content_types:
        sub = f"Content-Type {ct}"
        title = f"Mock response with {ct}"
        steps = f"1. Define response with Content-Type {ct}.\n2. Send matching request."
        expected = f"Response Content-Type = {ct}; body type rendered appropriately in viewer."
        t(DESK, "MR", sub, "Functional", title, "Mock running.", steps, ct, expected, "Medium")

    # Matching rules
    matching = [
        ("Method match", "Method-only routing",
         "1. Two endpoints same path: GET and POST.\n2. Send GET.",
         "GET endpoint matched; POST not invoked.", "High"),
        ("Path exact match", "Exact path",
         "1. /users vs /users/list; send /users.",
         "Only /users matched.", "High"),
        ("Path with :id param", "Param matching",
         "1. /users/:id; send /users/42.",
         "Matched; id=42 captured.", "High"),
        ("Path with {id} param", "Brace-style param",
         "1. /users/{id}; send /users/7.",
         "Matched; id=7 captured.", "High"),
        ("Path wildcard /*", "Wildcard",
         "1. /assets/*; send /assets/img/a.png.",
         "Matched; wildcard captured.", "Medium"),
        ("Two endpoints same path/method, different selectors", "Selector",
         "1. Endpoint A matches body.type='alpha'; B matches body.type='beta'.\n2. Send each.",
         "Each request routed to matching response.", "High"),
        ("No selector match - 404", "Fallthrough",
         "1. No matching response.",
         "Server returns 404 with informative body.", "Medium"),
        ("Header-based selector", "Match request header",
         "1. Selector header X-Tenant=A.\n2. Send with X-Tenant=A.",
         "Matched.", "Medium"),
        ("Query-based selector", "Match query param",
         "1. Selector ?env=staging.\n2. Send with env=staging.",
         "Matched.", "Medium"),
        ("Path param echo in response", "Use captured :id in body",
         "1. Response template '{\"id\":\"{id}\"}'.\n2. Send /users/9.",
         "Body returns {\"id\":\"9\"}.", "Medium"),
    ]
    for sub, title, steps, expected, prio in matching:
        t(DESK, "MR", "Matching", "Functional", f"Mock matching: {title}",
          "Mock running.", steps, "matching", expected, prio)
        t(WEB, "MR", "Matching", "Functional", f"Mock matching definition: {title}",
          "Web build.", steps.replace("Send", "Define and verify (no runtime)"),
          "matching", expected.replace("Matched", "Definition valid"), prio)

    # Delays / streaming
    delays = [
        ("Immediate response", "Default behavior",
         "1. Definition has no delay.\n2. Send.",
         "Response within typical local latency (<50ms).", "Medium"),
        ("Configured delay 2s", "Latency simulation",
         "1. Definition delay=2000ms.\n2. Send.",
         "Response after ~2s.", "Low"),
        ("Streamed response chunks", "Streaming",
         "1. Response defined as stream (if supported).\n2. Send.",
         "Body arrives in chunks; UI updates progressively.", "Low"),
        ("Random delay 100-500ms", "Jitter",
         "1. Definition delay range.\n2. Send 5 times.",
         "Each response within range; UI handles variance.", "Low"),
    ]
    for sub, title, steps, expected, prio in delays:
        t(DESK, "MR", "Timing", "Functional", f"Mock timing: {title}",
          "Mock running.", steps, "delay", expected, prio)

    # CLI mock
    t(DESK, "CL", "Mock Server", "Functional",
      "apicircle mock serves defined responses",
      "Workspace with mock; CLI installed.",
      "1. Run 'apicircle mock <ws>'.\n2. Hit each endpoint with curl.",
      "All endpoints",
      "Each endpoint returns the defined response; logs to stdout with timing.",
      "High")
    t(DESK, "CL", "Mock Server", "Functional",
      "apicircle mock multiple endpoints concurrent",
      "Mock has 5 endpoints.",
      "1. Run mock.\n2. curl all 5 in parallel.",
      "5 concurrent",
      "All respond correctly; logs differentiate concurrent requests.",
      "Medium")
    t(DESK, "CL", "Mock Server", "Negative",
      "apicircle mock with invalid workspace exits cleanly",
      "Corrupt workspace.json.",
      "1. Run.", "-",
      "Clear validation error; exit 1.", "Medium")


gen_mock_response_matrix()


# ---- JSON SCHEMA REFERENCES (JS) ----
def gen_json_schema_refs():
    scenarios = [
        ("Local $ref at root", "Functional",
         "$ref to top-level definition",
         "Schema has #/definitions/User and $ref to it from root properties.",
         "1. Set body schema; type body referencing User.",
         "Editor validates against User; mismatches highlighted.", "High"),
        ("Local $ref nested 2 levels", "Functional",
         "Medium-depth $ref resolution",
         "Schema has definitions A -> B -> C with $refs.",
         "1. Set body matching C.",
         "All levels resolved; validation works end-to-end.", "High"),
        ("Local $ref nested 5 levels", "Edge Case",
         "Deep $ref",
         "Schema with 5-level $ref chain.",
         "1. Body matching deepest.",
         "Resolution terminates; validation correct.", "Medium"),
        ("Local $ref with allOf", "Functional",
         "Composition allOf",
         "Schema combines two refs with allOf.",
         "1. Body matching combined schema.",
         "Both constraints applied; violations against either flagged.", "High"),
        ("Local $ref with anyOf", "Functional",
         "Composition anyOf",
         "anyOf two refs.",
         "1. Body matching one alternative.",
         "Accepted if matches at least one.", "Medium"),
        ("Local $ref with oneOf", "Functional",
         "Composition oneOf",
         "oneOf alternatives.",
         "1. Body matching exactly one.",
         "Accepted only if matches one.", "Medium"),
        ("Local $ref with not", "Edge Case",
         "Negative composition",
         "Schema with 'not' clause.",
         "1. Body that violates 'not'.",
         "Validation fails; clear error.", "Low"),
        ("Circular $ref", "Edge Case",
         "Self-referential schema",
         "Schema tree node $refs to itself for child nodes.",
         "1. Body with nested children.",
         "Resolution terminates; validation works for finite trees.", "Medium"),
        ("External $ref to file", "Functional",
         "$ref to global asset",
         "Schema $refs 'schemas/User.json' (another globalAssets entry).",
         "1. Body matching User.",
         "Cross-asset resolution succeeds; validation correct.", "High"),
        ("External $ref broken (asset missing)", "Negative",
         "Missing target",
         "Schema $refs deleted/missing asset.",
         "1. Open editor.",
         "Clear error 'cannot resolve $ref'; UI does not crash.", "Medium"),
        ("$ref with JSON Pointer escapes (~0, ~1)", "Edge Case",
         "Escaped pointer",
         "Schema has key with '/' or '~'.",
         "1. Reference via escaped pointer.",
         "Pointer decoded correctly.", "Low"),
        ("$ref to URL (http)", "Edge Case",
         "Remote $ref",
         "$ref to https://schemas.example/User.json.",
         "1. Open editor.",
         "Either fetched and cached, or warning that remote $ref not supported. Consistent behavior.", "Medium"),
        ("Schema validation - missing required field", "Negative",
         "Required field omitted",
         "Schema requires 'id'.",
         "1. Body omits id.",
         "Squiggle on missing field; sendable but flagged.", "High"),
        ("Schema validation - wrong type", "Negative",
         "Type mismatch",
         "Schema id: integer.",
         "1. Body id='abc'.",
         "Squiggle on field; sendable.", "Medium"),
        ("Schema with pattern regex", "Edge Case",
         "Pattern validation",
         "Schema field pattern '^[A-Z]+$'.",
         "1. Body field='aaa'.",
         "Squiggle.", "Low"),
        ("Schema with enum", "Functional",
         "Enum validation",
         "Schema enum [a,b,c].",
         "1. Body field='d'.", "Squiggle on field.", "Medium"),
        ("Schema with min/max length", "Functional",
         "Length constraints",
         "Schema minLength=3 maxLength=10.",
         "1. Body field='ab'.", "Squiggle on field.", "Low"),
        ("Schema with min/max value (number)", "Functional",
         "Numeric range",
         "Schema minimum=0 maximum=100.",
         "1. Body field=150.", "Squiggle.", "Low"),
        ("Schema applied to response body too (if supported)", "Edge Case",
         "Response validation",
         "responseSchemaId set.",
         "1. Receive non-conforming response.",
         "Tests panel reports schema mismatch.", "Medium"),
        ("Schema autocomplete in body editor", "UX/UI",
         "Property autocomplete",
         "Schema with known properties.",
         "1. Type '{' and Ctrl+Space.",
         "Suggestions include schema-defined properties.", "Medium"),
        ("Schema edit live re-validates body", "Functional",
         "Reactive validation",
         "Body matches schema.",
         "1. Edit schema to add required field.",
         "Body now shows squiggle without manual refresh.", "Medium"),
        ("Schema rename updates references", "Functional",
         "Rename safety",
         "Schema 'User' referenced by request bodies.",
         "1. Rename to 'Customer'.",
         "All bodySchemaId refs update; no orphans.", "Medium"),
        ("Schema delete with refs warns", "Negative",
         "Delete with references",
         "Schema referenced by 3 requests.",
         "1. Delete schema.",
         "Warning lists referencing requests; cancel preserves; confirm clears refs.", "Medium"),
    ]
    for sub, ttype, title, pre, steps, expected, prio in scenarios:
        t(BOTH, "JS", sub, ttype, title, pre, steps, "JSON Schema", expected, prio)


gen_json_schema_refs()


# ---- WORKSPACE RESTORE FROM GIT (WR) ----
def gen_workspace_restore_matrix():
    # Each data type round-trip: configure -> push -> wipe locally -> re-clone -> verify restore
    data_types = [
        ("Collections and folders (tree)", "10 collections, 30 folders, 100 requests", "High"),
        ("Request URL/method/headers/body/auth", "Configured request", "High"),
        ("Request scripts (pre/test) and assertions", "Scripts non-empty", "High"),
        ("Request docs (Markdown)", "Per-request docs", "Medium"),
        ("Request settings (timeout, redirect)", "Custom settings", "Medium"),
        ("Body schema reference", "bodySchemaId set", "Medium"),
        ("Environments and their variables", "3 envs, 20 vars each", "High"),
        ("Environment priority order", "Custom order", "Medium"),
        ("Active environment", "Set Active=Prod", "Medium"),
        ("Secret variables (encrypted at rest)", "Several secrets", "High"),
        ("Execution plans", "Plan with 10 steps", "High"),
        ("Plan step enable/disable flags", "Mix enabled/disabled", "Medium"),
        ("Mock servers", "2 mocks with 10 endpoints each", "High"),
        ("Mock endpoint matching rules", "Complex selectors", "Medium"),
        ("Mock response bodies and headers", "Diverse responses", "Medium"),
        ("Linked workspaces and overrides", "Linked WS + per-var override", "High"),
        ("Linked request overrides", "Override URL/headers", "Medium"),
        ("Global JSON Schemas", "Multiple schemas with $refs", "Medium"),
        ("Global GraphQL docs", "Persisted docs", "Medium"),
        ("Release ledger entries", "Published releases", "Medium"),
        ("Linked release ledger", "Cached versions", "Low"),
        ("Workspace passphrase metadata", "KDF + verifier", "High"),
        ("Secret keys / slots metadata", "Labels & salts", "High"),
    ]
    for name, descr, prio in data_types:
        sub = f"Round-trip: {name}"
        title = f"Restore {name} from Git clone"
        pre = f"Workspace linked; has {descr}; pushed."
        steps = (f"1. Delete workspace locally (or use a fresh device).\n"
                 f"2. Link to the same repo and pull.\n"
                 f"3. Verify {name} is restored identically.")
        expected = (f"All {name} restored intact; ids stable; references resolve; "
                    f"no data loss. Secrets remain encrypted until passphrase entered.")
        t(BOTH, "WR", sub, "Functional", title, pre, steps, name, expected, prio)

    # Edge cases
    edges = [
        ("Empty repo init", "Functional",
         "Empty workspace clone",
         "Repo has empty workspace.json.",
         "1. Link and pull.",
         "App initializes a clean empty workspace; no errors.", "Medium"),
        ("Schema version mismatch (older repo)", "Negative",
         "Older schema version",
         "Repo has older WorkspaceSynced schema.",
         "1. Pull.",
         "App migrates forward or refuses with clear message; never silently corrupts.", "High"),
        ("Schema version mismatch (newer repo)", "Negative",
         "Newer schema version",
         "Repo has newer schema.",
         "1. Pull.",
         "App informs user to upgrade; no silent partial load.", "High"),
        ("Repo with secrets references but no passphrase metadata", "Edge Case",
         "Mismatched secret state",
         "-", "1. Pull.",
         "Secret rows shown as locked; user prompted to set passphrase or skip.", "Medium"),
        ("Large workspace (10MB JSON)", "Performance",
         "Big workspace load time",
         "10MB workspace.json.", "1. Pull.",
         "Hydrate completes <5s; UI does not freeze.", "Medium"),
        ("Workspace with broken refs (request -> deleted folder)", "Negative",
         "Dangling reference",
         "Manually corrupt: request.folderId points to missing folder.",
         "1. Pull.",
         "App surfaces health warning and offers to fix (re-home requests).", "Medium"),
    ]
    for sub, ttype, title, pre, steps, expected, prio in edges:
        t(BOTH, "WR", sub, ttype, title, pre, steps, "restore edge", expected, prio)


gen_workspace_restore_matrix()


# ---- HISTORY REPLAY MATRIX (HR) ----
def gen_history_replay_matrix():
    scenarios = [
        ("Simple GET", "GET 200 in history", "1. Replay.", "Request restored; sends correctly.", "High", "Functional"),
        ("POST with JSON body", "JSON body run", "1. Replay.",
         "Body restored; Content-Type intact.", "High", "Functional"),
        ("POST with form-data + file attachment", "Form-data run",
         "1. Replay.",
         "If attachment still in IndexedDB, restored; if pruned, user warned.", "Medium", "Edge Case"),
        ("OAuth2 Bearer (token may be expired)", "Run with token",
         "1. Replay.",
         "Token re-fetched / refreshed if expired; new run uses fresh token.", "High", "Functional"),
        ("AWS SigV4 (timestamp sensitive)", "Signed run",
         "1. Replay.",
         "Re-signs with current timestamp; original signature replaced.", "High", "Functional"),
        ("Digest auth", "Digest run",
         "1. Replay.",
         "Challenge cycle repeats with current nonce.", "Medium", "Functional"),
        ("Variable interpolation since env changed", "Env value changed since run",
         "1. Replay after changing var.",
         "Re-resolves with current env; original wire bytes not reused.", "Medium", "Edge Case"),
        ("Request was deleted", "Source request deleted",
         "1. Replay from history.",
         "Replays into a transient editor or new request stub; original deletion preserved.", "Medium", "Edge Case"),
        ("Replay across workspace switch", "Different workspace active",
         "1. Replay run from workspace A while B is active.",
         "Either switches to A or refuses with clear message; no cross-workspace state leak.", "Medium", "Edge Case"),
        ("Replay run from before schema migration", "Old schema run",
         "1. Replay.",
         "Migrated forward at replay or user informed; replay does not corrupt.", "Low", "Negative"),
        ("Replay creates new history entry", "New run",
         "1. Replay.",
         "A new RequestRun row is added; original untouched.", "High", "Functional"),
        ("Replay 100 runs in batch (if supported)", "Bulk replay",
         "1. Select 100 runs; replay.",
         "Sequential replay completes; aggregate result visible.", "Low", "Performance"),
        ("Replay with assertions runs them too", "Run with assertions",
         "1. Replay.",
         "Assertions re-evaluated against new response.", "High", "Functional"),
        ("Replay GraphQL with stored schema",
         "GraphQL run", "1. Replay.",
         "Schema still drives completions if needed.", "Low", "Functional"),
        ("Replay GET response includes Set-Cookie",
         "Cookie-producing run", "1. Replay.",
         "Cookies updated in jar (jar represents current world, not snapshot).", "Medium", "Functional"),
        ("Replay shows raw old response and new response side by side",
         "Comparison", "1. Replay; compare snapshots.",
         "Timeline shows two snapshots; user can diff status/body/headers.", "Medium", "UX/UI"),
        ("Replay aborted mid-flight", "Cancel during replay",
         "1. Replay slow request; Cancel.",
         "New run marked Cancelled; old run untouched.", "Medium", "Negative"),
        ("Replay against unreachable host",
         "Endpoint down", "1. Replay.",
         "Network error in new run; original unaffected.", "Medium", "Negative"),
        ("Replay updates request UI to historical state",
         "Pre-replay editor state different", "1. Replay.",
         "Editor optionally restores historical values (with confirm) or keeps current; documented behavior.", "Medium", "UX/UI"),
    ]
    for sub, pre, steps, expected, prio, ttype in scenarios:
        title = f"Replay: {sub}"
        t(BOTH, "HR", sub, ttype, title, pre, steps, "replay", expected, prio)


gen_history_replay_matrix()


# ---- METHOD x AUTH x BODY TRIPLE MATRIX (MM extended) ----
TRIPLE_AUTHS = ["None", "Bearer", "Basic", "ApiKeyHeader", "ApiKeyQuery",
                "AwsSigV4", "JwtHS256", "OAuth2ClientCreds", "OAuth2PKCE",
                "Digest", "NTLM"]
TRIPLE_BODIES = ["none", "form-data", "x-www-form-urlencoded",
                 "raw-json", "raw-xml", "raw-text", "binary", "graphql"]


def gen_method_auth_body_triple():
    """For meaningful (method, auth, body) combinations, generate explicit tests.

    This is a deliberately wide matrix - many cells are unusual on purpose,
    because real APIs hit every weird corner. Each test should give the QA
    engineer specific things to verify.
    """
    for m in HTTP_METHODS:
        for auth in TRIPLE_AUTHS:
            for body in TRIPLE_BODIES:
                # Trim some clearly nonsensical combos to keep count sane.
                # Skip body=graphql with non-POST (covered in MM module).
                if body == "graphql" and m != "POST":
                    continue
                # Skip auth=None + bodyless methods + body=none (covered in MM).
                if auth == "None" and m in BODYLESS_METHODS and body == "none":
                    continue
                sub = f"{m} + {auth} + {body}"
                title = f"{m} with {auth} auth and {body} body"
                steps = (
                    f"1. New request; method {m}; URL to an echo endpoint.\n"
                    f"2. Auth type {auth}; configure minimum credentials.\n"
                    f"3. Body type {body}; provide minimum valid content.\n"
                    f"4. Send and inspect outgoing wire request (DevTools/proxy)."
                )
                # Heuristic expectations
                if auth == "AwsSigV4" and body != "none":
                    expected = (
                        f"AWS SigV4 signs canonical request including body bytes; "
                        f"x-amz-content-sha256 reflects SHA256 of body. Server returns expected status."
                    )
                    prio = "High"
                    ttype = "Functional"
                elif auth == "Digest" and m in ("POST", "PUT") and body != "none":
                    expected = (
                        f"Digest challenge handled even though body is buffered; "
                        f"retry sends same body. Final status OK if creds valid."
                    )
                    prio = "Medium"
                    ttype = "Functional"
                elif auth == "NTLM" and m in ("POST", "PUT") and body != "none":
                    expected = (
                        f"NTLM 3-way handshake; body re-transmitted on Authenticate step. "
                        f"App handles body buffering across handshake without truncation."
                    )
                    prio = "Medium"
                    ttype = "Edge Case"
                elif auth.startswith("OAuth2"):
                    expected = (
                        f"Access token acquired (or refreshed); attached as Bearer; "
                        f"{m} sent with {body} body; server returns expected status."
                    )
                    prio = "Medium"
                    ttype = "Functional"
                elif m in BODYLESS_METHODS and body != "none":
                    expected = (
                        f"App allows configuring {body} body for {m}; verifies wire "
                        f"behavior (sent or stripped) is consistent. Auth header sent regardless."
                    )
                    prio = "Low"
                    ttype = "Edge Case"
                elif auth == "None":
                    expected = (
                        f"No Authorization header on {m}; {body} body sent if applicable."
                    )
                    prio = "Low"
                    ttype = "Functional"
                else:
                    expected = (
                        f"{auth} auth applied correctly; {body} body sent with right "
                        f"Content-Type; server returns expected status."
                    )
                    prio = "Medium"
                    ttype = "Functional"
                t(BOTH, "MM", sub, ttype, title, "Editor open.", steps,
                  f"method={m} auth={auth} body={body}", expected, prio)


gen_method_auth_body_triple()


# ---- STATUS CODE x METHOD RESPONSE MATRIX ----
RESPONSE_STATUS_GROUPS = [
    ("200", "OK"), ("201", "Created"), ("202", "Accepted"), ("204", "No Content"),
    ("301", "Moved Permanently"), ("302", "Found"), ("304", "Not Modified"),
    ("307", "Temporary Redirect"), ("308", "Permanent Redirect"),
    ("400", "Bad Request"), ("401", "Unauthorized"), ("403", "Forbidden"),
    ("404", "Not Found"), ("405", "Method Not Allowed"), ("409", "Conflict"),
    ("410", "Gone"), ("418", "I'm a teapot"), ("422", "Unprocessable Entity"),
    ("429", "Too Many Requests"),
    ("500", "Internal Server Error"), ("501", "Not Implemented"),
    ("502", "Bad Gateway"), ("503", "Service Unavailable"), ("504", "Gateway Timeout"),
]


def gen_status_method_matrix():
    for code, descr in RESPONSE_STATUS_GROUPS:
        for m in ["GET", "POST", "PUT", "DELETE"]:
            sub = f"{code} on {m}"
            title = f"Response {code} ({descr}) for {m}"
            steps = (
                f"1. {m} to endpoint that returns {code}.\n"
                f"2. Inspect response panel."
            )
            expected = (
                f"Response panel shows status {code} with appropriate badge color "
                f"(2xx green, 3xx blue, 4xx orange, 5xx red); body and headers rendered. "
                f"History entry records correctly."
            )
            prio = "Medium" if int(code) in (200, 201, 204, 400, 401, 404, 500) else "Low"
            t(BOTH, "RP", "Status Matrix", "Functional",
              f"{title}", "Editor open; endpoint configured.",
              steps, f"status={code} method={m}", expected, prio)


gen_status_method_matrix()


# ---- COOKIE x METHOD x DOMAIN MATRIX ----
def gen_cookie_matrix():
    methods = ["GET", "POST", "PUT", "DELETE"]
    scenarios = [
        ("Cookie sent on same-domain request", "Cookie a=1 for example.com",
         "Send to example.com",
         "Cookie header includes a=1.", "High"),
        ("Cookie NOT sent to different domain", "Cookie a=1 for example.com",
         "Send to other.com",
         "Cookie NOT included.", "Medium"),
        ("Cookie sent on subdomain (Domain=.example.com)", "Cookie with Domain=.example.com",
         "Send to api.example.com",
         "Cookie included.", "Medium"),
        ("Cookie NOT sent to parent of subdomain", "Cookie for api.example.com",
         "Send to example.com",
         "Cookie NOT included.", "Low"),
        ("Path match: cookie path=/api sent to /api/v1", "Cookie path=/api",
         "Send to /api/v1",
         "Cookie included.", "Medium"),
        ("Path match: cookie path=/api NOT sent to /other", "Cookie path=/api",
         "Send to /other",
         "Cookie NOT included.", "Medium"),
        ("Secure cookie NOT sent over http", "Secure cookie",
         "Send http://example.com",
         "Cookie NOT included.", "High"),
        ("HttpOnly cookie sent like any other", "HttpOnly cookie",
         "Send.",
         "Included; only restriction is on JS access (response panel may flag).", "Low"),
        ("Multiple cookies for same domain", "3 cookies on example.com",
         "Send.",
         "All 3 in Cookie header, semicolon-separated.", "Medium"),
        ("Expired cookie not sent", "Cookie with past Expires",
         "Send.",
         "Cookie excluded; UI may show as expired.", "Medium"),
        ("Set-Cookie with SameSite=Strict", "Response sets SameSite=Strict",
         "Send.",
         "Cookie stored; flagged with SameSite attribute.", "Low"),
        ("Manual Cookie header overrides jar", "Jar has session=abc; Manual Cookie=session=xyz",
         "Send.",
         "Manual value sent; jar not appended.", "Medium"),
    ]
    for sub, pre, steps, expected, prio in scenarios:
        for m in methods:
            sub2 = f"{sub} ({m})"
            steps_m = f"1. {steps.replace('Send', m + ' send')}"
            t(BOTH, "CO", sub2, "Functional",
              f"{m}: {sub}", pre, steps_m, f"method={m}",
              expected, prio)


gen_cookie_matrix()


# ---- BODY x VARIABLE INTERPOLATION DEEP MATRIX ----
def gen_body_variable_matrix():
    """Variable interpolation across every body type."""
    scopes = ["Workspace var", "Env var", "Secret var"]
    body_targets = [
        ("raw-json", "JSON value", '{"name":"{{v}}"}'),
        ("raw-json", "JSON key", '{"{{v}}":"x"}'),
        ("raw-json", "nested object value", '{"u":{"name":"{{v}}"}}'),
        ("raw-json", "array item", '["a","{{v}}","c"]'),
        ("raw-xml", "element value", '<u name="x">{{v}}</u>'),
        ("raw-xml", "attribute value", '<u name="{{v}}"/>'),
        ("raw-text", "free text", 'hello {{v}}'),
        ("raw-html", "tag content", '<p>{{v}}</p>'),
        ("raw-html", "attribute", '<a href="{{v}}">x</a>'),
        ("x-www-form-urlencoded", "value", 'q={{v}}'),
        ("form-data", "text field value", 'name={{v}}'),
        ("form-data", "file field key", '{{v}}=<file>'),
        ("graphql", "query argument", 'query { user(id:"{{v}}"){ name } }'),
        ("graphql", "variables JSON", 'variables {"id":"{{v}}"}'),
    ]
    for scope in scopes:
        for body, where, snippet in body_targets:
            sub = f"{body} - {where} (scope: {scope})"
            title = f"Interpolate {scope} into {body} {where}"
            pre = f"Workspace has v in scope '{scope}'."
            steps = (
                f"1. Body type {body}.\n"
                f"2. Use snippet referencing {{{{v}}}}: `{snippet}`.\n"
                f"3. Send."
            )
            expected = (
                f"Variable resolves; outgoing body contains the resolved value at {where}; "
                f"Content-Type appropriate for {body}; server echo confirms."
            )
            ttype = "Security" if scope == "Secret var" else "Functional"
            prio = "High" if scope == "Secret var" else "Medium"
            t(BOTH, "BC", "Variable", ttype, title, pre, steps,
              f"body={body} scope={scope}", expected, prio)


gen_body_variable_matrix()


# ---- PLAN STEP x ASSERTION TYPE MATRIX ----
def gen_plan_step_matrix():
    step_kinds = [
        ("Single GET", "Single step with GET",
         "1. Plan with 1 step (GET).\n2. Run.",
         "Step passes if endpoint reachable.", "High"),
        ("Sequential 5 steps", "5 sequential requests",
         "1. Plan with 5 steps.\n2. Run.",
         "Each runs in order; aggregate result shown.", "High"),
        ("Step 2 depends on step 1 output", "Extract id; use {{id}} in step 2",
         "1. Step 1 sets pm.variables('id'); step 2 uses {{id}}.\n2. Run.",
         "Step 2 receives extracted id; succeeds.", "High"),
        ("Step with pre-script error", "Pre-script throws",
         "1. Step 2 has pre-script error.\n2. Run.",
         "Step 2 marked failed; plan continues or stops per setting.", "Medium"),
        ("Step with post-script error", "Tests script throws",
         "1. Step 2 tests throw.\n2. Run.",
         "Step 2 marked failed; HTTP response still shown.", "Medium"),
        ("Disabled step skipped", "1 of 5 disabled",
         "1. Disable step 3; Run.",
         "Step 3 'skipped'; 1,2,4,5 run.", "High"),
        ("Re-run idempotent for read-only", "All GETs",
         "1. Run twice.",
         "Same aggregate result; no side-effects.", "Medium"),
        ("Loop step (if supported)", "Step inside loop",
         "1. Define loop; Run.",
         "Each iteration recorded; aggregated.", "Low"),
        ("Conditional step (if supported)", "Step gated by assertion",
         "1. Step 2 runs only if step 1.status==200.",
         "Conditional honored.", "Low"),
        ("Step with empty assertions", "No tests defined",
         "1. Run.",
         "Step shows '0 assertions'; status shown.", "Low"),
        ("Step with 50 assertions", "Many assertions",
         "1. Run.",
         "All evaluated; aggregate counts.", "Medium"),
        ("Plan with parallel branch (if supported)", "Parallel",
         "1. Run.",
         "Parallel results joined.", "Low"),
        ("Step timeout", "Step timeout 5s; slow endpoint",
         "1. Run.",
         "Step times out; marked failed.", "Medium"),
        ("Step retry on failure", "Retry policy",
         "1. Step retries 3x.",
         "Retries until success or count exhausted.", "Medium"),
    ]
    assertion_combos = [
        ("status==X", "Status check"),
        ("status<Y", "Status range"),
        ("body.path == value", "JSON path"),
        ("body matches regex", "Regex"),
        ("header present", "Header"),
        ("header value matches", "Header value"),
        ("responseTime lt N", "Duration"),
        ("body schema valid", "Schema"),
    ]
    for sub, title_snippet, steps, expected, prio in step_kinds:
        for op_name, op_descr in assertion_combos:
            full_sub = f"{sub} / {op_descr}"
            full_title = f"Plan step: {title_snippet} with {op_descr} assertion"
            full_steps = f"{steps}\nAssertion: {op_name}"
            full_expected = (
                f"{expected} Assertion outcome (pass/fail) recorded in step report; "
                f"plan aggregate updates."
            )
            ttype = "Functional" if "error" not in sub.lower() and "timeout" not in sub.lower() else "Negative"
            t(BOTH, "AS", full_sub, ttype, full_title, "Plan configured.",
              full_steps, "plan", full_expected,
              prio if op_name in ("status==X", "body.path == value", "responseTime lt N") else "Low")


gen_plan_step_matrix()


# ---- IMPORT/EXPORT ROUND-TRIP MATRIX ----
def gen_import_export_roundtrip():
    formats = [
        ("Postman v2.1", "Postman Collection v2.1"),
        ("Insomnia", "Insomnia export"),
        ("OpenAPI 3.0 YAML", "OpenAPI 3.0 YAML"),
        ("OpenAPI 3.0 JSON", "OpenAPI 3.0 JSON"),
        ("Swagger 2.0", "Swagger 2.0"),
        ("cURL command", "cURL paste"),
        ("HAR file", "HAR (HTTP Archive)"),
    ]
    request_shapes = [
        ("GET no body", "Simple GET"),
        ("POST with JSON body", "POST JSON"),
        ("POST with form-data", "POST form-data"),
        ("POST with file upload", "POST file"),
        ("GraphQL query", "GraphQL POST"),
        ("Bearer auth", "Bearer"),
        ("Basic auth", "Basic"),
        ("OAuth2 (with token)", "OAuth2"),
        ("API key in header", "API key header"),
        ("API key in query", "API key query"),
        ("Custom headers", "Custom headers"),
        ("Path params (:id)", "Path params"),
        ("Query params", "Query params"),
        ("Variables in URL/body", "Variables"),
        ("Pre-request script", "Pre-script"),
        ("Tests / assertions", "Tests"),
    ]
    for fmt_label, fmt in formats:
        for shape_label, shape in request_shapes:
            sub = f"{fmt_label} <-> {shape_label}"
            title = f"Round-trip {fmt_label} for: {shape_label}"
            steps = (
                f"1. In source tool (or fixture), create request shape '{shape_label}'.\n"
                f"2. Export from source as {fmt_label}.\n"
                f"3. Import into API Circle.\n"
                f"4. Compare resulting request to source.\n"
                f"5. (Optional) Export from API Circle as same format; diff."
            )
            expected = (
                f"All compatible fields of '{shape_label}' preserved through import. "
                f"Any fields the format cannot represent are clearly flagged in the import warnings, "
                f"not silently dropped. Round-trip export back yields a structurally equivalent file."
            )
            ttype = "Functional"
            prio = "High" if shape_label in ("Simple GET", "POST JSON", "Bearer", "OAuth2") else "Medium"
            t(BOTH, "IE", "Round-trip", ttype, title, "-", steps,
              f"fmt={fmt} shape={shape}", expected, prio)


gen_import_export_roundtrip()


# ---- MOCK ROUTING DEEPER MATRIX ----
def gen_mock_routing_extras():
    methods = ["GET", "POST", "PUT", "PATCH", "DELETE"]
    scenarios = [
        ("Different methods same path", "/users handled by GET vs POST",
         "Endpoints for both.", "High"),
        ("Same method different paths", "/users vs /users/list",
         "Distinct.", "High"),
        ("Path with multiple params /:org/:repo", "Both captured",
         "Both captured.", "High"),
        ("Path with mixed :param and {param}", "Either supported",
         "Both syntaxes accepted.", "Medium"),
        ("Trailing slash handling", "/users vs /users/",
         "Documented behavior: either equal or distinct, consistent.", "Medium"),
        ("Case sensitivity in path", "/Users vs /users",
         "Distinct or equal per definition; consistent.", "Low"),
        ("Query-string ignored in match", "Path matches regardless of ?x=y",
         "Match ignores query; query can still be inspected.", "Medium"),
        ("Header-based selector OR", "Two endpoints; one matches X=A OR X=B",
         "Both invoke same response.", "Medium"),
        ("Body content selector", "Match request body field",
         "Routes by body value.", "Medium"),
        ("Selector priority order", "Multiple selectors; specific wins",
         "More specific selector picked.", "Medium"),
        ("Wildcard fallback selector", "Default response",
         "Used when no other match.", "Medium"),
    ]
    for sub, title, expected, prio in scenarios:
        for m in methods:
            full_sub = f"{sub} ({m})"
            steps = (
                f"1. Configure scenario: {title}.\n"
                f"2. Send {m} request matching one selector.\n"
                f"3. Send {m} request matching another."
            )
            full_expected = expected + f" (Verified for {m}.)"
            t(DESK, "MR", full_sub, "Functional",
              f"{title} - {m}", "Mock running.", steps,
              f"matching method={m}", full_expected, prio)
            t(WEB, "MR", full_sub, "Functional",
              f"{title} definition - {m}", "Web build.", steps.replace("Send", "Define"),
              f"matching method={m}", full_expected, prio)


gen_mock_routing_extras()


# ---- CLI MOCK SCENARIO MATRIX ----
def gen_cli_mock_scenarios():
    scenarios = [
        ("Boot with single endpoint", "Workspace with 1 mock endpoint",
         "Run 'apicircle mock <ws>'", "Endpoint reachable; logs to stdout.", "High"),
        ("Boot with multiple mocks", "5 mocks defined",
         "Run.", "All start on distinct ports; ports printed.", "Medium"),
        ("Endpoint with path :id", "1 endpoint /users/:id",
         "Curl /users/42.", "200 with body containing 42.", "High"),
        ("Endpoint with multiple responses", "Selector-based",
         "Curl with selector header A vs B.", "Each request routed correctly.", "High"),
        ("Endpoint with delay", "delay=2000ms",
         "Curl; measure time.", "~2s elapsed.", "Low"),
        ("404 fallthrough", "Path not defined",
         "Curl unknown path.", "404 with informative body.", "Medium"),
        ("Method not allowed", "Path exists for GET, send POST",
         "Curl POST.", "405 with Allow header.", "Medium"),
        ("Large response body", "10MB body defined",
         "Curl.", "Bytes received; CLI stays responsive.", "Low"),
        ("Streaming response (SSE)", "text/event-stream",
         "Curl with -N.", "Chunks received progressively.", "Low"),
        ("Binary response", "PDF body",
         "Curl > out.pdf.", "File saved; bytes match.", "Low"),
        ("Multiple concurrent requests", "5 parallel curls",
         "All clients curl simultaneously.", "All respond correctly; logs interleave.", "Medium"),
        ("Re-load definitions without restart (if supported)", "Hot reload",
         "Edit definition; observe.", "Either auto-reload or documented manual restart.", "Low"),
        ("Graceful shutdown on Ctrl+C", "Mock running",
         "Press Ctrl+C.", "All listeners stopped; ports freed; exit 0.", "Medium"),
        ("Workspace path passed as relative", "./ws",
         "Run with relative path.", "Resolves correctly.", "Low"),
        ("Workspace path passed as absolute", "/full/path",
         "Run.", "Works.", "Low"),
        ("Vault secret key env var resolves secrets", "Encrypted vars referenced",
         "APICIRCLE_VAULT_SECRET_KEY=... apicircle mock <ws>.",
         "Secrets decrypted; mock returns interpolated values.", "High"),
        ("Vault secret missing surfaces error", "No env var",
         "Run without env var on workspace with secrets.",
         "Clear error 'VAULT_SECRET_KEY required'; exit non-zero.", "High"),
        ("Custom port via flag", "--port 9000",
         "Run; curl 9000.", "Bound to 9000.", "Medium"),
        ("Port already in use", "Re-run on same port",
         "Run twice.", "Second fails clearly; exit non-zero.", "Medium"),
        ("Log format flag (json/text)", "--log-format json",
         "Run; inspect logs.", "Logs in chosen format.", "Low"),
        ("Verbose flag", "--verbose",
         "Run; inspect.", "More log detail.", "Low"),
        ("Help flag prints usage", "--help",
         "Run.", "Usage and flags shown.", "Low"),
    ]
    for sub, pre, steps, expected, prio in scenarios:
        t(DESK, "CL", f"Mock - {sub}", "Functional",
          f"CLI mock: {sub}", pre, steps, "cli mock", expected, prio)


gen_cli_mock_scenarios()


# ---- ENV / WORKSPACE / EXECUTION EDIT -> CHANGES-TO-PUSH SCENARIOS ----
def gen_edit_to_push_scenarios():
    """Specific user-facing flow tests: edit something, view Changes-to-Push.
    These complement CP module by being narrative."""
    scenarios = [
        # Editor changes
        ("Editor: rename request", "Modified request entry",
         "1. Rename request 'Login' to 'SignIn'.\n2. Open Workspace panel.",
         "Strip ~1 modified; modal lists 'SignIn' (label uses current name) with kind 'modified'.",
         "High"),
        ("Editor: change URL", "Modified request URL",
         "1. Change URL.\n2. Inspect strip.",
         "Strip +0/~1/-0; modal diff shows old vs new URL.", "High"),
        ("Editor: change method", "Modified method",
         "1. GET -> POST.\n2. Strip.",
         "Strip ~1 modified; diff shows method change.", "Medium"),
        ("Editor: add header", "Modified headers",
         "1. Add header.\n2. Strip.",
         "Strip ~1 modified; diff shows new header row.", "Medium"),
        ("Editor: change body", "Modified body",
         "1. Edit body.\n2. Strip.",
         "Strip ~1; diff shows body change.", "Medium"),
        ("Editor: change auth type", "Modified auth",
         "1. Change auth.\n2. Strip.",
         "Strip ~1; diff shows auth change.", "Medium"),
        ("Editor: add pre-request script", "Modified script",
         "1. Add script.\n2. Strip.",
         "Strip ~1; diff shows script.", "Medium"),
        ("Editor: change settings (timeout)", "Modified settings",
         "1. Change timeout.\n2. Strip.",
         "Strip ~1; diff shows setting.", "Low"),
        # Environment changes
        ("Env: create new env", "Added env",
         "1. New env Staging.\n2. Strip.",
         "Strip +1; modal Added env 'Staging'.", "High"),
        ("Env: rename env", "Modified env rename",
         "1. Dev -> Development.\n2. Strip.",
         "Strip shows env rename (added new + removed old, or modified — per impl).", "Medium"),
        ("Env: delete env", "Removed env",
         "1. Delete env.\n2. Strip.",
         "Strip -1.", "Medium"),
        ("Env: add var", "Modified env (var added)",
         "1. Add var.\n2. Strip.",
         "Strip ~1 (env-level); diff shows var added.", "High"),
        ("Env: change var value", "Modified env value",
         "1. Change var value.\n2. Strip.",
         "Strip ~1; diff shows value change.", "High"),
        ("Env: change secret var value", "Secret modified - masked",
         "1. Change secret value.\n2. Strip.",
         "Strip ~1; diff shows '<secret changed>' (never plaintext).", "High"),
        ("Env: toggle secret flag", "Toggled secret",
         "1. Flip flag.\n2. Strip.",
         "Strip ~1; diff records flag change.", "Medium"),
        ("Env: change active env", "Modified active",
         "1. Switch active.\n2. Strip.",
         "Strip ~1 'Active environment'.", "Medium"),
        ("Env: reorder priority", "Modified priority",
         "1. Drag reorder.\n2. Strip.",
         "Strip ~1 'Environment priority'.", "Medium"),
        # Execution plan changes
        ("Plan: create plan", "Added plan",
         "1. New plan.\n2. Strip.",
         "Strip +1 plan.", "High"),
        ("Plan: add step", "Modified plan steps",
         "1. Add step.\n2. Strip.",
         "Strip ~1; diff shows steps array change.", "High"),
        ("Plan: reorder steps", "Modified plan order",
         "1. Drag step.\n2. Strip.",
         "Strip ~1.", "Medium"),
        ("Plan: disable step", "Modified plan enable flag",
         "1. Disable step.\n2. Strip.",
         "Strip ~1.", "Medium"),
        ("Plan: delete plan", "Removed plan",
         "1. Delete.\n2. Strip.",
         "Strip -1.", "Medium"),
        # Mock changes
        ("Mock: create server", "Added mock",
         "1. Create.\n2. Strip.",
         "Strip +1 mock.", "High"),
        ("Mock: rename server", "Modified mock",
         "1. Rename.\n2. Strip.",
         "Strip ~1.", "Medium"),
        ("Mock: add endpoint", "Modified mock endpoints",
         "1. Add endpoint.\n2. Strip.",
         "Strip ~1.", "High"),
        ("Mock: change endpoint path", "Modified endpoint",
         "1. Change path.\n2. Strip.",
         "Strip ~1; diff shows path.", "Medium"),
        ("Mock: add response", "Modified response",
         "1. Add response.\n2. Strip.",
         "Strip ~1.", "Medium"),
        ("Mock: change response body", "Modified body",
         "1. Edit body.\n2. Strip.",
         "Strip ~1; diff shows body change.", "Medium"),
        ("Mock: delete server", "Removed mock",
         "1. Delete.\n2. Strip.",
         "Strip -1.", "Medium"),
        # Linked workspace changes
        ("Linked WS: link a workspace", "Added linked WS",
         "1. Link.\n2. Strip.",
         "Strip +1 linked WS.", "High"),
        ("Linked WS: unlink", "Removed linked WS",
         "1. Unlink.\n2. Strip.",
         "Strip -1.", "Medium"),
        ("Linked WS: add override on request", "Added override",
         "1. Override URL.\n2. Strip.",
         "Strip +1 'linked request override (<key>)'.", "Medium"),
        ("Linked WS: add override on env var", "Added env override",
         "1. Override var value.\n2. Strip.",
         "Strip +1 'linked env var override (<key>)'.", "Medium"),
        ("Linked WS: delete override", "Removed override",
         "1. Delete.\n2. Strip.",
         "Strip -1.", "Low"),
        # Global assets
        ("Schema: add", "Added schema",
         "1. New schema.\n2. Strip.",
         "Strip +1 global schema.", "Medium"),
        ("Schema: edit content", "Modified schema",
         "1. Edit.\n2. Strip.",
         "Strip ~1.", "Medium"),
        ("Schema: rename", "Modified schema name",
         "1. Rename.\n2. Strip.",
         "Strip ~1.", "Low"),
        ("Schema: delete", "Removed schema",
         "1. Delete.\n2. Strip.",
         "Strip -1.", "Medium"),
        ("GraphQL doc: add", "Added GraphQL doc",
         "1. Add.\n2. Strip.",
         "Strip +1 global graphql.", "Low"),
        ("GraphQL doc: edit", "Modified GraphQL doc",
         "1. Edit.\n2. Strip.",
         "Strip ~1.", "Low"),
        ("GraphQL doc: delete", "Removed GraphQL doc",
         "1. Delete.\n2. Strip.",
         "Strip -1.", "Low"),
        # Secrets
        ("Secret keys: add slot", "Added secret slot",
         "1. Add.\n2. Strip.",
         "Strip +1.", "Medium"),
        ("Secret keys: rename label", "Modified label",
         "1. Rename.\n2. Strip.",
         "Strip ~1.", "Low"),
        ("Secret keys: delete slot", "Removed slot",
         "1. Delete.\n2. Strip.",
         "Strip -1.", "Medium"),
        ("Workspace passphrase: set new", "Modified passphrase",
         "1. Set.\n2. Strip.",
         "Strip ~1 'Workspace passphrase'.", "High"),
        # Releases
        ("Release: publish version", "Modified release self",
         "1. Publish.\n2. Strip.",
         "Strip ~1 'Release ledger'.", "Medium"),
        # Combined
        ("Combined: 5 different bucket changes", "Modified mixed",
         "1. Edit request + env + mock + schema + plan in one session.\n2. Strip.",
         "Strip count = 5 across buckets; modal lists each correctly.", "High"),
        ("Combined: undo one of N", "Edit then partial revert",
         "1. Make 3 edits.\n2. Revert edit 2 manually.\n3. Strip.",
         "Strip drops the reverted one; remaining 2 still listed.", "Medium"),
        # Push verification
        ("After push, strip resets to empty", "Push cleanup",
         "1. Push.\n2. Strip.",
         "Strip empty; 'No unpushed changes'.", "High"),
        ("Strip remains across reload", "Persistence",
         "1. Edits made.\n2. Reload app.\n3. Strip.",
         "Strip restored from baseline diff.", "High"),
    ]
    for sub, title_snip, steps, expected, prio in scenarios:
        t(BOTH, "CP", "Edit Flow", "Functional",
          f"Edit -> Changes-to-Push: {sub}",
          "Linked workspace; clean state.", steps,
          "edit-to-push", expected, prio)


gen_edit_to_push_scenarios()


# ---- RESPONSE BODY x VIEWER MATRIX ----
def gen_response_viewer_matrix():
    sizes = [("0 B (empty)", "empty"), ("1 KB", "small"), ("100 KB", "medium"),
             ("1 MB", "large"), ("10 MB", "huge"), ("100 MB", "extreme")]
    content_types = ["application/json", "application/xml", "text/html",
                     "text/plain", "application/octet-stream", "text/csv",
                     "image/png", "image/jpeg", "application/pdf",
                     "text/event-stream", "application/x-yaml"]
    viewers = ["Pretty", "Raw", "Preview"]
    for size_label, size_class in sizes:
        for ct in content_types:
            for viewer in viewers:
                sub = f"{ct} {size_label} ({viewer})"
                title = f"Response body {ct} at {size_label} in {viewer} viewer"
                steps = (
                    f"1. Send request returning {ct} body of size {size_label}.\n"
                    f"2. Switch viewer to {viewer}.\n"
                    f"3. Inspect rendering and responsiveness."
                )
                expected = (
                    f"{viewer} renders appropriately for {ct} at {size_label}. "
                    f"For sizes >= 1 MB, preview cap may kick in with download CTA. "
                    f"UI remains responsive; no freeze."
                )
                ttype = "Performance" if size_class in ("large", "huge", "extreme") else "Functional"
                prio = "High" if size_class in ("small", "medium") and ct == "application/json" else (
                    "Medium" if size_class != "extreme" else "Low"
                )
                t(BOTH, "RP", "Viewer Matrix", ttype, title, "-", steps,
                  f"ct={ct} size={size_label} view={viewer}", expected, prio)


gen_response_viewer_matrix()


# ---- PARAM x METHOD MATRIX ----
def gen_params_method_matrix():
    scenarios = [
        ("Single param", "name=alice", "Echo confirms ?name=alice."),
        ("Multiple params", "a=1, b=2, c=3", "All in URL."),
        ("Duplicate keys", "foo=1, foo=2", "Both sent."),
        ("Unicode value", "q=测试", "Percent-encoded; server confirms."),
        ("Empty value", "a=", "Sent as a=."),
        ("Reserved chars in value", "q=a&b=c", "Encoded as a%26b%3Dc."),
        ("Disabled row", "a=1 disabled, b=2", "Only b=2 sent."),
        ("Variable in value", "q={{search}}", "Resolved value transmitted."),
        ("Path with query already", "URL?x=1; add y=2", "?x=1&y=2."),
        ("Long param value (4KB)", "q=<4KB>", "Sent; may be rejected by server."),
        ("Null value (treat as empty)", "a=null literal", "Sent as 'null' string."),
        ("Boolean-ish value", "active=true", "Sent verbatim 'true'."),
    ]
    for m in ["GET", "POST", "PUT", "PATCH", "DELETE"]:
        for sub, descr, expected_core in scenarios:
            full_sub = f"{sub} ({m})"
            title = f"Query params: {sub} on {m}"
            steps = f"1. Build request with params: {descr}.\n2. Send {m}."
            expected = expected_core + f" Verified on {m}."
            t(BOTH, "RE", "Params Matrix", "Functional",
              title, "-", steps, f"method={m} params={sub}", expected,
              "Medium" if sub in ("Single param", "Multiple params") else "Low")


gen_params_method_matrix()


# ---- AUTH REFRESH x SCENARIO MATRIX ----
def gen_auth_refresh_matrix():
    scenarios = [
        ("OAuth2 token unexpired", "Token still valid",
         "1. Send.",
         "Token used as-is; no refresh call.", "High"),
        ("OAuth2 token expired with refresh_token", "Expired",
         "1. Send.",
         "Auto-refresh; new token used.", "High"),
        ("OAuth2 token expired without refresh_token", "Expired no refresh",
         "1. Send.",
         "Send fails with auth error or prompts user; clear UI.", "High"),
        ("OAuth2 refresh fails (IdP 4xx)", "IdP rejects refresh",
         "1. Send.",
         "Error surfaced; user prompted to re-authorize.", "High"),
        ("OAuth2 token returned with no expiry", "No expires_in",
         "1. Send.",
         "Sent as-is; no proactive refresh.", "Medium"),
        ("OAuth2 token nearly expiring (skew)", "Within skew window",
         "1. Send.",
         "Refresh proactively per documented skew (e.g., <60s).", "Medium"),
        ("Bearer token interpolated from secret var", "Secret-backed token",
         "1. Send.",
         "Secret decrypted; used as token; not in plaintext history.", "High"),
        ("AWS SigV4 with rotating session token", "Session token in env",
         "1. Rotate token.\n2. Send.",
         "New token used; signature reflects new creds.", "Medium"),
        ("Digest replay on stale nonce", "Server returns stale=true",
         "1. Send after nonce rotation.",
         "Client retries with new nonce.", "Medium"),
        ("NTLM handshake interrupted", "Network drops mid-handshake",
         "1. Drop network between Challenge and Authenticate.",
         "Clean error; no half-state.", "Medium"),
        ("OAuth2 401 on resource after token acquired", "Token rejected by server",
         "1. Send.",
         "App offers to clear token / re-auth; does not loop.", "High"),
        ("Bearer token edited after acquisition", "User manually overrides",
         "1. Edit token field.",
         "Send uses user-edited token; refresh skipped.", "Medium"),
        ("OAuth2 multiple requests share token", "Two requests same OAuth config",
         "1. Send A; send B.",
         "Both use same access_token; only one token fetch.", "Medium"),
        ("OAuth2 different configs same provider", "Same IdP, different clientIds",
         "1. Each gets own token.",
         "Tokens scoped per request config; no cross-contamination.", "Medium"),
    ]
    for m in ["GET", "POST"]:
        for sub, pre, steps, expected, prio in scenarios:
            full_sub = f"{sub} on {m}"
            t(BOTH, "AU", "Refresh Matrix", "Functional",
              f"Auth refresh: {sub} ({m})", pre, steps,
              f"refresh method={m}", expected, prio)


gen_auth_refresh_matrix()


# ---- FOLDER/REQUEST/COLLECTION RENAME -> DOWNSTREAM REFERENCE CHECK ----
def gen_rename_reference_safety():
    """When entity X is renamed, downstream references should remain valid."""
    scenarios = [
        ("Rename request",
         "Plan steps referencing request still resolve",
         "1. Add request to plan.\n2. Rename request.\n3. Run plan.",
         "Plan still works; step displays new name.", "High"),
        ("Rename request",
         "History entries display old or current name consistently",
         "1. Run; rename.\n2. History.",
         "History entries point by id; show current name.", "Medium"),
        ("Rename folder",
         "Child request paths/IDs unchanged",
         "1. Rename folder.\n2. Verify children intact.",
         "Tree refreshes; no orphaned refs.", "High"),
        ("Rename folder",
         "Folder-level auth still inherited",
         "1. Folder had Bearer; rename.\n2. Child Inherit send.",
         "Bearer still applied.", "Medium"),
        ("Rename collection",
         "Children intact",
         "1. Rename collection.\n2. Verify.",
         "Tree shows new name; children OK.", "Medium"),
        ("Rename env",
         "Active env resolution still works",
         "1. Activate; rename.\n2. Send a request using env vars.",
         "Vars still resolve; resolution by id.", "High"),
        ("Rename env",
         "Linked env priority entry retained",
         "1. Reorder; rename one.",
         "Order preserved.", "Medium"),
        ("Rename env var",
         "Existing {{var}} refs need updating",
         "1. Rename baseUrl -> apiBase.\n2. Send requests using {{baseUrl}}.",
         "Old refs no longer resolve (empty); UI warns of refs to changed name.", "High"),
        ("Rename plan",
         "Plan run history retained",
         "1. Run plan; rename.\n2. History.",
         "History entries linked by id; show new name.", "Medium"),
        ("Rename mock server",
         "Endpoints inside intact",
         "1. Rename.\n2. Verify endpoints.",
         "Children intact.", "Medium"),
        ("Rename mock endpoint path",
         "Selector rules still apply",
         "1. Change path.\n2. Hit new path (desktop).",
         "Routes match new path; old path 404.", "Medium"),
        ("Rename JSON Schema",
         "bodySchemaId refs update",
         "1. Rename schema.\n2. Inspect a referencing request.",
         "Reference updates; validation still works.", "High"),
        ("Rename secret slot label",
         "Encrypted bytes unchanged",
         "1. Rename label.\n2. Verify decryption still works.",
         "Label updates; encryption intact.", "Medium"),
        ("Rename linked workspace alias",
         "Overrides keyed by id intact",
         "1. Rename.\n2. Verify overrides.",
         "Overrides keyed by id; intact.", "Medium"),
    ]
    for sub, title, steps, expected, prio in scenarios:
        t(BOTH, "CR", "Reference Safety", "Regression",
          title, "-", steps, sub, expected, prio)


gen_rename_reference_safety()


# ---- DELETE -> DOWNSTREAM IMPACT ----
def gen_delete_reference_safety():
    scenarios = [
        ("Delete request used in plan",
         "Plan handles missing step",
         "1. Add to plan.\n2. Delete.\n3. Run plan.",
         "Step shown as 'missing'; run continues (or stops) per design; clear message.", "High"),
        ("Delete folder with children",
         "Cascade deletes children",
         "1. Folder w/ 5 requests.\n2. Delete.",
         "Confirm mentions child count; cascade.", "High"),
        ("Delete collection with everything",
         "Full cascade",
         "1. Collection with nested.\n2. Delete.",
         "Full cascade; history references unlinked.", "High"),
        ("Delete env that is active",
         "Active env fallback",
         "1. Delete active.",
         "Fallback to no env or next priority.", "High"),
        ("Delete env that has linked overrides",
         "Overrides become orphaned",
         "1. Delete env on linked WS source.\n2. Consumer pulls.",
         "Consumer shows orphaned override warnings.", "Medium"),
        ("Delete env var referenced by {{var}}",
         "Refs resolve empty",
         "1. Delete var.\n2. Send referencing request.",
         "Empty resolution; UI shows reference warning where var appears.", "High"),
        ("Delete schema referenced by request",
         "Schema ref orphaned",
         "1. Delete.\n2. Open referencing request.",
         "Warning; body editor reverts to no-schema validation.", "High"),
        ("Delete mock endpoint",
         "Other endpoints unaffected",
         "1. Delete one.\n2. Verify others.",
         "Only that endpoint removed.", "Medium"),
        ("Delete mock server while running",
         "Stop runtime first or warn",
         "1. Mock running; delete server.",
         "App stops runtime before delete; or refuses with clear message.", "Medium"),
        ("Delete linked workspace",
         "Overrides removed",
         "1. Unlink.",
         "Overrides cleaned up; UI no longer shows linked tree.", "Medium"),
        ("Delete secret slot in use",
         "Secret references break",
         "1. Delete slot referenced by secret var.",
         "Variable shows '<missing slot>'; UI warns.", "High"),
    ]
    for sub, title, steps, expected, prio in scenarios:
        t(BOTH, "CR", "Delete Safety", "Negative",
          title, "-", steps, sub, expected, prio)


gen_delete_reference_safety()


# ---- ENCODING / CHARSET MATRIX ----
def gen_encoding_matrix():
    encodings = ["UTF-8", "UTF-16 LE BOM", "UTF-16 BE BOM", "ISO-8859-1",
                 "Windows-1252", "GBK", "Shift_JIS"]
    for enc in encodings:
        for body in ["raw-json", "raw-xml", "raw-text"]:
            sub = f"{enc} body ({body})"
            title = f"Body in {enc} encoding ({body})"
            steps = (f"1. Set body type {body}.\n"
                     f"2. Provide content encoded as {enc} (paste or import).\n"
                     f"3. Send.")
            expected = (
                f"App preserves byte sequence; Content-Type optionally annotates charset; "
                f"server receives intended bytes; response body decode honors response Content-Type charset."
            )
            prio = "Medium" if enc == "UTF-8" else "Low"
            t(BOTH, "BC", "Encoding", "Edge Case", title, "-", steps,
              f"enc={enc} body={body}", expected, prio)
        # Response side
        sub = f"Response in {enc}"
        title = f"Decode response body in {enc}"
        steps = (f"1. Send to endpoint that returns {enc}-encoded body.\n"
                 f"2. View response Pretty/Preview.")
        expected = (
            f"Response charset detected from Content-Type or BOM; rendered text correct; "
            f"raw bytes preserved on Download."
        )
        t(BOTH, "RP", "Encoding", "Edge Case", title, "-", steps,
          f"resp-enc={enc}", expected, "Low" if enc != "UTF-8" else "Medium")


gen_encoding_matrix()


# ---- CONCURRENT OPERATIONS MATRIX ----
def gen_concurrent_ops_matrix():
    scenarios = [
        ("Send + Send", "Two requests in flight",
         "1. Open Request A.\n2. Send A.\n3. Switch to Request B; Send.",
         "Both run independently; each updates its own response panel.", "High"),
        ("Send + Switch workspace", "Switch mid-flight",
         "1. Send.\n2. Switch workspace.",
         "Either request continues to completion in original workspace, or aborts cleanly. Documented behavior.", "Medium"),
        ("Send + Delete request", "Delete the request being sent",
         "1. Send.\n2. Delete the request.",
         "In-flight completes or aborts; history retains entry (or is sanitized) per design.", "Medium"),
        ("Send + Push", "Push during request",
         "1. Send.\n2. Click Push.",
         "Push queued or runs in parallel; final state has both completed.", "Medium"),
        ("Push + Pull race", "Sequential push/pull",
         "1. Push.\n2. Immediately Pull.",
         "Pull sees pushed commit; no resync conflict.", "Medium"),
        ("Edit + Send same request", "Concurrent edits",
         "1. Type in body.\n2. Send before debounce flush.",
         "Send uses latest editor state; nothing lost.", "Medium"),
        ("Plan running + Edit step", "Edit during plan run",
         "1. Run plan.\n2. Edit step 3 mid-run.",
         "Either pause until end, or edit applied to next iteration; clear behavior.", "Medium"),
        ("Edit env var + Send request using var", "Mid-typing variable change",
         "1. Type new value.\n2. Send before debounce flush.",
         "Send uses latest unflushed value (in-memory).", "Medium"),
        ("Refresh during local edits", "Pull with unsaved",
         "1. Edit; click Refresh.",
         "Pull merges or warns about unsaved.", "Medium"),
        ("Two tab edits in web app", "Same workspace two tabs",
         "1. Type in A and B.",
         "Last-write wins; no IDB corruption.", "Medium"),
        ("Mock running + edit definition", "Edit while running",
         "1. Mock running; edit endpoint.",
         "Edit applied on next request (auto-reload) or warn to restart.", "Medium"),
    ]
    for sub, title, steps, expected, prio in scenarios:
        t(BOTH, "CC", "Concurrency", "Edge Case",
          f"Concurrency: {title}", "-", steps, sub, expected, prio)


gen_concurrent_ops_matrix()


# ---- LARGE-SCALE STRESS / BOUNDARY MATRIX ----
def gen_stress_matrix():
    boundaries = [
        ("Workspace with 10000 requests", "Performance",
         "1. Open workspace.", "Tree virtualized; load <10s; scrolling smooth.", "Medium"),
        ("Workspace with 100 collections", "Performance",
         "1. Open.", "Tree loads <3s.", "Low"),
        ("Workspace with 500 folders", "Performance",
         "1. Open.", "Tree responsive.", "Low"),
        ("Workspace with 50 environments", "Performance",
         "1. Switch env.", "Picker responsive; switch <1s.", "Low"),
        ("Workspace with 100 environment vars", "Performance",
         "1. Open Variables panel.", "Loads <1s; filtering responsive.", "Low"),
        ("Workspace with 10 mock servers", "Functional",
         "1. List mocks.", "All shown.", "Low"),
        ("Mock server with 200 endpoints", "Performance",
         "1. Open mock; scroll endpoints.", "Smooth.", "Low"),
        ("Plan with 100 steps", "Performance",
         "1. Run.", "Each step shown; UI does not freeze.", "Medium"),
        ("Request with 50 headers", "Edge Case",
         "1. Send.", "All transmitted.", "Low"),
        ("Request with 50 query params", "Edge Case",
         "1. Send.", "URL contains all.", "Low"),
        ("Body 100MB raw text", "Performance",
         "1. Send.", "Either uploads or surfaces size warning; no crash.", "Low"),
        ("Response 100MB binary", "Performance",
         "1. Send.", "Preview cap; download streams.", "Medium"),
        ("History with 5000 runs", "Performance",
         "1. Open history; filter.", "Virtualized; filter responsive.", "Medium"),
        ("Workspace with 1000 history runs and switch", "Performance",
         "1. Switch workspaces.", "<2s.", "Medium"),
        ("Unicode-heavy workspace (all names CJK)", "Edge Case",
         "1. Open.", "All names render; search works.", "Low"),
        ("Very deep folder nesting (20 levels)", "Edge Case",
         "1. Open.", "Renders; no recursion limit hit.", "Low"),
        ("Request name with 500 chars", "Edge Case",
         "1. Create.", "Truncated or accepted; tooltip shows full.", "Low"),
        ("URL with 16KB length", "Edge Case",
         "1. Send.", "Either rejected by client/server with clear error or sent.", "Low"),
    ]
    for sub, ttype, steps, expected, prio in boundaries:
        t(BOTH, "PE", "Boundaries", ttype, f"Stress: {sub}", "-", steps, sub, expected, prio)


gen_stress_matrix()


# ---- TRUST / PERMISSION FLOWS (GIT/GITHUB) ----
def gen_github_flow_matrix():
    flows = [
        ("Link to public repo", "Public repo owned by user",
         "1. Link.", "Repo accepted; working branch created.", "High"),
        ("Link to private repo with personal token", "Private with full scope",
         "1. Link.", "Accepted; private indicator shown.", "High"),
        ("Link to org repo (member, write)", "Org member write",
         "1. Link.", "Accepted.", "Medium"),
        ("Link to org repo (member, read-only)", "Read-only",
         "1. Link; attempt Push.", "Push fails with clear permission error; read still works.", "Medium"),
        ("Link to org repo (non-member, public)", "Non-member public",
         "1. Link.", "Allowed for read; push fails clearly.", "Low"),
        ("Link to org repo (non-member, private)", "Private not accessible",
         "1. Link.", "Permission error.", "Medium"),
        ("Link to repo with branch protection", "Main protected",
         "1. Link.", "Working branch policy honored; push to main blocked.", "Medium"),
        ("Link to archived repo", "Archived repo",
         "1. Link.", "Read OK; push blocked with clear error.", "Low"),
        ("Link to forked repo (upstream PR)", "User fork",
         "1. Link.", "Working branch on fork; PR opens against upstream.", "Medium"),
        ("OAuth scope downgrade after linking", "User revoked scope",
         "1. Push.", "Detects insufficient scope; prompts re-auth with needed scopes.", "Medium"),
        ("Repo deleted after linking", "Remote repo missing",
         "1. Pull.", "Clear error 'repo not found'; offer unlink.", "Medium"),
        ("Repo renamed by owner", "Path moved",
         "1. Pull.", "Detect rename via GitHub redirect; update link or prompt.", "Medium"),
        ("Repo transferred to another owner", "Ownership change",
         "1. Pull.", "Detect transfer; prompt to update link.", "Medium"),
        ("Branch protection requires status checks", "Required CI",
         "1. PR open.", "PR remains open until checks; user informed.", "Low"),
        ("Push to branch with PR draft", "PR draft",
         "1. Push.", "Adds commits to PR; PR updated.", "Medium"),
        ("PR merged via squash on GitHub", "Squash merge",
         "1. PR merged.\n2. Refresh.",
         "Working branch retired; local snapshot updated to squashed main; "
         "user prompted for next branch.", "High"),
        ("PR merged via rebase on GitHub", "Rebase merge",
         "1. Merge.\n2. Refresh.",
         "History updated; identical content; retire flow works.", "Medium"),
        ("PR merged via merge commit", "Merge commit",
         "1. Merge.\n2. Refresh.",
         "Same.", "Medium"),
        ("Direct push to main by collaborator", "Direct main edits",
         "1. Refresh.",
         "Local pulls main delta; conflict if local diverged.", "Medium"),
        ("Force-push on working branch", "Remote force push",
         "1. Pull.",
         "Detected; user warned; reset path offered.", "Medium"),
        ("Concurrent push from two devices", "Race",
         "1. Both push.",
         "Second fails; pull+merge required.", "High"),
        ("OAuth token revoked on github.com mid-session", "Revoked",
         "1. Push.",
         "401 detected; re-auth prompted; edits preserved.", "High"),
        ("Network drops during push (large)", "Mid-upload drop",
         "1. Push large; drop network.",
         "Push fails cleanly; edits intact; retry button.", "Medium"),
        ("Network drops during pull", "Mid-download drop",
         "1. Drop.",
         "Pull fails cleanly; existing state intact.", "Medium"),
        ("Workspace push includes secrets metadata only (not values)", "Privacy",
         "1. Push; inspect commit.",
         "Commit has secret slot metadata and ciphertext-as-stored; "
         "no plaintext secret values.", "High"),
        ("Push of conflict resolution", "After resolving",
         "1. Push resolved state.",
         "New commit pushed; sync snapshot advanced.", "Medium"),
        ("Open PR shows in workspace UI", "PR badge",
         "1. PR open.",
         "App shows 'PR open' badge with link to GitHub.", "Low"),
    ]
    for sub, pre, steps, expected, prio in flows:
        title = f"GitHub flow: {sub}"
        t(BOTH, "GT", "GitHub Flow", "Functional", title, pre, steps,
          sub, expected, prio)


gen_github_flow_matrix()


# ---- ASSERTION OPERATOR x TARGET MATRIX ----
def gen_assertion_operator_matrix():
    operators = ["equals", "not-equals", "contains", "not-contains",
                 "starts-with", "ends-with", "matches (regex)",
                 "lt", "lte", "gt", "gte", "exists", "not-exists",
                 "is-type (string/number/boolean/null/array/object)",
                 "length-equals", "length-gt", "length-lt"]
    targets = [
        ("HTTP status", "status code (e.g., 200)"),
        ("Response time (ms)", "responseTime"),
        ("Response size (bytes)", "responseSize"),
        ("Response header", "header named Content-Type"),
        ("Response body (JSON path)", "body.user.id"),
        ("Response body (raw text)", "body as string"),
        ("Response body matches schema", "JSON schema validation"),
        ("Cookie value", "cookie 'session'"),
    ]
    for target_name, target_descr in targets:
        for op in operators:
            sub = f"{target_name} {op}"
            title = f"Assertion: {target_name} {op}"
            steps = (
                f"1. Add assertion: target='{target_descr}', operator='{op}', "
                f"expected=<a value that should pass>.\n"
                f"2. Send and observe Tests result."
            )
            expected = (
                f"Pass case: assertion green; actual matches expected per operator. "
                f"Failure case (separately): red row with clear actual vs expected delta. "
                f"Operator '{op}' applies correct semantics."
            )
            ttype = "Functional"
            prio = "High" if op in ("equals", "contains", "matches (regex)",
                                     "lt", "gt", "exists") and target_name in (
                "HTTP status", "Response body (JSON path)", "Response time (ms)") else "Medium"
            t(BOTH, "SC", "Assertion Matrix", ttype, title,
              "-", steps, f"target={target_name} op={op}", expected, prio)


gen_assertion_operator_matrix()


# ---- PRE-REQUEST / TEST SCRIPT API SURFACE MATRIX ----
def gen_script_api_matrix():
    api_surfaces = [
        ("pm.variables.set(name, value)", "Variable set",
         "1. Pre-script sets var.\n2. Use {{var}} in body.\n3. Send.",
         "Body uses the set value.", "High"),
        ("pm.variables.get(name)", "Variable get",
         "1. Pre-script reads existing var.\n2. console.log it.",
         "Logged value matches stored.", "High"),
        ("pm.variables.unset(name)", "Variable unset",
         "1. Unset; reference {{var}}.",
         "Resolution empty.", "Medium"),
        ("pm.environment.set/get", "Env access",
         "1. Pre-script reads/writes env.",
         "Env updated; reflects in active env panel.", "High"),
        ("pm.request.headers.add", "Modify request headers",
         "1. Pre-script adds header.\n2. Send.",
         "Header present in outgoing request.", "High"),
        ("pm.request.headers.remove", "Remove header",
         "1. Pre-script removes Authorization.\n2. Send.",
         "Header not sent (auth disabled by script).", "Medium"),
        ("pm.request.url.set", "Mutate URL",
         "1. Pre-script overrides URL.\n2. Send.",
         "Modified URL used.", "High"),
        ("pm.request.body.set", "Mutate body",
         "1. Pre-script sets body.\n2. Send.",
         "Body matches scripted value.", "High"),
        ("pm.request.auth.set", "Mutate auth at runtime",
         "1. Pre-script changes auth type.\n2. Send.",
         "Auth applied per script.", "Medium"),
        ("pm.sendRequest (helper)", "Async sub-request",
         "1. Pre-script does pm.sendRequest({...}, cb).",
         "Sub-request completes before main; cb invoked.", "High"),
        ("pm.response.json()", "Parse JSON in tests",
         "1. Test script reads body as JSON.",
         "Object usable.", "High"),
        ("pm.response.text()", "Raw text",
         "1. Test script reads as string.",
         "String returned.", "Medium"),
        ("pm.response.headers.get", "Read response header",
         "1. Test reads header.",
         "Value returned (case-insensitive).", "Medium"),
        ("pm.response.code", "Status code",
         "1. Test references .code.", "Integer status.", "Medium"),
        ("pm.response.responseTime", "Duration", "1. Test reads.", "ms number.", "Low"),
        ("pm.test('name', fn)", "Assertion wrapper",
         "1. Test wraps assertion.", "Result tracked under name.", "High"),
        ("pm.expect", "Chai expect-like",
         "1. Test uses pm.expect(x).to.equal(1).",
         "Standard expect semantics.", "High"),
        ("console.log / warn / error", "Console",
         "1. Each level.",
         "Each renders with level color.", "Medium"),
        ("setTimeout / setInterval", "Timers",
         "1. setTimeout in pre-script.",
         "Either supported with timeout limit, or rejected.", "Low"),
        ("crypto operations (subtle)", "Crypto API",
         "1. Use crypto.subtle.digest.",
         "Works in sandbox.", "Low"),
        ("JSON.parse / stringify", "Standard JSON",
         "1. Use.",
         "Works.", "Medium"),
        ("Date / Math", "Standard globals",
         "1. Use Date.now().",
         "Works.", "Low"),
        ("Throw inside pre-script", "Error",
         "1. Throw.",
         "Send aborted; error surfaces.", "High"),
        ("Async/await", "Promises",
         "1. await fetch (or pm.sendRequest).",
         "Resolves before send.", "Medium"),
        ("Long-running script (5s)", "Slow script",
         "1. Pre-script with 5s loop.",
         "Script timeout per limit; surfaced as error.", "Medium"),
        ("Script accesses fetch (if allowed)", "Sandbox fetch",
         "1. fetch() call.",
         "Either allowed via sandbox or rejected with clear note.", "Low"),
        ("Script accesses localStorage (should be blocked)",
         "Security",
         "1. localStorage.setItem.",
         "Blocked or no-op in sandbox.", "High"),
        ("Script imports modules (should be blocked)", "Imports",
         "1. import ...",
         "Blocked / not supported; clear error.", "Medium"),
        ("Script with infinite loop", "Hang",
         "1. while(true){}.",
         "Detected by timeout; send aborted.", "High"),
        ("Script across pre and tests sees same variables", "Shared scope",
         "1. Pre sets x; tests reads x.",
         "Same context; value visible.", "Medium"),
    ]
    for sub, descr, steps, expected, prio in api_surfaces:
        ttype = "Security" if "Security" in descr or "blocked" in expected.lower() else "Functional"
        t(BOTH, "SC", "Script API", ttype,
          f"Script API: {sub}", "-", steps, sub, expected, prio)


gen_script_api_matrix()


# ---- UI NAVIGATION x WORKSPACE STATE MATRIX ----
def gen_ui_state_matrix():
    panels = ["Workspace", "Editor", "Environments", "Execution", "Mocks",
              "History", "Help", "MCP"]
    states = ["Empty workspace", "Loaded workspace with data",
              "Workspace with active edits", "Workspace mid-pull",
              "Workspace with conflict modal open",
              "Workspace with secrets locked (no passphrase)"]
    for panel in panels:
        for state in states:
            sub = f"{panel} ({state})"
            title = f"Open {panel} panel while: {state}"
            steps = f"1. Set workspace state: {state}.\n2. Open {panel} panel."
            expected = (
                f"Panel renders without error; relevant data shown or empty state; "
                f"actions appropriately enabled/disabled for state '{state}'."
            )
            ttype = "UX/UI"
            prio = "Medium" if state in ("Loaded workspace with data",
                                           "Workspace with secrets locked (no passphrase)") else "Low"
            t(BOTH, "CC", "Panel State", ttype, title, "-", steps,
              f"panel={panel} state={state}", expected, prio)


gen_ui_state_matrix()


# ---- THEMING / FONT / SIZE COMBINATIONS ----
def gen_theme_matrix():
    themes = ["studio-dark", "workbench-light", "dracula", "nord",
              "tokyo-night", "monokai-pro", "high-contrast", "solarized-light",
              "github-light", "rose-pine"]
    surfaces = ["Editor panel", "Workspace explorer", "Response panel",
                "Variables panel", "Mocks list", "Help"]
    for theme in themes:
        for surf in surfaces:
            sub = f"{theme} / {surf}"
            title = f"Render {surf} under theme '{theme}'"
            steps = (f"1. Set theme {theme}.\n2. Navigate to {surf}.\n"
                     f"3. Verify text/icon contrast and color tokens.")
            expected = (
                f"Theme tokens applied consistently; text legible against background; "
                f"focus rings still visible; no contrast regressions."
            )
            t(BOTH, "ST", "Theme Matrix", "UX/UI", title, "-", steps,
              f"theme={theme} surface={surf}", expected,
              "Medium" if theme in ("studio-dark", "workbench-light", "high-contrast") else "Low")


gen_theme_matrix()


# =====================================================================
# MCP (MODEL CONTEXT PROTOCOL) - desktop only (CLI binary apicircle-mcp)
# =====================================================================
# 50 tools across 8 namespaces, 8+ AI client integrations.

MCP_TOOL_CATALOG = [
    # Imports (5)
    ("import.curl", "Parse cURL command to Request",
     "{ source: '<curl text>' }", "request"),
    ("import.openapi", "Parse OpenAPI/Swagger to requests",
     "{ source, format }", "requests[]"),
    ("import.postman", "Parse Postman v2/v2.1",
     "{ source }", "requests[]"),
    ("import.insomnia", "Parse Insomnia v4",
     "{ source }", "requests[]"),
    ("import.har", "Parse HAR file",
     "{ source }", "requests[]"),
    # Codegen (1)
    ("generate.code", "Codegen for request",
     "{ id, language }", "code string"),
    # Workspace (2)
    ("workspace.read", "Snapshot full workspace",
     "{}", "{ synced, local }"),
    ("workspace.write", "Bulk overwrite synced/local",
     "{ synced, local }", "{ ok }"),
    # Request CRUD (4)
    ("request.create", "Create new request",
     "{ name, method, url, folderId? }", "{ id, changedIds }"),
    ("request.read", "Get one or list",
     "{ id? }", "request | request[]"),
    ("request.update", "Patch a request",
     "{ id, patch }", "{ changedIds }"),
    ("request.delete", "Delete a request",
     "{ id }", "{ changedIds }"),
    # Folder CRUD (4)
    ("folder.create", "Create folder",
     "{ name, parentId? }", "{ id, changedIds }"),
    ("folder.read", "Get one or list",
     "{ id? }", "folder | folder[]"),
    ("folder.update", "Patch folder",
     "{ id, parentId? }", "{ changedIds }"),
    ("folder.delete", "Delete folder",
     "{ id }", "{ changedIds }"),
    # Environment CRUD (8)
    ("environment.create", "Create environment",
     "{ name, variables }", "{ changedIds }"),
    ("environment.read", "Get one or list",
     "{ id? }", "env | env[]"),
    ("environment.update", "Patch env",
     "{ id, patch }", "{ changedIds }"),
    ("environment.delete", "Delete env",
     "{ id }", "{ changedIds }"),
    ("environment.set_active", "Set active env",
     "{ id }", "{ changedIds }"),
    ("environment.set_priority", "Set priority order",
     "{ id, order }", "{ changedIds }"),
    ("environment.export", "Export env",
     "{ id, format }", "string"),
    ("environment.import", "Import env",
     "{ name, format, content }", "{ changedIds }"),
    # Plan CRUD (9)
    ("plan.create", "Create plan",
     "{ name, stepRequestIds, envPriorityOrder? }", "{ ok, id, changedIds }"),
    ("plan.read", "Get one or list",
     "{ id? }", "plan | plan[]"),
    ("plan.update", "Patch plan",
     "{ id, patch }", "{ changedIds }"),
    ("plan.delete", "Delete plan",
     "{ id }", "{ changedIds }"),
    ("plan.run", "Execute all steps",
     "{ id, env?, variables? }", "{ runId, results[] }"),
    ("plan.add_step", "Append step",
     "{ planId, requestId }", "{ changedIds }"),
    ("plan.remove_step", "Remove step by index",
     "{ planId, stepIndex }", "{ changedIds }"),
    ("plan.reorder_steps", "Reorder steps",
     "{ planId, newOrder }", "{ changedIds }"),
    ("plan.set_variables", "Set plan-scoped vars",
     "{ planId, variables }", "{ changedIds }"),
    # Assertion CRUD (4)
    ("assertion.create", "Create assertion",
     "{ requestId, assertion }", "{ id, changedIds }"),
    ("assertion.read", "Get one or list",
     "{ id? }", "assertion | assertion[]"),
    ("assertion.update", "Patch assertion",
     "{ id, patch }", "{ changedIds }"),
    ("assertion.delete", "Delete assertion",
     "{ id }", "{ changedIds }"),
    # History (4)
    ("history.list_runs", "List run summaries",
     "{ requestId?, ok?, since?, until?, limit<=500 }", "runs[]"),
    ("history.get_run", "Get full run detail",
     "{ id }", "run | { found:false }"),
    ("history.delete_run", "Delete a run",
     "{ id }", "{ deleted:N }"),
    ("history.purge_by_age", "Purge older than N days",
     "{ olderThanDays }", "{ purgedCount }"),
    # Codebase (1)
    ("codebase.extract_collection", "Detect routes from source",
     "{ source, frameworks? }", "candidates[]"),
    # Prompt-driven (13)
    ("prompt.create_environment", "LLM-shaped env",
     "{ json }", "{ id, changedIds }"),
    ("prompt.create_assertion", "LLM-shaped assertion",
     "{ requestId, json }", "{ id, changedIds }"),
    ("prompt.create_plan", "LLM-shaped plan",
     "{ json }", "{ id, changedIds }"),
    ("prompt.create_request", "LLM-shaped request",
     "{ json }", "{ id, changedIds }"),
    ("prompt.update_request", "LLM-shaped update",
     "{ id, json }", "{ changedIds }"),
    ("prompt.create_folder_tree", "LLM-shaped tree",
     "{ json }", "{ ids[], changedIds }"),
    ("prompt.add_plan_steps", "LLM-shaped add steps",
     "{ planId, json }", "{ changedIds }"),
    ("prompt.set_plan_variables", "LLM-shaped plan vars",
     "{ planId, json }", "{ changedIds }"),
    ("prompt.create_mock_server", "LLM-shaped mock",
     "{ json }", "{ id, changedIds }"),
    ("prompt.add_mock_endpoint", "LLM-shaped endpoint",
     "{ mockId, json }", "{ changedIds }"),
    ("prompt.set_endpoint_validation_rules", "LLM rules",
     "{ mockId, endpointId, json }", "{ changedIds }"),
    ("prompt.set_endpoint_response_rules", "LLM rules",
     "{ mockId, endpointId, json }", "{ changedIds }"),
    ("prompt.set_endpoint_multipliers", "LLM multipliers",
     "{ mockId, endpointId, json }", "{ changedIds }"),
    # Mock CRUD + lifecycle (15)
    ("mock.create_from_openapi", "OpenAPI -> Mock",
     "{ name, spec, format }", "{ id, endpointCount, warnings }"),
    ("mock.create_from_postman", "Postman -> Mock",
     "{ name, collection }", "{ id, endpointCount, warnings }"),
    ("mock.create_from_insomnia", "Insomnia -> Mock",
     "{ name, export }", "{ id, endpointCount, warnings }"),
    ("mock.create_manual", "Manual mock",
     "{ name, endpoints }", "{ id }"),
    ("mock.list", "List mocks + status",
     "{}", "mocks[]"),
    ("mock.list_endpoints", "List endpoints of a mock",
     "{ id }", "endpoints[]"),
    ("mock.start", "Boot runtime",
     "{ id, port? }", "{ port, pid, startedAt }"),
    ("mock.stop", "Stop runtime",
     "{ id }", "{ ok }"),
    ("mock.delete", "Delete mock",
     "{ id }", "{ changedIds }"),
    ("mock.add_endpoint", "Add endpoint",
     "{ id, endpoint }", "{ changedIds }"),
    ("mock.update_endpoint", "Patch endpoint",
     "{ id, endpointId, patch }", "{ changedIds }"),
    ("mock.delete_endpoint", "Delete endpoint",
     "{ id, endpointId }", "{ changedIds }"),
    ("mock.set_validation_rules", "Endpoint validation",
     "{ id, endpointId, rules }", "{ changedIds }"),
    ("mock.set_response_rules", "Endpoint response rules",
     "{ id, endpointId, rules }", "{ changedIds }"),
    ("mock.set_multipliers", "Response multipliers",
     "{ id, endpointId, multipliers }", "{ changedIds }"),
    ("mock.import_postman_mock_collection", "Postman mock collection",
     "{ name, collection }", "{ id, endpointCount }"),
]

AI_CLIENTS = [
    ("claude-desktop", "Claude Desktop", "~/Library/Application Support/Claude/claude_desktop_config.json (macOS)"),
    ("cursor", "Cursor", "~/.cursor/mcp.json"),
    ("continue", "Continue", "~/.continue/config.json"),
    ("zed", "Zed", "~/.config/zed/settings.json"),
    ("claude-code", "Claude Code", "~/.claude.json"),
    ("cline", "Cline", "VS Code extension config"),
    ("windsurf", "Windsurf", "Windsurf config"),
    ("github-copilot", "GitHub Copilot", "Copilot config"),
    ("chatgpt", "ChatGPT", "ChatGPT config"),
    ("generic", "Generic MCP client", "stdio command line"),
]


def gen_mcp_lifecycle():
    cases = [
        ("Boot with --workspace flag", "Functional",
         "1. Run `apicircle-mcp --workspace ./ws` from terminal.",
         "Process stays alive on stdio; stderr prints 'apicircle-mcp ready · workspace=<dir>'; exit 0 on Ctrl+C.",
         "High"),
        ("Boot with APICIRCLE_WORKSPACE env var", "Functional",
         "1. APICIRCLE_WORKSPACE=./ws apicircle-mcp.",
         "Resolves env var; ready message.", "High"),
        ("Boot with cwd fallback", "Functional",
         "1. cd into workspace dir; run `apicircle-mcp` with no args.",
         "Uses cwd; ready.", "Medium"),
        ("Flag precedence over env var", "Edge Case",
         "1. APICIRCLE_WORKSPACE=./a; run with --workspace ./b.",
         "Uses ./b (flag wins).", "Medium"),
        ("Boot with missing workspace dir creates it", "Functional",
         "1. --workspace ./new-dir (does not exist).",
         "ensureWorkspace initializes files; ready.", "Medium"),
        ("Boot with unwritable workspace dir", "Negative",
         "1. --workspace /read-only-mount.",
         "stderr 'apicircle-mcp boot error: ...'; exit 1.", "High"),
        ("Boot with corrupt workspace.json", "Negative",
         "1. Corrupt synced.json; boot.",
         "stderr parse error; exit 1; protocol never opens.", "High"),
        ("Boot with empty --workspace value", "Negative",
         "1. --workspace ''.",
         "Validation error; exit 1.", "Low"),
        ("Boot with workspace containing only synced.json (no local)", "Edge Case",
         "1. Boot.",
         "Initializes missing local store; ready.", "Medium"),
        ("Ready message goes to stderr (not stdout)", "Regression",
         "1. Boot; capture stdout vs stderr.",
         "stdout is pure JSON-RPC; stderr has human log lines.", "High"),
        ("Stdout never emits non-JSON-RPC bytes during runtime", "Regression",
         "1. Boot; observe stdout.",
         "Only JSON-RPC frames; nothing else (would break MCP framing).", "High"),
        ("Graceful shutdown when stdin closes", "Functional",
         "1. Close stdin from client.",
         "Server closes via host.close(); exit 0; no orphan handles.", "Medium"),
        ("Force-kill leaves no zombie processes", "Edge Case",
         "1. kill -9 the binary.",
         "Process terminates; no listeners persist; child mocks (if any) cleaned up by OS reaper.", "Medium"),
        ("Workspace symlink resolved", "Edge Case",
         "1. --workspace points to a symlink.",
         "Resolves and uses target dir.", "Low"),
        ("Workspace path with spaces", "Edge Case",
         "1. --workspace '/path with spaces/ws'.",
         "Handled correctly.", "Low"),
        ("Workspace path with unicode", "Edge Case",
         "1. --workspace ./测试-ws.",
         "Handled.", "Low"),
        ("Two MCP servers on same workspace concurrently", "Edge Case",
         "1. Run two binaries on same dir.",
         "File-backed provider locks or last-write-wins per design; no corruption.", "High"),
        ("MCP boot does not require network", "Functional",
         "1. Disconnect network; boot.",
         "Ready; runtime works for read-only ops until network needed (run requests).", "Medium"),
    ]
    for sub, ttype, steps, expected, prio in cases:
        t(DESK, "MC", "Lifecycle", ttype, f"MCP lifecycle: {sub}",
          "CLI installed.", steps, "mcp lifecycle", expected, prio)


gen_mcp_lifecycle()


def gen_mcp_protocol_compliance():
    """JSON-RPC 2.0 + MCP spec compliance."""
    cases = [
        ("Handshake initialize", "Functional",
         "1. Send `initialize` JSON-RPC with capabilities.\n2. Read response.",
         "Server replies with serverInfo, protocolVersion, capabilities; matches MCP spec.",
         "High"),
        ("initialized notification after initialize", "Functional",
         "1. Send `notifications/initialized`.",
         "No response (notification); server now ready for tool calls.", "High"),
        ("tools/list returns full catalog", "Functional",
         "1. Send tools/list.",
         "Response.tools has 50 entries (all namespaces); each has name, description, inputSchema.",
         "High"),
        ("tools/list pagination (cursor) if supported", "Edge Case",
         "1. Send tools/list with cursor.",
         "Either pagination supported or single page returned.", "Low"),
        ("tools/call with valid args", "Functional",
         "1. Call request.create with valid args.",
         "Result content[0].type='text'; text is JSON-stringified result; isError absent.", "High"),
        ("tools/call with missing required arg", "Negative",
         "1. Call request.create without 'name'.",
         "Response has isError:true; content text 'Validation failed: name: ...'.", "High"),
        ("tools/call with wrong type arg", "Negative",
         "1. Call request.create with method=42 (number not string).",
         "isError; validation message with path and expected type.", "High"),
        ("tools/call with unknown tool", "Negative",
         "1. tools/call name='nope.bogus'.",
         "JSON-RPC error -32601 Method not found, or tool isError; protocol-compliant.", "Medium"),
        ("Malformed JSON-RPC frame", "Negative",
         "1. Write invalid JSON to stdin.",
         "Server emits -32700 Parse error; connection remains alive.", "High"),
        ("Request with missing jsonrpc field", "Negative",
         "1. Send {id:1, method:'tools/list'}.",
         "-32600 Invalid Request.", "Medium"),
        ("Request with wrong jsonrpc version", "Negative",
         "1. jsonrpc='1.0'.",
         "-32600 Invalid Request.", "Medium"),
        ("Notification (no id) gets no reply", "Functional",
         "1. Send a notification.",
         "No response sent; server still processes.", "Medium"),
        ("Concurrent requests interleaved by id", "Edge Case",
         "1. Send 3 requests with id 1,2,3 quickly.",
         "Responses come back; ids match; order may differ but matches by id.", "Medium"),
        ("Large response (>1MB JSON)", "Performance",
         "1. workspace.read on large workspace.",
         "Streamed via stdio without framing errors; client receives intact.", "Medium"),
        ("Binary not embedded (only text content)", "Regression",
         "1. Any tool result.",
         "content[].type is 'text'; no binary blobs in MCP frame.", "Medium"),
        ("Server stays alive after tool error", "Functional",
         "1. Trigger isError; then call another tool.",
         "Second call succeeds.", "High"),
        ("Capability negotiation - no resources", "Functional",
         "1. initialize.",
         "capabilities.resources absent; only tools advertised.", "Medium"),
        ("Capability negotiation - no prompts (if not exposed)", "Functional",
         "1. initialize.",
         "capabilities.prompts absent or empty per design.", "Low"),
        ("Tool result JSON pretty-printed", "Regression",
         "1. Any tool call.",
         "content.text is JSON.stringify(result, null, 2).", "Low"),
        ("Process never writes log lines to stdout", "Regression",
         "1. Trigger any internal warning.",
         "All logs to stderr; stdout pure protocol.", "High"),
        ("Cancel notification (notifications/cancelled) honored", "Functional",
         "1. Start plan.run; send cancelled.",
         "Run aborts gracefully; response returns partial/aborted result.", "Medium"),
        ("Progress notifications during plan.run (if supported)", "Edge Case",
         "1. plan.run on long plan.",
         "Either progress emitted or single final response; documented.", "Low"),
    ]
    for sub, ttype, steps, expected, prio in cases:
        t(DESK, "MC", "Protocol", ttype, f"MCP protocol: {sub}",
          "MCP server running with test workspace.", steps, "json-rpc / mcp", expected, prio)


gen_mcp_protocol_compliance()


def gen_mcp_tools_per_tool():
    """For every tool, generate a happy-path + 1-2 negative/edge tests."""
    for tool_name, descr, input_shape, output_shape in MCP_TOOL_CATALOG:
        namespace = tool_name.split(".")[0]
        sub = f"{tool_name}"
        # Happy path
        steps = (f"1. Boot MCP server with test workspace.\n"
                 f"2. tools/call name='{tool_name}' arguments={input_shape}.\n"
                 f"3. Inspect response.")
        expected = (f"Tool returns expected shape: {output_shape}. "
                    f"isError absent. For mutations: changedIds includes affected entities; "
                    f"subsequent workspace.read reflects the change.")
        t(DESK, "MC", namespace, "Functional", f"MCP tool {tool_name}: happy path",
          "MCP server running.", steps, f"tool={tool_name}", expected,
          "High" if namespace in ("request", "environment", "plan", "mock", "workspace") else "Medium")

        # Validation
        steps_v = (f"1. tools/call name='{tool_name}' arguments={{}} (empty/wrong shape).")
        expected_v = (f"Validation error: isError:true with 'Validation failed: ...' message; "
                      f"server stays alive; workspace state unchanged.")
        t(DESK, "MC", namespace, "Negative", f"MCP tool {tool_name}: validation",
          "MCP server running.", steps_v, f"tool={tool_name}", expected_v, "Medium")

        # CRUD-specific edges
        if tool_name.endswith(".update") or tool_name.endswith(".delete"):
            steps_e = f"1. Call {tool_name} with id='non-existent'."
            expected_e = ("Tool returns { found:false } or graceful error; isError set "
                          "or ok:false with descriptive details.")
            t(DESK, "MC", namespace, "Negative",
              f"MCP tool {tool_name}: missing target",
              "MCP server running.", steps_e, f"tool={tool_name}", expected_e, "Medium")

        if tool_name.endswith(".read"):
            steps_l = f"1. Call {tool_name} with no id (list mode)."
            expected_l = "Returns array; respects any pagination or limit defaults."
            t(DESK, "MC", namespace, "Functional",
              f"MCP tool {tool_name}: list mode",
              "MCP server running.", steps_l, f"tool={tool_name}", expected_l, "Medium")

        if tool_name == "plan.run":
            t(DESK, "MC", "plan", "Functional", "plan.run with env override",
              "Plan exists.", "1. Call plan.run with env='Staging'.",
              "tool=plan.run env=Staging",
              "Run uses staging env values; results capture per-step status.", "High")
            t(DESK, "MC", "plan", "Functional", "plan.run with variables map override",
              "Plan exists.", "1. Call plan.run with variables={token:'abc'}.",
              "tool=plan.run variables=...",
              "Plan steps see {{token}}=abc.", "High")
            t(DESK, "MC", "plan", "Negative", "plan.run on plan with missing referenced request",
              "Plan references deleted request.", "1. Call plan.run.",
              "tool=plan.run",
              "Returns ok:false with details listing missing request ids; no partial commit.", "High")

        if tool_name == "history.list_runs":
            t(DESK, "MC", "history", "Edge Case", "history.list_runs hard cap at 500",
              "Workspace has >500 runs.", "1. Call with limit=10000.",
              "tool=history.list_runs",
              "Result respects max of 500; further pages would need follow-ups.", "Medium")

        if tool_name == "import.openapi":
            t(DESK, "MC", "import", "Edge Case", "import.openapi with circular $ref",
              "OpenAPI source with circular references.", "1. Call.",
              "tool=import.openapi",
              "Either resolves with safe depth limit or returns warnings; no infinite loop.", "Medium")

        if tool_name.startswith("mock.start"):
            t(DESK, "MC", "mock", "Edge Case", "mock.start with port already in use",
              "Port occupied.", "1. Call with port=X (in use).",
              "tool=mock.start",
              "isError or auto-pick next free port per design; documented behavior.", "Medium")
            t(DESK, "MC", "mock", "Functional", "mock.start without port (auto)",
              "Mock defined.", "1. Call without port.",
              "tool=mock.start",
              "Picks free port; response includes assigned port/pid/startedAt.", "High")

        if tool_name == "mock.stop":
            t(DESK, "MC", "mock", "Edge Case", "mock.stop when not running",
              "Mock not running.", "1. Call.",
              "tool=mock.stop",
              "Either ok:true (idempotent) or graceful 'not running' error.", "Low")


gen_mcp_tools_per_tool()


def gen_mcp_clients():
    for client_id, client_name, path_hint in AI_CLIENTS:
        t(DESK, "MC", "Clients", "Functional",
          f"Generate config snippet for {client_name}",
          "Desktop app.", f"1. Settings -> MCP -> select {client_name}; click Copy snippet.",
          f"client={client_id}",
          f"Clipboard has valid JSON for {client_name} format. Path hint: {path_hint}.",
          "High" if client_id in ("claude-desktop", "cursor", "claude-code") else "Medium")
        t(DESK, "MC", "Clients", "Functional",
          f"Path shown for {client_name} matches OS convention",
          "Desktop app.", f"1. Settings -> MCP -> {client_name}; observe path.",
          f"client={client_id}",
          f"Path matches OS convention (macOS/Win/Linux); 'Open in file manager' link works.", "Medium")
        t(DESK, "MC", "Clients", "Compatibility",
          f"End-to-end: paste snippet into {client_name} and verify it loads",
          "Real AI client installed.",
          f"1. Paste snippet.\n2. Restart {client_name}.\n3. Inspect MCP servers / tool list.",
          f"client={client_name}",
          f"{client_name} discovers 50 apicircle tools; calls succeed.",
          "High" if client_id in ("claude-desktop", "cursor") else "Medium")


gen_mcp_clients()


def gen_mcp_security():
    cases = [
        ("Workspace path traversal denied", "Security",
         "1. --workspace '../../etc'.",
         "App refuses with clear error or scopes to a safe canonical path.", "High"),
        ("MCP cannot decrypt secrets without passphrase", "Security",
         "1. Workspace has encrypted vars; MCP boot without passphrase.\n2. Tool reads env var.",
         "Ciphertext remains; tool returns enc:v1:... literal; never returns plaintext.", "High"),
        ("MCP cannot spawn arbitrary shells via tool args", "Security",
         "1. Inject shell metacharacters into name/url args.",
         "Args stored as data; never passed to shell; no command execution.", "High"),
        ("Tool args cannot write outside workspace dir", "Security",
         "1. workspace.write with paths attempting traversal.",
         "Constrained to workspace dir; refused otherwise.", "High"),
        ("history payloads do not leak Authorization", "Security",
         "1. history.get_run on run with Bearer auth.",
         "Authorization header redacted in result (matches main app history view).", "High"),
        ("Stdin DoS - giant frame rejected", "Security",
         "1. Feed multi-GB JSON frame to stdin.",
         "Bounded buffer; server returns parse error or refuses; does not OOM.", "Medium"),
        ("Concurrent tool calls do not race-corrupt state", "Security",
         "1. Send 50 parallel request.create.",
         "All succeed or fail cleanly; workspace JSON consistent (no torn writes).", "High"),
        ("Tool catalog read-only - no dynamic injection", "Security",
         "1. tools/list before and after various calls.",
         "Catalog identical; no per-session tool injection.", "Medium"),
        ("No network access from MCP server itself", "Security",
         "1. Block outbound; run codegen/import tools.",
         "Local-only tools work; only request/plan run needs network (and fails clearly).", "Medium"),
        ("Workspace file integrity after crash", "Security",
         "1. kill -9 mid-write of workspace.write.",
         "synced.json valid (atomic write) or remains at previous state; never partial.", "High"),
    ]
    for sub, ttype, steps, expected, prio in cases:
        t(DESK, "MC", "Security", ttype, f"MCP security: {sub}",
          "MCP running.", steps, "mcp security", expected, prio)


gen_mcp_security()


def gen_mcp_vault_passphrase():
    cases = [
        ("Workspace passphrase unlocked via desktop unlocks MCP child", "Functional",
         "Desktop app unlocked.",
         "1. Desktop spawns MCP subprocess; pass passphrase via IPC handshake.",
         "MCP can decrypt secrets for that session.", "High"),
        ("Standalone CLI MCP without passphrase keeps ciphertext", "Negative",
         "Workspace has secrets.",
         "1. Run apicircle-mcp standalone.\n2. environment.read.",
         "Encrypted values returned as enc:v1:... ciphertext.", "High"),
        ("Passphrase changed in desktop while MCP running", "Edge Case",
         "MCP running.",
         "1. Change passphrase.",
         "MCP either uses cached key (existing session) or fails on next decrypt; documented.", "Medium"),
        ("Plain vars accessible to MCP without passphrase", "Functional",
         "Workspace has plaintext vars.",
         "1. environment.read.",
         "Plain values returned; secrets stay encrypted.", "High"),
        ("Importing encrypted var via environment.import preserves cipher", "Edge Case",
         "Export from one workspace.",
         "1. environment.import.",
         "Cipher imported intact; new workspace needs source passphrase to decrypt.", "Medium"),
    ]
    for sub, ttype, pre, steps, expected, prio in cases:
        t(DESK, "MC", "Vault", ttype, f"MCP vault: {sub}",
          pre, steps, "mcp vault", expected, prio)


gen_mcp_vault_passphrase()


def gen_mcp_performance():
    cases = [
        ("workspace.read on 10K-request workspace", "Performance",
         "1. tools/call workspace.read.",
         "Response within ~2s; JSON valid; client receives full doc.", "Medium"),
        ("history.list_runs with 50K runs and limit=500", "Performance",
         "1. Call.",
         "Returns 500 newest in <1s.", "Medium"),
        ("import.openapi on 5MB spec", "Performance",
         "1. Call.",
         "Completes in reasonable time; warnings array populated.", "Medium"),
        ("100 sequential request.create", "Performance",
         "1. Loop create.",
         "All succeed; workspace.json grows correctly.", "Medium"),
        ("plan.run with 50 steps", "Performance",
         "1. Run.",
         "Results array has 50 entries; aggregate timing reasonable.", "Medium"),
        ("Long-running session (1h, 1000 calls)", "Performance",
         "1. Drive load.",
         "Memory stays bounded; no FD leaks; logs do not grow unbounded.", "Medium"),
        ("Concurrent 20 mock.start", "Edge Case",
         "1. Start 20 mocks.",
         "Each gets distinct port; pids tracked.", "Low"),
    ]
    for sub, ttype, steps, expected, prio in cases:
        t(DESK, "MC", "Performance", ttype, f"MCP performance: {sub}",
          "MCP running.", steps, "mcp perf", expected, prio)


gen_mcp_performance()


def gen_mcp_changedids_consistency():
    """Every mutating tool must return changedIds so the desktop UI can refresh."""
    mutating = [n for n, _, _, _ in MCP_TOOL_CATALOG
                if any(n.endswith(s) for s in (".create", ".update", ".delete",
                                                 ".set_active", ".set_priority",
                                                 ".add_step", ".remove_step",
                                                 ".reorder_steps", ".set_variables",
                                                 ".import", ".write",
                                                 ".add_endpoint", ".update_endpoint",
                                                 ".delete_endpoint",
                                                 ".set_validation_rules",
                                                 ".set_response_rules",
                                                 ".set_multipliers"))
                or n in ("plan.run", "mock.start", "mock.stop")]
    for tool in mutating:
        t(DESK, "MC", "changedIds", "Regression",
          f"{tool} returns changedIds (or run/start equivalent)",
          "MCP running.",
          f"1. Mutate via {tool}.\n2. Inspect response.",
          f"tool={tool}",
          f"Response contains changedIds (or runId/pid for run/start); desktop UI subscribes and refreshes the affected entities.",
          "High" if tool.startswith(("request.", "environment.", "plan.")) else "Medium")


gen_mcp_changedids_consistency()


# =====================================================================
# OAUTH2 IdP COMPATIBILITY MATRIX (OI)
# =====================================================================
OAUTH_IDPS = [
    ("Auth0", "https://<tenant>.auth0.com", "High"),
    ("Okta", "https://<org>.okta.com", "High"),
    ("Google Identity", "https://accounts.google.com", "High"),
    ("GitHub", "https://github.com/login/oauth", "High"),
    ("Microsoft Identity (Entra)", "https://login.microsoftonline.com", "High"),
    ("AWS Cognito", "https://<pool>.auth.<region>.amazoncognito.com", "Medium"),
    ("Keycloak", "https://<host>/auth/realms/<realm>", "Medium"),
    ("Firebase Auth", "https://securetoken.google.com", "Medium"),
    ("PingFederate", "https://<host>/idp", "Low"),
    ("Generic OIDC", "Generic OpenID Connect provider", "Medium"),
]
OAUTH_GRANTS = ["client_credentials", "authorization_code",
                "authorization_code+PKCE", "password",
                "device_code", "refresh_token", "implicit"]


def gen_oauth_idp_matrix():
    for idp, descr, base_prio in OAUTH_IDPS:
        for grant in OAUTH_GRANTS:
            sub = f"{idp} / {grant}"
            title = f"OAuth2 against {idp} using {grant}"
            steps = (f"1. Configure OAuth2: authUrl/tokenUrl per {idp}; clientId/secret/scope.\n"
                     f"2. Initiate {grant}.\n"
                     f"3. Approve in IdP UI (or supply credentials for non-interactive grants).\n"
                     f"4. Verify access_token + (optional refresh_token) stored.\n"
                     f"5. Send a downstream protected request.")
            expected = (
                f"Token acquired against {idp} successfully; obtainedScope reflects "
                f"granted scope; downstream request returns 200; refresh path (if returned) "
                f"works on next expiry."
            )
            if grant == "implicit" and idp in ("Google Identity", "GitHub"):
                expected += " (Note: many providers have deprecated implicit; if not supported, IdP returns clear error.)"
                prio = "Low"
            elif grant == "device_code":
                expected += " Device flow user_code displayed; polling respected; activation works."
                prio = "Medium"
            else:
                prio = base_prio
            t(BOTH, "OI", idp, "Compatibility", title,
              "OAuth2 credentials for the IdP.", steps, f"idp={idp} grant={grant}",
              expected, prio)


gen_oauth_idp_matrix()


# =====================================================================
# NETWORK SECURITY / TLS (NS) and PROXY (PR)
# =====================================================================
def gen_network_security_matrix():
    tls = [
        ("Valid HTTPS cert", "Functional",
         "1. Send to https endpoint with valid cert.",
         "200; lock badge shown if any.", "High"),
        ("Self-signed cert", "Negative",
         "1. Send to self-signed.",
         "Error with cert details; allow-once option (if implemented).", "High"),
        ("Expired cert", "Negative",
         "1. Send.",
         "Clear error 'certificate expired'.", "High"),
        ("Hostname mismatch (SNI)", "Negative",
         "1. Send to cert valid for other host.",
         "Error 'hostname mismatch'.", "High"),
        ("Revoked cert (OCSP/CRL)", "Edge Case",
         "1. Send.",
         "If revocation checked, error; otherwise documented behavior.", "Medium"),
        ("TLS 1.0 endpoint (deprecated)", "Negative",
         "1. Send.",
         "Either upgraded to 1.2+ or refused; clear message.", "Medium"),
        ("TLS 1.2 endpoint", "Functional", "1. Send.", "200.", "High"),
        ("TLS 1.3 endpoint", "Functional", "1. Send.", "200.", "High"),
        ("Custom CA cert uploaded", "Functional",
         "1. Upload custom CA; send to internal endpoint.",
         "Cert chain validates against uploaded CA; 200.", "Medium"),
        ("Per-request 'allow insecure'", "Edge Case",
         "1. Enable insecure flag; send to self-signed.",
         "Bypass with explicit user opt-in; visual warning shown.", "Medium"),
        ("mTLS with client cert", "Functional",
         "1. Configure mTLS cert; send to mTLS endpoint.",
         "Client cert presented; 200.", "Medium"),
        ("mTLS missing client cert", "Negative",
         "1. mTLS required; no cert.",
         "Server returns 400/403; clear error.", "Medium"),
        ("Cipher suite restrictions", "Edge Case",
         "1. Server only supports older ciphers.",
         "App connects or refuses per OpenSSL defaults; documented.", "Low"),
        ("Certificate transparency logs (CT)", "Compatibility",
         "1. Send to CT-mandated endpoint.",
         "Connection works if CT honored.", "Low"),
        ("DNS-over-HTTPS (DoH)", "Compatibility",
         "1. If app supports DoH.", "Resolves; otherwise OS DNS.", "Low"),
        ("IPv6 endpoint", "Compatibility",
         "1. Send to https://[::1]:8443/.", "Works.", "Medium"),
        ("Endpoint behind Cloudflare / WAF", "Functional",
         "1. Send.", "Works; potential 403 if WAF triggers - clear feedback.", "Medium"),
        ("Endpoint with HSTS", "Functional",
         "1. Send http after HSTS pin.",
         "Upgraded to https or honors HSTS per browser/desktop semantics.", "Low"),
    ]
    for sub, ttype, steps, expected, prio in tls:
        t(BOTH, "NS", "TLS", ttype, f"TLS: {sub}", "-", steps, sub, expected, prio)


gen_network_security_matrix()


def gen_proxy_matrix():
    proxies = [
        ("System proxy (OS-level)", "Functional",
         "1. Set OS proxy; send.", "Honored.", "Medium"),
        ("Manual HTTP proxy", "Functional",
         "1. Configure proxy URL.\n2. Send.", "Outgoing via proxy.", "Medium"),
        ("Manual SOCKS5 proxy", "Functional",
         "1. SOCKS5 URL.\n2. Send.", "Works (desktop typically; browser limits).", "Medium"),
        ("HTTPS proxy", "Functional",
         "1. Proxy URL https://.\n2. Send.", "CONNECT tunneling; works.", "Medium"),
        ("Proxy with Basic auth", "Functional",
         "1. user:pass@proxy.\n2. Send.", "407 challenge handled.", "Medium"),
        ("Proxy with NTLM auth", "Functional",
         "1. NTLM proxy.\n2. Send.", "Handshake completes (desktop).", "Low"),
        ("PAC file URL", "Compatibility",
         "1. Provide PAC URL.\n2. Send.", "If supported, PAC evaluates per-host; documented if not.", "Low"),
        ("Proxy bypass list", "Functional",
         "1. Configure bypass for *.local.\n2. Send to api.local and api.example.com.",
         "api.local bypasses proxy; example.com uses proxy.", "Low"),
        ("Per-request proxy override", "Functional",
         "1. Request-level proxy.\n2. Send.", "Overrides global proxy.", "Low"),
        ("Proxy unreachable", "Negative",
         "1. Invalid proxy host.\n2. Send.", "Clear error; not silent timeout.", "Medium"),
        ("Proxy refuses connection", "Negative",
         "1. Proxy down.\n2. Send.", "Clear error.", "Medium"),
        ("Captive portal interception", "Edge Case",
         "1. Send through captive portal.",
         "Detects HTML interception; warns user.", "Low"),
        ("Network change to/from proxy", "Edge Case",
         "1. Toggle network.", "Reloads proxy config or warns.", "Low"),
    ]
    for sub, ttype, steps, expected, prio in proxies:
        t(BOTH, "PR", "Proxy", ttype, f"Proxy: {sub}", "-", steps, sub, expected, prio)


gen_proxy_matrix()


# =====================================================================
# HTTP PROTOCOL VERSIONS (HV)
# =====================================================================
def gen_http_version_matrix():
    cases = [
        ("HTTP/1.0", "Functional", "Endpoint serves HTTP/1.0", "1. Send.", "Connection: close per spec; 200.", "Low"),
        ("HTTP/1.1 keepalive", "Functional", "Default", "1. Send 2 requests.", "Same connection reused.", "Medium"),
        ("HTTP/1.1 Transfer-Encoding chunked", "Functional", "Chunked resp",
         "1. Send.", "Chunks reassembled; size reported.", "Medium"),
        ("HTTP/2 multiplexed", "Functional", "h2 endpoint",
         "1. Send 5 in parallel.", "Multiplexed over one connection; each completes.", "High"),
        ("HTTP/2 header compression (HPACK)", "Edge Case", "-",
         "1. Inspect wire.", "Headers compressed; logically equivalent.", "Low"),
        ("HTTP/2 server push (if used)", "Edge Case", "-",
         "1. Server pushes resource.", "Documented behavior; not silently dropped.", "Low"),
        ("HTTP/3 (QUIC)", "Compatibility", "h3 endpoint",
         "1. Send.", "If supported, works over QUIC; else falls back to h2.", "Medium"),
        ("Alt-Svc upgrade", "Compatibility", "Server advertises h3",
         "1. Subsequent send.", "Upgraded if supported.", "Low"),
        ("HTTP/1.1 100-Continue", "Edge Case", "Large body upload",
         "1. POST with Expect: 100-continue.", "Client waits for 100 before body.", "Medium"),
        ("Connection close mid-response", "Negative", "Server resets",
         "1. Send.", "Clear error; partial body marked truncated.", "Medium"),
        ("HTTP/1.1 pipelining (if supported)", "Edge Case", "-",
         "1. Send pipelined.", "Either supported or sequential; consistent.", "Low"),
        ("Keep-Alive with idle timeout", "Edge Case", "Endpoint with 5s idle",
         "1. Send; wait; send again.", "New connection if idle exceeded.", "Low"),
        ("Response without Content-Length and without chunked", "Edge Case",
         "Server uses connection-close framing.", "1. Send.", "Body read until close.", "Low"),
        ("Trailer headers (HTTP/1.1)", "Edge Case", "Server uses trailers",
         "1. Send.", "Trailers shown alongside headers.", "Low"),
    ]
    for sub, ttype, pre, steps, expected, prio in cases:
        t(BOTH, "HV", sub, ttype, f"HTTP version: {sub}", pre, steps, sub, expected, prio)


gen_http_version_matrix()


# =====================================================================
# COMPRESSION & CONTENT ENCODING (CE)
# =====================================================================
def gen_compression_matrix():
    encodings = ["gzip", "deflate", "br (Brotli)", "zstd",
                 "identity (none)", "gzip+chained"]
    for enc in encodings:
        for direction in ("Request", "Response"):
            sub = f"{direction} - {enc}"
            steps = (
                f"1. {'Set Content-Encoding ' + enc + ' on request' if direction == 'Request' else 'Send to endpoint returning Content-Encoding: ' + enc}.\n"
                f"2. Inspect viewer/body."
            )
            expected = (
                f"{'Body compressed correctly on send' if direction == 'Request' else 'Body decompressed transparently; preview shows decoded content'}; "
                f"Content-Length reflects {'compressed' if direction == 'Request' else 'decompressed when known'} size."
            )
            ttype = "Functional"
            prio = "High" if enc in ("gzip", "identity (none)") else "Medium"
            t(BOTH, "CE", enc, ttype, f"Compression: {sub}",
              "-", steps, f"enc={enc} dir={direction}", expected, prio)
    # Accept-Encoding negotiation
    accepts = [
        ("Accept-Encoding: gzip", "Server picks gzip."),
        ("Accept-Encoding: gzip, br;q=0.8", "Quality-weighted negotiation."),
        ("Accept-Encoding: *", "Any encoding accepted."),
        ("No Accept-Encoding header", "Server returns identity."),
    ]
    for ae, expected_core in accepts:
        steps = f"1. Set {ae}.\n2. Send."
        expected = expected_core + " App decodes correctly."
        t(BOTH, "CE", "Negotiation", "Functional", f"Accept-Encoding: {ae}",
          "-", steps, ae, expected, "Medium")
    # Failures
    t(BOTH, "CE", "Failure", "Negative", "Corrupt gzip response",
      "-", "1. Server sends corrupt gzip.", "corrupt",
      "Decompression error surfaced clearly; raw bytes available for inspection.", "Medium")
    t(BOTH, "CE", "Failure", "Negative", "Unsupported encoding from server",
      "-", "1. Server returns Content-Encoding: xyz.", "unsupported",
      "Clear error; raw bytes shown.", "Low")


gen_compression_matrix()


# =====================================================================
# SCHEMA MIGRATION & VERSIONING (SM)
# =====================================================================
def gen_schema_migration_matrix():
    cases = [
        ("Open v1 workspace.json in current app", "Compatibility",
         "1. Pull a workspace saved by an older app version.",
         "App migrates forward in-memory; subsequent push writes current schema; no data loss.", "High"),
        ("Open v2 workspace.json", "Compatibility", "1. Pull.", "Migrates; data preserved.", "High"),
        ("Open future-schema workspace.json", "Negative",
         "1. Pull workspace marked with future schema version.",
         "App refuses with clear 'please update' message; no partial load.", "High"),
        ("Migration adds new optional fields", "Functional",
         "1. Old workspace missing newer fields.",
         "Defaults applied silently; UI shows new sections empty.", "Medium"),
        ("Migration renames a field", "Functional",
         "1. Old field name still present.",
         "Renamed on load; persisted under new name on next push.", "Medium"),
        ("Migration removes a deprecated field", "Functional",
         "1. Old field present.",
         "Dropped on load; warning logged once.", "Medium"),
        ("Migration is reversible (or one-way clearly stated)", "Regression",
         "1. Inspect docs/code.", "Migration direction documented; downgrade path noted.", "Medium"),
        ("Two devices on different app versions", "Edge Case",
         "1. Older device pulls newer schema; pushes back.",
         "Older device refuses or warns; never silently downgrades.", "High"),
        ("CLI MCP version mismatch with desktop", "Edge Case",
         "1. Newer mcp binary against older workspace, or vice versa.",
         "Version negotiated; clear error if incompatible.", "Medium"),
        ("Telemetry: migration event tracked once per upgrade", "Regression",
         "1. First migration boot.",
         "One event; subsequent boots silent.", "Low"),
        ("Encrypted secrets survive migration", "Security",
         "1. Old schema with secrets; migrate.",
         "Cipher format preserved; passphrase unlock still works.", "High"),
        ("Linked workspace ledger migration", "Edge Case",
         "1. Old release ledger shape.",
         "Migrated; version pinning preserved.", "Medium"),
    ]
    for sub, ttype, steps, expected, prio in cases:
        t(BOTH, "SM", sub, ttype, f"Schema migration: {sub}", "-", steps, sub, expected, prio)


gen_schema_migration_matrix()


# =====================================================================
# MULTI-USER CONCURRENCY (MU)
# =====================================================================
def gen_multi_user_matrix():
    scenarios = [
        ("Two devs add request to same folder simultaneously", "Edge Case",
         "1. Both push.", "Both adds merge; no id collision.", "High"),
        ("Two devs edit different fields of same request", "Functional",
         "1. Both push.", "Auto-merged; both edits applied.", "High"),
        ("Two devs edit same field of same request", "Negative",
         "1. Both push.", "Conflict modal on second pull; explicit resolution.", "High"),
        ("Dev A deletes; dev B edits", "Negative",
         "1. A deletes; B pushes edit.", "Conflict; B chooses to resurrect or accept delete.", "High"),
        ("Three devs in succession", "Edge Case",
         "1. A pushes; B pulls and pushes; C pulls and pushes.",
         "Each pulls latest; all changes preserved.", "Medium"),
        ("Dev forgets to pull and pushes after main moved", "Negative",
         "1. Push.", "Rejected; prompted to pull-merge-push.", "High"),
        ("PR review on GitHub (out-of-band edit)", "Edge Case",
         "1. Reviewer edits via GitHub web.\n2. Dev pulls.",
         "Reviewer's commit pulled cleanly.", "Medium"),
        ("Branch protection requires linear history", "Compatibility",
         "1. App attempts merge commit; protection blocks.",
         "App falls back to rebase or surfaces clear error.", "Medium"),
        ("Two devs on different branches, then merge", "Functional",
         "1. Independent branches; merge.", "Cross-branch merge clean.", "Medium"),
        ("Force-push by another dev", "Edge Case",
         "1. Co-dev force-pushes; user pulls.",
         "History rewrite detected; warning; reset option.", "Medium"),
        ("Pull while another push in progress on remote", "Edge Case",
         "1. Race the operations.", "Last-writer wins on remote; user pulls latest.", "Medium"),
        ("Linked workspace owned by another team updates", "Edge Case",
         "1. Source linked WS published v2.",
         "Consumer banner 'Update available'; opt-in adoption.", "High"),
        ("Encrypted vars on multi-user repo", "Security",
         "1. Two users share passphrase.",
         "Both can decrypt; commits keep ciphertext only; safe.", "High"),
        ("Encrypted vars unlocked on one device, locked on other", "Security",
         "1. User without passphrase pulls.",
         "Cipher visible; UI flags as locked; no plaintext leak.", "High"),
        ("Different secret slots per user (named)", "Edge Case",
         "1. User A creates secret 'apiToken' slot A.\n2. User B creates own slot.",
         "Both coexist; each user uses own slot.", "Medium"),
    ]
    for sub, ttype, steps, expected, prio in scenarios:
        t(BOTH, "MU", sub, ttype, f"Multi-user: {sub}", "-", steps, sub, expected, prio)


gen_multi_user_matrix()


# =====================================================================
# LOCALE & I18N (LO)
# =====================================================================
def gen_locale_matrix():
    locales = [
        ("en-US", "12h time, MM/DD/YYYY"),
        ("en-GB", "24h time, DD/MM/YYYY"),
        ("de-DE", "DD.MM.YYYY"),
        ("ja-JP", "YYYY/MM/DD"),
        ("fr-FR", "DD/MM/YYYY"),
        ("es-ES", "DD/MM/YYYY"),
        ("zh-CN", "YYYY/MM/DD"),
        ("ar-EG", "RTL"),
        ("he-IL", "RTL"),
        ("pt-BR", "DD/MM/YYYY"),
        ("ko-KR", "YYYY.MM.DD"),
        ("ru-RU", "DD.MM.YYYY"),
    ]
    for loc, fmt in locales:
        t(BOTH, "LO", loc, "Compatibility", f"Locale {loc} - date/time format",
          f"OS locale {loc}.",
          "1. Send a request; inspect history timestamp; inspect date displays.",
          f"locale={loc}",
          f"Dates formatted per locale ({fmt}); times respect 12/24h; relative phrases ('2m ago') localized if i18n supported.",
          "Medium")
        t(BOTH, "LO", loc, "Compatibility", f"Locale {loc} - number format",
          f"OS locale {loc}.",
          "1. Inspect numeric fields (response time, size, counts).",
          f"locale={loc}",
          f"Numbers respect locale grouping/decimal separators (e.g., 1.234,56 vs 1,234.56).",
          "Low")
        if "ar" in loc or "he" in loc:
            t(BOTH, "LO", loc, "Compatibility", f"Locale {loc} - RTL layout",
              "OS RTL locale.",
              "1. Open app; inspect direction of panels, text, scrollbars.",
              f"locale={loc}",
              "Layout mirrored; text right-aligned; icons reflect direction; keyboard nav still logical.",
              "Medium")
    # Generic i18n
    t(BOTH, "LO", "Strings", "Compatibility", "All visible UI strings translatable",
      "i18n manifest exists or planned.", "1. Audit hardcoded strings.",
      "i18n",
      "No hardcoded English in components; all keys flow through translation lookup.", "Low")
    t(BOTH, "LO", "Strings", "Compatibility", "Missing translation falls back to English",
      "Locale missing keys.", "1. Switch locale; load.",
      "missing-keys",
      "Falls back; UI does not show key names; warning logged.", "Low")
    t(BOTH, "LO", "Strings", "Edge Case", "Long translation strings (German)",
      "de-DE.", "1. Inspect button labels.",
      "long-text",
      "Layout adapts; no clipped text; tooltips for truncation.", "Low")


gen_locale_matrix()


# =====================================================================
# OS / PLATFORM COMPATIBILITY (OP)
# =====================================================================
def gen_os_platform_matrix():
    desktops = [
        ("macOS 14", "Sonoma"), ("macOS 15", "Sequoia"),
        ("Windows 10", "21H2+"), ("Windows 11", "23H2"),
        ("Ubuntu 22.04", "Jammy"), ("Ubuntu 24.04", "Noble"),
        ("Fedora 40", ""), ("Arch", "rolling"),
    ]
    for os_name, ver in desktops:
        t(DESK, "OP", os_name, "Compatibility", f"Smoke on {os_name} {ver}",
          "Fresh install.", "1. Install; launch.\n2. Run smoke (create-send-push).",
          f"os={os_name}",
          "App installs; launches; smoke passes; menus & shortcuts native; auto-update works.",
          "High" if "macOS 15" in os_name or "Windows 11" in os_name or "Ubuntu 24.04" in os_name else "Medium")
        t(DESK, "OP", os_name, "Compatibility", f"Keychain integration on {os_name}",
          "-", "1. Save secret.", f"os={os_name}",
          "OS-native keychain used (Keychain / Credential Manager / libsecret).",
          "High" if "macOS" in os_name or "Windows 11" in os_name else "Medium")
        t(DESK, "OP", os_name, "Compatibility", f"Window state on {os_name}",
          "-", "1. Resize/move; restart.", f"os={os_name}",
          "Bounds restored; multi-monitor handled.", "Medium")
    # Web browsers x OS
    browsers = [
        ("Chrome stable", "Mac/Win/Linux"), ("Chrome canary", "Mac/Win"),
        ("Firefox stable", "Mac/Win/Linux"), ("Firefox ESR", "Win"),
        ("Safari", "macOS"), ("Edge stable", "Mac/Win"),
    ]
    for br, os_targets in browsers:
        for os_t in os_targets.split("/"):
            t(WEB, "OP", br, "Compatibility", f"Smoke {br} on {os_t}",
              "Fresh.", "1. Smoke flow.", f"browser={br} os={os_t}",
              "Works; OAuth popup works; persistence stable.",
              "High" if br in ("Chrome stable", "Firefox stable", "Safari", "Edge stable") else "Low")
    # ARM vs x86
    archs = [("Apple Silicon (arm64)", "macOS"), ("Intel (x86_64)", "macOS"),
             ("x86_64", "Windows"), ("ARM64", "Windows"),
             ("x86_64", "Linux"), ("arm64", "Linux")]
    for arch, os_t in archs:
        t(DESK, "OP", "Architecture", "Compatibility",
          f"{arch} on {os_t}",
          "-", "1. Install native build; smoke.", f"arch={arch} os={os_t}",
          "Binary correct arch; native modules load.", "Medium")


gen_os_platform_matrix()


# =====================================================================
# LINKED WORKSPACE VERSIONING (LV)
# =====================================================================
def gen_linked_versioning_matrix():
    cases = [
        ("Link to latest version", "Functional",
         "1. Link without pinning.", "Tracks latest published release.", "High"),
        ("Pin to specific version", "Functional",
         "1. Pin v1.2.0.", "Stays on v1.2.0 even when v1.3.0 published.", "High"),
        ("Update banner when source publishes new version", "Functional",
         "1. Source publishes v2.", "Banner 'v2 available'; preview diff.", "High"),
        ("Adopt new version", "Functional",
         "1. Click adopt; review changes.",
         "Local workspace updates references; existing overrides preserved when keys still exist.",
         "High"),
        ("Decline new version (stay pinned)", "Functional",
         "1. Decline.", "No change.", "Medium"),
        ("Breaking change in new version (removed env var)", "Negative",
         "1. Source removes var.", "Banner flags breaking; consumer warned which overrides break.",
         "High"),
        ("Renamed entity in new version", "Edge Case",
         "1. Source renames request.", "Override migrates if rename detected, else flagged.",
         "Medium"),
        ("Multiple linked workspaces with conflicting var names", "Edge Case",
         "1. Two linked WS define 'token'.", "Resolution follows priority order; UI shows source per var.",
         "Medium"),
        ("Override per linked-version", "Functional",
         "1. Pin v1 with overrides; switch to v2.",
         "Either v2-scoped overrides or warning to re-apply; documented.",
         "Medium"),
        ("Unlink preserves local copies (optional)", "Edge Case",
         "1. Unlink with 'keep local copies' on.",
         "Linked resources duplicated into workspace as native entities.",
         "Medium"),
        ("Source unpublished a version we pinned", "Negative",
         "1. Source unpublishes v1.2.0.",
         "Consumer warned; current data still usable; cannot re-fetch.", "Medium"),
        ("Linked release ledger refresh", "Functional",
         "1. Refresh.", "Ledger cache updates; new version visible.", "Medium"),
        ("Compare diff between linked versions", "UX/UI",
         "1. Open version diff.", "Side-by-side or change list.", "Medium"),
        ("Release notes Markdown rendered", "UX/UI",
         "1. View release notes.", "Markdown safely rendered.", "Low"),
        ("Linked WS that itself links to another WS (chain)", "Edge Case",
         "1. A -> B -> C.", "Transitive references resolved or blocked at depth N.",
         "Low"),
    ]
    for sub, ttype, steps, expected, prio in cases:
        t(BOTH, "LV", sub, ttype, f"Linked versioning: {sub}", "-", steps, sub, expected, prio)


gen_linked_versioning_matrix()


# =====================================================================
# WEBSOCKET / SSE / STREAMING (WK)
# =====================================================================
def gen_streaming_matrix():
    cases = [
        ("WS connect to ws://", "Functional",
         "1. New WS request; ws://echo.\n2. Connect.",
         "Status Connected; ping/pong; chat panel ready.", "Medium"),
        ("WS connect to wss://", "Functional",
         "1. wss endpoint.", "Connected.", "High"),
        ("WS send text frame", "Functional",
         "1. Send 'hello'.", "Echo received.", "High"),
        ("WS send binary frame", "Functional",
         "1. Send binary.", "Echo received as binary.", "Medium"),
        ("WS receive server-pushed frames", "Functional",
         "1. Subscribe.", "Frames stream into panel.", "Medium"),
        ("WS subprotocol negotiation", "Edge Case",
         "1. Sec-WebSocket-Protocol on connect.",
         "Server-chosen subprotocol shown.", "Low"),
        ("WS auth via header on upgrade", "Functional",
         "1. Authorization header.", "Server accepts.", "Medium"),
        ("WS auth via cookie", "Functional",
         "1. Cookie sent on upgrade.", "Auth works.", "Medium"),
        ("WS auto-reconnect on drop", "Edge Case",
         "1. Drop connection.", "If supported, reconnects with backoff.", "Low"),
        ("WS close 1000 vs 1006", "Negative",
         "1. Compare normal vs abnormal close.", "Reason shown.", "Low"),
        ("WS large frame (1MB)", "Performance",
         "1. Send.", "Works; buffered correctly.", "Low"),
        ("WS history persisted", "Functional",
         "1. Send messages; reload.", "Per-session history visible.", "Medium"),
        ("SSE connect", "Functional",
         "1. text/event-stream endpoint.", "Stream open; events arrive.", "Medium"),
        ("SSE event with id and reconnect", "Functional",
         "1. Drop; reconnect.", "Last-Event-Id sent on reconnect.", "Medium"),
        ("SSE multiline data", "Edge Case",
         "1. Server sends multi-data lines.", "Reassembled per spec.", "Low"),
        ("gRPC unary call (if supported)", "Functional",
         "1. .proto loaded; unary call.", "Response decoded.", "Medium"),
        ("gRPC server streaming", "Functional",
         "1. Server stream.", "Frames stream into UI.", "Medium"),
        ("gRPC client streaming", "Functional",
         "1. Send multiple frames.", "Server final response.", "Medium"),
        ("gRPC bidirectional", "Functional",
         "1. Bidi.", "Both directions stream.", "Medium"),
        ("gRPC reflection", "Functional",
         "1. Server with reflection enabled.", "Method list discovered.", "Low"),
        ("gRPC over TLS", "Security",
         "1. TLS endpoint.", "Handshake OK; calls succeed.", "Medium"),
    ]
    for sub, ttype, steps, expected, prio in cases:
        t(BOTH, "WK", sub, ttype, f"Streaming: {sub}", "-", steps, sub, expected, prio)


gen_streaming_matrix()


# =====================================================================
# CACHING / ETAG / CONDITIONAL (CA)
# =====================================================================
def gen_caching_matrix():
    cases = [
        ("If-None-Match with stored ETag", "Functional",
         "1. First send stores ETag.\n2. Send again with If-None-Match.",
         "304 Not Modified or 200 with new ETag.", "Medium"),
        ("If-Modified-Since", "Functional",
         "1. Send.", "Server returns 304 if not modified.", "Medium"),
        ("Cache-Control: no-store respected", "Functional",
         "1. Inspect history.", "Response not cached or marked accordingly.", "Low"),
        ("Cache-Control: private vs public", "Edge Case",
         "1. Inspect.", "Documented behavior.", "Low"),
        ("Vary header tracked", "Edge Case",
         "1. Send with different Accept.", "Cached entries keyed by Vary.", "Low"),
        ("ETag changes on body change", "Regression",
         "1. Re-send; compare.", "ETag rotates appropriately.", "Low"),
        ("If-Match for optimistic concurrency", "Functional",
         "1. PUT with If-Match.", "412 Precondition Failed handled.", "Medium"),
        ("Stale-while-revalidate (if respected)", "Edge Case",
         "1. Stale entry.", "Used while revalidating per RFC 5861.", "Low"),
        ("Expires header old (HTTP/1.0)", "Edge Case",
         "1. Send.", "Honored.", "Low"),
        ("304 with no body shown correctly", "UX/UI",
         "1. 304 response.", "Status shown; body section explains.", "Low"),
        ("Disable cache per request", "Functional",
         "1. Toggle 'fresh-only'.", "If-None-Match not sent; force network.", "Medium"),
    ]
    for sub, ttype, steps, expected, prio in cases:
        t(BOTH, "CA", sub, ttype, f"Caching: {sub}", "-", steps, sub, expected, prio)


gen_caching_matrix()


# =====================================================================
# TELEMETRY & PRIVACY (TP)
# =====================================================================
def gen_telemetry_matrix():
    cases = [
        ("Telemetry default off (or first-run consent)", "Security",
         "1. Fresh install.", "Either no events or explicit consent prompt before any event.", "High"),
        ("Disable telemetry from settings", "Functional",
         "1. Toggle off.", "No events sent; setting persists.", "High"),
        ("Event payload contains no PII", "Security",
         "1. Inspect events (DevTools/proxy).",
         "No URLs/bodies/secrets/PII; only counts, error class names, durations.", "High"),
        ("Crash reports opt-in", "Security",
         "1. Trigger crash with reporting off.",
         "No upload; setting respected.", "High"),
        ("Crash reports include stack but no user data", "Security",
         "1. Trigger crash with reporting on.",
         "Stack and version uploaded; no workspace data or secrets.", "High"),
        ("Anonymous install id", "Security",
         "1. Inspect persisted id.",
         "UUID v4 with no derivation from email/machine; can be reset.", "Medium"),
        ("Reset install id", "Functional",
         "1. Settings -> Reset.",
         "New id; old events not retroactively re-attributed.", "Medium"),
        ("Network for telemetry only when enabled", "Security",
         "1. Telemetry off; sniff network.",
         "No telemetry endpoint contacted.", "High"),
        ("Privacy policy link visible", "UX/UI",
         "1. Settings / About.",
         "Link to docs; up-to-date.", "Low"),
        ("Workspace data never sent to telemetry endpoint", "Security",
         "1. Audit code paths.",
         "Code reviewed; never sends workspace JSON to telemetry.", "High"),
    ]
    for sub, ttype, steps, expected, prio in cases:
        t(BOTH, "TP", sub, ttype, f"Telemetry: {sub}", "-", steps, sub, expected, prio)


gen_telemetry_matrix()


# =====================================================================
# BACKUP & RESTORE (BK)
# =====================================================================
def gen_backup_restore_matrix():
    cases = [
        ("Export full workspace JSON (synced)", "Functional",
         "1. Export.", "Valid JSON; secrets excluded (or marked).", "High"),
        ("Export includes attachments (or references)", "Edge Case",
         "1. Workspace has file attachments; export.",
         "Either attachments bundled into archive or references listed.", "Medium"),
        ("Export does not include history (local-only)", "Regression",
         "1. Export; inspect.", "History excluded.", "Medium"),
        ("Re-import exported workspace creates equivalent state", "Functional",
         "1. Import the file.", "Round-trip equal; ids preserved or remapped consistently.", "High"),
        ("Import into existing workspace merges vs overwrites", "Functional",
         "1. Import with both options.", "Both modes documented; merge preserves existing data.", "Medium"),
        ("Backup before destructive op (rest workspace, delete)", "Functional",
         "1. Trigger destructive; observe.",
         "Automatic snapshot created in app-data dir before destructive op.", "High"),
        ("Restore from auto-snapshot", "Functional",
         "1. Settings -> Snapshots.",
         "List of recent snapshots; restore reverts to chosen one.", "Medium"),
        ("Disk full during export", "Negative",
         "1. Simulate.", "Clear error; partial file cleaned up.", "Low"),
        ("Selective restore: only environments", "Functional",
         "1. Restore env section only.", "Other sections untouched.", "Medium"),
        ("Backup encrypted with passphrase", "Security",
         "1. Export with 'encrypt with passphrase'.",
         "Backup is encrypted blob; import requires passphrase.", "High"),
        ("Backup file integrity check (checksum)", "Edge Case",
         "1. Truncate backup; import.",
         "Detected via checksum; import refused.", "Medium"),
        ("Cross-version backup compatibility", "Compatibility",
         "1. Import backup from older app version.",
         "Migrated forward at import; warnings shown.", "Medium"),
    ]
    for sub, ttype, steps, expected, prio in cases:
        t(BOTH, "BK", sub, ttype, f"Backup: {sub}", "-", steps, sub, expected, prio)


gen_backup_restore_matrix()


# =====================================================================
# CODE GENERATION (CG)
# =====================================================================
CODEGEN_LANGS = ["curl", "fetch (JS)", "axios (JS)",
                 "Python requests", "Python httpx", "Go net/http",
                 "Rust reqwest", "Java OkHttp", "C# HttpClient",
                 "PHP curl", "Ruby Net::HTTP", "Node http",
                 "Kotlin OkHttp", "Swift URLSession"]
CODEGEN_INPUTS = [
    ("Simple GET", "Just URL + method"),
    ("POST JSON", "Body + headers"),
    ("POST form-data", "Multipart"),
    ("POST urlencoded", "x-www-form-urlencoded"),
    ("File upload", "Binary"),
    ("GraphQL", "Query + variables"),
    ("Bearer auth", "Authorization header"),
    ("Basic auth", "user:pass"),
    ("API Key header", "Custom header"),
    ("API Key query", "Query param"),
    ("OAuth2 (with token)", "Token attached"),
    ("AWS SigV4", "Signed request"),
    ("Custom headers", "User-set"),
    ("Path params", "URL with {id}"),
    ("Query params with arrays", "?a=1&a=2"),
    ("Cookies", "Cookie header"),
    ("Variable interpolation", "Resolved values"),
]


def gen_codegen_matrix():
    for lang in CODEGEN_LANGS:
        for shape, descr in CODEGEN_INPUTS:
            sub = f"{lang} - {shape}"
            title = f"Codegen {lang}: {shape}"
            steps = (f"1. Set up request shape: {descr}.\n"
                     f"2. Generate code for {lang} (UI Copy-as / MCP generate.code).\n"
                     f"3. Run the generated snippet in a {lang} runtime against the same endpoint.")
            expected = (
                f"Generated code is syntactically valid for {lang}; "
                f"executes against the same endpoint and produces the same response. "
                f"Variables are interpolated to their resolved values (not as literal '{{{{var}}}}'); "
                f"secrets are masked or warning shown when copying."
            )
            prio = "High" if lang in ("curl", "fetch (JS)", "Python requests") and shape in (
                "Simple GET", "POST JSON", "Bearer auth", "OAuth2 (with token)") else (
                "Medium" if lang in ("curl", "fetch (JS)", "axios (JS)",
                                     "Python requests", "Go net/http",
                                     "Java OkHttp", "C# HttpClient") else "Low")
            t(BOTH, "CG", lang, "Functional", title, "-", steps,
              f"lang={lang} shape={shape}", expected, prio)
    # Edge cases
    t(BOTH, "CG", "Edge", "Security", "Codegen redacts secret values when copying to clipboard",
      "Bearer with secret token.", "1. Copy as curl/axios.", "-",
      "Token masked in clipboard preview; full token may be included if user explicitly confirmed.",
      "High")
    t(BOTH, "CG", "Edge", "Edge Case", "Codegen handles binary body via file path placeholder",
      "Binary body.", "1. Generate curl.", "-",
      "curl --data-binary @/path/to/file or language-specific equivalent.", "Medium")
    t(BOTH, "CG", "Edge", "Edge Case", "Codegen for GraphQL preserves query+variables",
      "GraphQL.", "1. Generate.", "-",
      "Body has both fields; correct quoting.", "Medium")
    t(BOTH, "CG", "Edge", "Edge Case", "Codegen for AWS SigV4 includes signing helper or note",
      "AWS request.", "1. Generate.", "-",
      "Either uses SDK call or includes inline signing; not just hardcoded signature.",
      "Medium")


gen_codegen_matrix()


# =====================================================================
# CLI BEYOND MOCK (extends CL)
# =====================================================================
def gen_cli_extended():
    commands = [
        ("apicircle run <plan-id>", "Functional",
         "1. Run a plan from CLI.", "Plan executes; per-step results to stdout; exit 0/1 on pass/fail.",
         "High"),
        ("apicircle run --env Staging", "Functional",
         "1. Override env.", "Staging values used.", "High"),
        ("apicircle run --reporter junit > report.xml", "Functional",
         "1. JUnit reporter.", "Valid JUnit XML; usable in CI.", "Medium"),
        ("apicircle run --reporter json", "Functional",
         "1. JSON reporter.", "Structured output.", "Medium"),
        ("apicircle run --bail", "Functional",
         "1. Stop on first failure.", "Subsequent steps not run.", "Medium"),
        ("apicircle export <workspace> --format postman", "Functional",
         "1. Export.", "Postman v2.1 file produced.", "Medium"),
        ("apicircle export --format openapi", "Functional",
         "1. Export.", "OpenAPI spec produced.", "Medium"),
        ("apicircle import <file> <workspace>", "Functional",
         "1. Import.", "Workspace updated; summary printed.", "Medium"),
        ("apicircle lint <workspace>", "Functional",
         "1. Run lints (if supported).", "Issues printed; exit non-zero on findings.", "Low"),
        ("apicircle --version", "Functional",
         "1. Version flag.", "Version + commit + build info.", "Low"),
        ("apicircle --help", "Functional",
         "1. Help.", "Top-level usage + subcommand list.", "Medium"),
        ("apicircle <subcommand> --help", "Functional",
         "1. Subcommand help.", "Subcommand-specific usage.", "Medium"),
        ("Invalid subcommand", "Negative",
         "1. apicircle bogus.", "Error + suggestions; exit non-zero.", "Low"),
        ("CLI in CI with no TTY", "Compatibility",
         "1. Run non-interactive.", "No prompts; honors flags; exits cleanly.", "High"),
        ("CLI honors NO_COLOR env var", "Compatibility",
         "1. NO_COLOR=1 apicircle run.", "Plain output; no ANSI.", "Low"),
        ("CLI honors HTTP_PROXY env var", "Compatibility",
         "1. HTTP_PROXY=... apicircle run.", "Uses proxy.", "Medium"),
        ("CLI logs structured JSON when --json-logs", "Functional",
         "1. Run with flag.", "Each log line is JSON object.", "Medium"),
        ("CLI exit codes are stable", "Regression",
         "1. Exercise pass/fail/error paths.",
         "0=ok, 1=tests failed, 2=invalid usage, 3=runtime error (stable spec).", "High"),
        ("Globbed workspace path", "Edge Case",
         "1. apicircle run ./ws*.", "Shell-expanded; clear error if multiple.", "Low"),
        ("CLI handles Ctrl+C mid-run", "Functional",
         "1. SIGINT.", "Aborts; partial summary; cleanup; exit non-zero.", "High"),
        ("CLI on Windows PowerShell quoting", "Compatibility",
         "1. Special chars in args.", "Quoted correctly per shell.", "Low"),
        ("CLI on macOS zsh", "Compatibility", "1. Run.", "Works.", "Medium"),
        ("CLI on Linux bash", "Compatibility", "1. Run.", "Works.", "Medium"),
        ("CLI on Windows cmd.exe", "Compatibility", "1. Run.", "Works.", "Low"),
    ]
    for sub, ttype, steps, expected, prio in commands:
        t(DESK, "CL", "Extended", ttype, f"CLI: {sub}", "CLI installed.", steps,
          sub, expected, prio)


gen_cli_extended()


print(f"After full senior expansion: {len(TESTS)} tests")


# =====================================================================
# WORKBOOK BUILD
# =====================================================================

PLATFORM_TITLES = {
    "web": "API Circle Studio - Web App Manual Test Cases",
    "desktop": "API Circle Studio - Desktop App Manual Test Cases",
}

PLATFORM_BLURBS = {
    "web": ("Covers the API Circle Studio web build (apps/web). OAuth2 uses popup + "
            "BroadcastChannel relay. Mock-server runtime is disabled (definitions remain "
            "editable). Data persists in IndexedDB; secrets are encrypted via "
            "passphrase-derived AES-GCM with the master JWK stored wrapped in IndexedDB."),
    "desktop": ("Covers the API Circle Studio desktop build (Electron). OAuth2 uses a "
                "localhost callback HTTP server bound to 127.0.0.1. Mock-server runtime "
                "runs in the main process. Auto-updater enabled. Secrets are wrapped via "
                "the OS keychain (safeStorage). IPC gated by assertTrustedSender. "
                "Window bounds persist."),
}


def build_workbook(platform, out_path):
    relevant = [x for x in TESTS if platform in x[0]]
    wb = Workbook()

    # README
    ws = wb.active
    ws.title = "README"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 120

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value = PLATFORM_TITLES[platform]
    c.font = TITLE_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    ws.row_dimensions[1].height = 30

    readme = [
        ("Platform", platform),
        ("Document owner", "QA Engineering"),
        ("Build under test", "Fill in version before each cycle."),
        ("Test environment", "OS, browser/Electron version, network, vault state."),
        ("Last updated", "2026-05-14"),
        ("Total test cases", str(len(relevant))),
        ("", ""),
        ("Scope", PLATFORM_BLURBS[platform]),
        ("", ""),
        ("Combinatorial coverage included", ""),
        ("HTTP Method x Body Type matrix", "7 methods x 9 body types (every cell)"),
        ("Auth x Method matrix", "17 auth schemes x 4 representative methods"),
        ("Body content variations", "JSON/XML/form-data/binary/urlencoded/GraphQL edge cases"),
        ("Variable interpolation matrix", "15 surfaces x 7 scopes"),
        ("Changes-to-Push view", "21 entity buckets x 3 operations (Added/Modified/Removed) plus per-field modifies"),
        ("Git conflict matrix", "18 buckets x 10 conflict shapes + 3 resolution strategies"),
        ("Mock response matrix", "15 status codes x 8 content types + matching rules + timing"),
        ("Headers deep matrix", "17 standard headers + 20 edge cases"),
        ("JSON Schema references", "$ref depth, composition, circular, external, validation"),
        ("Workspace restore round-trip", "23 data types verified after re-clone"),
        ("History replay matrix", "19 replay scenarios including auth/var/cookie semantics"),
        ("HTTP method edge cases", "GET-with-body, DELETE-with-body, HEAD-with-body, OPTIONS preflight, custom methods"),
        ("", ""),
        ("How to use this workbook", ""),
        ("1.", "Open the 'Test Cases' sheet. Filters are enabled on every column."),
        ("2.", "Pick a module (column B), Test Type (column D), or priority (column L) to scope your session."),
        ("3.", "For each row, follow the steps in 'Test Steps' (column G) using 'Test Data' (column H)."),
        ("4.", "Compare what you see against 'Expected Result' (column I)."),
        ("5.", "Record what you saw in 'Actual Result' (column J)."),
        ("6.", "Set 'Status' (column K) to Pass / Fail / Blocked / Skipped / Not Run."),
        ("7.", "Add 'Tester' (column M), 'Test Date' (column N), and a defect link in 'Notes' (column O)."),
        ("", ""),
        ("Status values", "Pass, Fail, Blocked, Skipped, Not Run"),
        ("Priority values", "High, Medium, Low"),
        ("Test Type values", "Functional, Negative, Edge Case, Security, A11y, Performance, Compatibility, UX/UI, Regression"),
        ("", ""),
        ("Summary dashboard", "The 'Summary' sheet auto-aggregates counts by module, priority, "
                             "and test type via formulas. Do not edit manually."),
        ("", ""),
        ("Prioritized smoke pass", "Filter Priority='High' for a smoke pass (covers OAuth, push/pull, "
                                    "every body type happy path, auth happy paths, JSON schema basics)."),
        ("Defect tracking", "Use column O to link the issue tracker for failed rows."),
    ]

    r = 2
    for label, val in readme:
        ws.cell(row=r, column=1, value=label).font = BOLD_FONT
        ws.cell(row=r, column=1).alignment = WRAP
        ws.cell(row=r, column=2, value=val).font = BASE_FONT
        ws.cell(row=r, column=2).alignment = WRAP
        r += 1

    # Test Cases
    tc_ws = wb.create_sheet("Test Cases")
    headers = [
        "TC ID", "Module", "Sub-Feature", "Test Type", "Test Case Title",
        "Pre-conditions", "Test Steps (How to Execute)", "Test Data",
        "Expected Result", "Actual Result", "Status",
        "Priority", "Tester", "Test Date", "Notes / Defect ID",
    ]
    widths = [14, 32, 32, 14, 42, 30, 50, 22, 56, 30, 12, 10, 14, 12, 28]
    for col_idx, (h, w) in enumerate(zip(headers, widths), 1):
        c = tc_ws.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
        tc_ws.column_dimensions[get_column_letter(col_idx)].width = w

    tc_ws.row_dimensions[1].height = 32
    tc_ws.freeze_panes = "F2"

    counters = {m[0]: 0 for m in MODULES}
    row = 2
    for plats, mod, sub, ttype, title, pre, steps, data, expected, prio in relevant:
        counters[mod] += 1
        tc_id = f"TC-{mod}-{counters[mod]:04d}"
        values = [
            tc_id, MODULE_NAME[mod], sub, ttype, title,
            pre, steps, data, expected,
            "", "Not Run", prio, "", "", "",
        ]
        for col_idx, v in enumerate(values, 1):
            c = tc_ws.cell(row=row, column=col_idx, value=v)
            c.font = BASE_FONT
            c.alignment = WRAP
            c.border = BORDER
        tc_ws.row_dimensions[row].height = 72
        row += 1
    last_row = row - 1

    # Data validations
    status_dv = DataValidation(
        type="list", formula1='"Pass,Fail,Blocked,Skipped,Not Run"',
        allow_blank=True, errorTitle="Invalid status",
        error="Choose one of: Pass, Fail, Blocked, Skipped, Not Run",
    )
    tc_ws.add_data_validation(status_dv)
    status_dv.add(f"K2:K{last_row}")

    prio_dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
    tc_ws.add_data_validation(prio_dv)
    prio_dv.add(f"L2:L{last_row}")

    type_dv = DataValidation(
        type="list",
        formula1='"Functional,Negative,Edge Case,Security,A11y,Performance,Compatibility,UX/UI,Regression"',
        allow_blank=True,
    )
    tc_ws.add_data_validation(type_dv)
    type_dv.add(f"D2:D{last_row}")

    # Conditional formatting
    status_range = f"K2:K{last_row}"
    for val, fill in [
        ("Pass", PASS_FILL), ("Fail", FAIL_FILL), ("Blocked", BLOCKED_FILL),
        ("Not Run", NOT_RUN_FILL), ("Skipped", SKIPPED_FILL),
    ]:
        tc_ws.conditional_formatting.add(
            status_range, CellIsRule(operator="equal", formula=[f'"{val}"'], fill=fill))

    prio_range = f"L2:L{last_row}"
    for val, fill in [
        ("High", PRIORITY_HIGH_FILL),
        ("Medium", PRIORITY_MED_FILL),
        ("Low", PRIORITY_LOW_FILL),
    ]:
        tc_ws.conditional_formatting.add(
            prio_range, CellIsRule(operator="equal", formula=[f'"{val}"'], fill=fill))

    tc_ws.auto_filter.ref = f"A1:O{last_row}"

    # Summary
    sm = wb.create_sheet("Summary", 1)
    sm.column_dimensions["A"].width = 40
    for ch in "BCDEFGH":
        sm.column_dimensions[ch].width = 13

    sm.merge_cells("A1:H1")
    c = sm["A1"]
    c.value = f"Test Execution Summary - {platform.title()}"
    c.font = TITLE_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    sm.row_dimensions[1].height = 28

    sm_headers = ["Module", "Total", "Not Run", "Pass", "Fail", "Blocked", "Skipped", "Pass %"]
    for i, h in enumerate(sm_headers, 1):
        c = sm.cell(row=3, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

    present_modules = sorted(
        {x[1] for x in relevant},
        key=lambda m: [code for code, _ in MODULES].index(m),
    )

    r = 4
    for mod_code in present_modules:
        mod_name = MODULE_NAME[mod_code]
        sm.cell(row=r, column=1, value=mod_name).font = BOLD_FONT
        sm.cell(row=r, column=1).alignment = WRAP
        sm.cell(row=r, column=1).border = BORDER
        sm.cell(row=r, column=2, value=f'=COUNTIF(\'Test Cases\'!B:B,"{mod_name}")')
        sm.cell(row=r, column=3, value=f'=COUNTIFS(\'Test Cases\'!B:B,"{mod_name}",\'Test Cases\'!K:K,"Not Run")')
        sm.cell(row=r, column=4, value=f'=COUNTIFS(\'Test Cases\'!B:B,"{mod_name}",\'Test Cases\'!K:K,"Pass")')
        sm.cell(row=r, column=5, value=f'=COUNTIFS(\'Test Cases\'!B:B,"{mod_name}",\'Test Cases\'!K:K,"Fail")')
        sm.cell(row=r, column=6, value=f'=COUNTIFS(\'Test Cases\'!B:B,"{mod_name}",\'Test Cases\'!K:K,"Blocked")')
        sm.cell(row=r, column=7, value=f'=COUNTIFS(\'Test Cases\'!B:B,"{mod_name}",\'Test Cases\'!K:K,"Skipped")')
        sm.cell(row=r, column=8, value=f'=IFERROR(D{r}/(D{r}+E{r}),"-")')

        for col in range(2, 9):
            cell = sm.cell(row=r, column=col)
            cell.font = BASE_FONT
            cell.alignment = CENTER
            cell.border = BORDER
            if col == 8:
                cell.number_format = "0.0%;0.0%;-"
        r += 1

    total_row = r
    sm.cell(row=total_row, column=1, value="TOTAL").font = SECTION_FONT
    sm.cell(row=total_row, column=1).fill = SECTION_FILL
    sm.cell(row=total_row, column=1).border = BORDER

    for col in range(2, 8):
        col_letter = get_column_letter(col)
        sm.cell(row=total_row, column=col,
                value=f"=SUM({col_letter}4:{col_letter}{total_row-1})")
        cell = sm.cell(row=total_row, column=col)
        cell.font = BOLD_FONT
        cell.fill = PatternFill("solid", start_color="EDE9FE")
        cell.alignment = CENTER
        cell.border = BORDER

    sm.cell(row=total_row, column=8,
            value=f"=IFERROR(D{total_row}/(D{total_row}+E{total_row}),\"-\")")
    cell = sm.cell(row=total_row, column=8)
    cell.font = BOLD_FONT
    cell.fill = PatternFill("solid", start_color="EDE9FE")
    cell.alignment = CENTER
    cell.border = BORDER
    cell.number_format = "0.0%;0.0%;-"

    # By priority
    r = total_row + 3
    sm.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    c = sm.cell(row=r, column=1, value="Counts by Priority")
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = CENTER
    sm.row_dimensions[r].height = 22
    r += 1

    for i, h in enumerate(sm_headers, 1):
        h2 = "Priority" if i == 1 else h
        c = sm.cell(row=r, column=i, value=h2)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    r += 1

    for prio in ["High", "Medium", "Low"]:
        sm.cell(row=r, column=1, value=prio).font = BOLD_FONT
        sm.cell(row=r, column=1).alignment = CENTER
        sm.cell(row=r, column=1).border = BORDER
        sm.cell(row=r, column=1).fill = (
            PRIORITY_HIGH_FILL if prio == "High"
            else PRIORITY_MED_FILL if prio == "Medium"
            else PRIORITY_LOW_FILL
        )
        sm.cell(row=r, column=2, value=f'=COUNTIF(\'Test Cases\'!L:L,"{prio}")')
        sm.cell(row=r, column=3, value=f'=COUNTIFS(\'Test Cases\'!L:L,"{prio}",\'Test Cases\'!K:K,"Not Run")')
        sm.cell(row=r, column=4, value=f'=COUNTIFS(\'Test Cases\'!L:L,"{prio}",\'Test Cases\'!K:K,"Pass")')
        sm.cell(row=r, column=5, value=f'=COUNTIFS(\'Test Cases\'!L:L,"{prio}",\'Test Cases\'!K:K,"Fail")')
        sm.cell(row=r, column=6, value=f'=COUNTIFS(\'Test Cases\'!L:L,"{prio}",\'Test Cases\'!K:K,"Blocked")')
        sm.cell(row=r, column=7, value=f'=COUNTIFS(\'Test Cases\'!L:L,"{prio}",\'Test Cases\'!K:K,"Skipped")')
        sm.cell(row=r, column=8, value=f'=IFERROR(D{r}/(D{r}+E{r}),"-")')
        for col in range(2, 9):
            cell = sm.cell(row=r, column=col)
            cell.font = BASE_FONT
            cell.alignment = CENTER
            cell.border = BORDER
            if col == 8:
                cell.number_format = "0.0%;0.0%;-"
        r += 1

    # By test type
    r += 2
    sm.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    c = sm.cell(row=r, column=1, value="Counts by Test Type")
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = CENTER
    sm.row_dimensions[r].height = 22
    r += 1

    for i, h in enumerate(sm_headers, 1):
        h2 = "Test Type" if i == 1 else h
        c = sm.cell(row=r, column=i, value=h2)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER
    r += 1

    for ttype in ["Functional", "Negative", "Edge Case", "Security", "A11y",
                  "Performance", "Compatibility", "UX/UI", "Regression"]:
        sm.cell(row=r, column=1, value=ttype).font = BOLD_FONT
        sm.cell(row=r, column=1).alignment = CENTER
        sm.cell(row=r, column=1).border = BORDER
        sm.cell(row=r, column=2, value=f'=COUNTIF(\'Test Cases\'!D:D,"{ttype}")')
        sm.cell(row=r, column=3, value=f'=COUNTIFS(\'Test Cases\'!D:D,"{ttype}",\'Test Cases\'!K:K,"Not Run")')
        sm.cell(row=r, column=4, value=f'=COUNTIFS(\'Test Cases\'!D:D,"{ttype}",\'Test Cases\'!K:K,"Pass")')
        sm.cell(row=r, column=5, value=f'=COUNTIFS(\'Test Cases\'!D:D,"{ttype}",\'Test Cases\'!K:K,"Fail")')
        sm.cell(row=r, column=6, value=f'=COUNTIFS(\'Test Cases\'!D:D,"{ttype}",\'Test Cases\'!K:K,"Blocked")')
        sm.cell(row=r, column=7, value=f'=COUNTIFS(\'Test Cases\'!D:D,"{ttype}",\'Test Cases\'!K:K,"Skipped")')
        sm.cell(row=r, column=8, value=f'=IFERROR(D{r}/(D{r}+E{r}),"-")')
        for col in range(2, 9):
            cell = sm.cell(row=r, column=col)
            cell.font = BASE_FONT
            cell.alignment = CENTER
            cell.border = BORDER
            if col == 8:
                cell.number_format = "0.0%;0.0%;-"
        r += 1

    wb.save(out_path)
    return len(relevant)


os.makedirs(OUT_DIR, exist_ok=True)
web_count = build_workbook("web", os.path.join(OUT_DIR, "web-app-manual-test-cases.xlsx"))
desktop_count = build_workbook("desktop", os.path.join(OUT_DIR, "desktop-app-manual-test-cases.xlsx"))
print(f"Web file: {web_count} tests")
print(f"Desktop file: {desktop_count} tests")
print(f"Total unique entries: {len(TESTS)}")
