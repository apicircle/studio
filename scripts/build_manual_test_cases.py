"""Build APICircle Studio Manual Test Cases workbook."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

OUT = r"C:/Local Development/APICircle/studio/docs/qa/manual-test-cases.xlsx"

# Style constants
FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="4F46E5")  # indigo
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=14)
SECTION_FILL = PatternFill("solid", start_color="6D28D9")  # purple
SECTION_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=12)
BASE_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, bold=True, size=10)
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
CENTER = Alignment(wrap_text=True, vertical="center", horizontal="center")
THIN = Side(border_style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PRIORITY_HIGH_FILL = PatternFill("solid", start_color="FEE2E2")
PRIORITY_MED_FILL = PatternFill("solid", start_color="FEF3C7")
PRIORITY_LOW_FILL = PatternFill("solid", start_color="DCFCE7")

PASS_FILL = PatternFill("solid", start_color="BBF7D0")
FAIL_FILL = PatternFill("solid", start_color="FECACA")
BLOCKED_FILL = PatternFill("solid", start_color="FED7AA")
NOT_RUN_FILL = PatternFill("solid", start_color="E5E7EB")
SKIPPED_FILL = PatternFill("solid", start_color="DBEAFE")

# Module definitions: (code, name, color)
MODULES = [
    ("WS", "Workspace Management"),
    ("CR", "Collections & Requests"),
    ("RE", "Request Editor"),
    ("BE", "Body Editor"),
    ("AU", "Authentication"),
    ("O2", "OAuth2 Flows"),
    ("VR", "Variables & Environments"),
    ("RP", "Response Panel"),
    ("SC", "Pre-request Scripts & Tests"),
    ("HS", "History"),
    ("CO", "Cookies"),
    ("MK", "Mock Servers"),
    ("ST", "Settings & Theming"),
    ("IE", "Import / Export"),
    ("CL", "CLI"),
    ("DS", "Desktop-Specific"),
    ("GT", "Git Integration"),
    ("GQ", "GraphQL"),
    ("AS", "Assertions & Plans"),
    ("DC", "Documentation Viewer"),
    ("SE", "Search & Marketplace"),
    ("KB", "Keyboard Shortcuts"),
    ("AL", "Accessibility & Theming"),
    ("CC", "Cross-Cutting (Toasts/Dialogs/DnD)"),
]

# Test case schema: (module_code, sub_feature, title, preconditions, steps, test_data, expected, priority)
T = []  # list of test cases


def add(mod, sub, title, pre, steps, data, expected, prio="Medium"):
    T.append((mod, sub, title, pre, steps, data, expected, prio))


# ============================================================
# WORKSPACE MANAGEMENT (WS)
# ============================================================
add("WS", "Workspace Create", "Create a new local workspace",
    "App is launched; no workspace currently open.",
    "1. Open the Workspace switcher in the top bar.\n2. Click 'Create new workspace'.\n3. Enter name 'QA-Smoke-WS' and confirm.",
    "Name: QA-Smoke-WS",
    "A new empty workspace 'QA-Smoke-WS' is created and becomes the active workspace; the explorer shows zero collections and the workspace name appears in the top bar.",
    "High")

add("WS", "Workspace Create", "Reject blank workspace name",
    "Create workspace dialog is open.",
    "1. Leave name field empty.\n2. Click Create.",
    "Name: (empty)",
    "Create button is disabled or validation toast 'Workspace name is required' appears; no workspace is created.",
    "Medium")

add("WS", "Workspace Create", "Workspace name supports unicode and emoji",
    "Create workspace dialog is open.",
    "1. Enter name 'Café 🚀 测试'.\n2. Click Create.",
    "Name: Café 🚀 测试",
    "Workspace is created with the unicode/emoji name rendered correctly in the top bar and switcher.",
    "Low")

add("WS", "Workspace Switcher", "Switch between two workspaces",
    "Two workspaces 'A' and 'B' exist.",
    "1. Click Workspace switcher.\n2. Select workspace 'B'.\n3. Wait for hydrate.",
    "Existing workspaces A & B",
    "Active workspace changes to 'B'; the explorer, environments, history reflect B's data; top-bar title updates.",
    "High")

add("WS", "Workspace Switcher", "Recent workspaces persist after app restart",
    "User has opened 3 workspaces this session.",
    "1. Quit and re-launch the app.\n2. Open Workspace switcher.",
    "3 prior workspaces",
    "All 3 workspaces appear in the recent list in last-opened order; the last-active workspace is loaded automatically.",
    "High")

add("WS", "Workspace Delete", "Delete workspace requires confirmation",
    "Workspace 'Disposable' is active.",
    "1. Open Workspace settings.\n2. Click 'Delete workspace'.\n3. Observe confirm dialog.\n4. Click Cancel.",
    "Workspace: Disposable",
    "A confirm dialog with destructive styling appears; on Cancel, no deletion occurs and the workspace remains.",
    "High")

add("WS", "Workspace Delete", "Confirm deletion removes workspace from registry",
    "Workspace 'Disposable' is active.",
    "1. Open Workspace settings.\n2. Click 'Delete workspace'.\n3. Confirm the destructive action.",
    "Workspace: Disposable",
    "Workspace 'Disposable' is removed from the registry; app switches to another workspace or empty state; deleted workspace does not reappear after restart.",
    "High")

add("WS", "Workspace Link to Git", "Link a fresh workspace to a GitHub repo",
    "Workspace 'QA-Link' has no git link; user has GitHub account.",
    "1. Click 'Link workspace' in Workspace panel.\n2. Authorize GitHub OAuth.\n3. Pick or create a target repo.\n4. Confirm linking.",
    "GitHub repo: test/qa-link",
    "Workspace is linked; a new working branch is created in the repo; Push/Pull buttons become enabled; status badge shows 'Linked'.",
    "High")

add("WS", "Workspace Link to Git", "OAuth scope rejection blocks linking",
    "Workspace is unlinked.",
    "1. Begin link flow.\n2. On GitHub consent screen, click 'Cancel' or deny scopes.",
    "—",
    "App surfaces an error toast 'GitHub authorization was cancelled' or similar; workspace remains unlinked; no working branch is created.",
    "Medium")

add("WS", "Workspace Push", "Push edits to working branch",
    "Workspace linked to git; user has made local edits to at least one request.",
    "1. Click Push button in Workspace panel.\n2. Wait for completion.",
    "Modified request URL",
    "A new commit is pushed to the working branch; commit metadata contains app version and timestamp; sync snapshot updates; 'Push' button returns to idle state.",
    "High")

add("WS", "Workspace Push", "Push with no local changes is a no-op",
    "Workspace linked; no local edits.",
    "1. Click Push.",
    "—",
    "Push button shows 'No changes' or completes silently; no new commit is created on the remote branch.",
    "Medium")

add("WS", "Workspace Pull", "Pull updates from remote branch",
    "Working branch was updated remotely (e.g., from another device).",
    "1. Click Pull.\n2. Wait for completion.",
    "Remote commit on working branch",
    "Latest changes are applied to the local workspace; sync snapshot updates; UI reflects the new collections/envs.",
    "High")

add("WS", "Workspace Refresh", "Refresh detects retired branch (PR merged)",
    "Working branch was merged into main on GitHub and deleted.",
    "1. Click Refresh.",
    "Merged PR",
    "App detects branch retirement and prompts user to create a new working branch or switch back to main; retired branch is recorded.",
    "High")

add("WS", "Workspace Reset", "Reset discards local edits",
    "Workspace has uncommitted local edits.",
    "1. Open Workspace settings.\n2. Click 'Reset workspace to last sync'.\n3. Confirm.",
    "Edited request",
    "All local edits are discarded; workspace returns to the last-synced state; the response history is preserved (history is local-only).",
    "High")

add("WS", "Offline Workspace", "Create offline workspace has no git affordances",
    "Workspace switcher is open.",
    "1. Choose 'Create offline workspace'.\n2. Name it 'Local-Only'.",
    "Name: Local-Only",
    "Workspace is created with no Push/Pull buttons visible; 'Link to git' option is offered; data persists in IndexedDB only.",
    "Medium")

add("WS", "Workspace Hydrate", "Workspace requiring passphrase prompts on open",
    "Workspace has at least one secret variable; user has logged out / new session.",
    "1. Open the workspace.\n2. Observe passphrase prompt.\n3. Enter correct passphrase.",
    "Correct passphrase",
    "Passphrase modal appears before secrets are decrypted; on correct entry, secrets unlock; UI proceeds normally.",
    "High")

add("WS", "Workspace Hydrate", "Wrong passphrase keeps secrets locked",
    "Workspace has secrets and a passphrase.",
    "1. On passphrase modal, enter an incorrect passphrase.\n2. Submit.",
    "Wrong passphrase",
    "Error message 'Incorrect passphrase' is shown; secrets remain encrypted; user can retry or skip (with secrets unavailable).",
    "High")

# ============================================================
# COLLECTIONS & REQUESTS (CR)
# ============================================================
add("CR", "Collection CRUD", "Create a new collection at root",
    "Workspace is open.",
    "1. In Editor sidebar, click + Add → 'Collection'.\n2. Name it 'Smoke Suite'.",
    "Name: Smoke Suite",
    "Collection appears at root of the tree, selected, empty; rename input is auto-focused on initial create.",
    "High")

add("CR", "Collection CRUD", "Rename collection inline",
    "A collection 'Old Name' exists.",
    "1. Right-click collection → Rename (or double-click).\n2. Enter 'New Name'.\n3. Press Enter.",
    "New Name: New Name",
    "Collection name updates in the tree and any open request tabs reflect the new collection name in breadcrumbs.",
    "Medium")

add("CR", "Collection CRUD", "Delete empty collection",
    "Empty collection 'Trash' exists.",
    "1. Right-click 'Trash' → Delete.\n2. Confirm.",
    "—",
    "Collection is removed from tree; no orphaned requests remain.",
    "Medium")

add("CR", "Collection CRUD", "Delete collection with children warns and cascades",
    "Collection with 2 folders and 5 requests exists.",
    "1. Right-click collection → Delete.\n2. Read confirm message.\n3. Confirm.",
    "Collection: Parent (2 folders, 5 requests)",
    "Confirm dialog explicitly mentions child count; on confirm, all descendants are removed; tree updates; no zombie history entries reference deleted IDs.",
    "High")

add("CR", "Folder CRUD", "Create folder under collection",
    "Collection 'API v1' exists.",
    "1. Right-click 'API v1' → Add folder.\n2. Name 'Users'.",
    "Name: Users",
    "Folder 'Users' is created as a child of 'API v1'; can be expanded/collapsed.",
    "High")

add("CR", "Folder CRUD", "Nested folders to depth 5",
    "Empty collection exists.",
    "1. Create folder A under collection.\n2. Create folder B under A.\n3. Create C under B; D under C; E under D.",
    "5-level nesting",
    "All 5 levels render; tree indent is visually distinct at each level; expand/collapse works at each level.",
    "Low")

add("CR", "Request CRUD", "Create new request via Ctrl+N",
    "Editor panel is focused; collection 'API v1' is selected.",
    "1. Press Ctrl/Cmd+N.",
    "Shortcut: Ctrl+N",
    "A new GET request named 'Untitled Request' is created under the selected collection; editor opens with focus on URL bar.",
    "High")

add("CR", "Request CRUD", "Duplicate request copies all fields",
    "A request with body, headers, auth, tests exists.",
    "1. Right-click request → Duplicate.",
    "Source request fully configured",
    "A new request is created with name 'X (copy)' and all fields (URL, method, headers, body, auth, scripts, tests) cloned; the original is unmodified.",
    "Medium")

add("CR", "Request CRUD", "Delete request removes from tree only",
    "Request 'Get User' exists with prior run history.",
    "1. Right-click request → Delete.\n2. Confirm.",
    "Request: Get User",
    "Request is removed from tree; existing history entries for that request remain viewable in History panel (showing URL/method) even though the source request no longer exists.",
    "Medium")

add("CR", "Tree Reorder", "Drag request to different folder",
    "Two folders A and B; request R in A.",
    "1. Drag request R from A and drop on B.",
    "Request R, folders A and B",
    "Request R appears as a child of B; A no longer contains it; tree state persists after refresh/push.",
    "Medium")

add("CR", "Tree Search/Filter", "Filter requests by name substring",
    "Collection with many requests.",
    "1. Type 'login' in the explorer search box.",
    "Search: login",
    "Tree filters to only show folders/requests whose name contains 'login' (case-insensitive); matches are highlighted; clearing search restores full tree.",
    "Medium")

# ============================================================
# REQUEST EDITOR (RE)
# ============================================================
add("RE", "URL Bar", "Send simple GET request",
    "Request open in editor; method GET.",
    "1. Enter URL https://httpbin.org/get.\n2. Click Send.",
    "URL: https://httpbin.org/get",
    "Request is sent; response panel shows 200 status, time in ms, size in bytes; body shows JSON pretty-printed.",
    "High")

add("RE", "URL Bar", "Variable interpolation in URL",
    "Environment has var baseUrl = https://httpbin.org",
    "1. Enter URL '{{baseUrl}}/get'.\n2. Hover the variable.\n3. Click Send.",
    "{{baseUrl}}/get",
    "Hover popover shows the resolved value 'https://httpbin.org'; on Send, request resolves and returns 200.",
    "High")

add("RE", "URL Bar", "Undefined variable resolves to empty string",
    "No environment vars set.",
    "1. Enter URL 'https://{{missing}}.example.com'.\n2. Send.",
    "{{missing}}",
    "URL becomes 'https://.example.com'; request is attempted and fails with a meaningful network error (toast or response error panel).",
    "Medium")

add("RE", "Method Picker", "Switch method GET → POST",
    "Request is GET.",
    "1. Open method dropdown.\n2. Select POST.",
    "Method: POST",
    "Method updates to POST; Body tab becomes more prominent; if no body was set, default body type is 'none'.",
    "High")

add("RE", "Method Picker", "All supported methods present",
    "Method dropdown is open.",
    "1. Scroll/inspect the list.",
    "—",
    "GET, POST, PUT, PATCH, DELETE, HEAD, OPTIONS are all available; custom methods either supported or clearly absent.",
    "Medium")

add("RE", "Params Tab", "Add query param updates URL",
    "URL is https://httpbin.org/get.",
    "1. Open Params tab.\n2. Add row key=name, value=alice.\n3. Observe URL bar.",
    "name=alice",
    "URL bar shows 'https://httpbin.org/get?name=alice'; on Send, response echoes the query param.",
    "High")

add("RE", "Params Tab", "Toggle disable a param removes it from URL",
    "Request has params name=alice and age=30 enabled.",
    "1. Uncheck 'age'.",
    "—",
    "URL bar shows only 'name=alice'; sent request does not include 'age'.",
    "Medium")

add("RE", "Params Tab", "URL-paste parses query into rows",
    "Params tab is empty.",
    "1. Paste 'https://api.example.com/items?status=open&limit=10' into URL bar.",
    "URL with 2 params",
    "Params tab auto-populates with two rows: status=open, limit=10; URL bar shows the path without query string in synced state.",
    "Medium")

add("RE", "Headers Tab", "Add custom header",
    "Request open.",
    "1. Headers tab → add row 'X-Trace-Id' = 'abc123'.\n2. Send.",
    "Header X-Trace-Id: abc123",
    "Request is sent with the header; server response (httpbin echo) shows it received the header.",
    "High")

add("RE", "Headers Tab", "Header autocomplete suggests standard names",
    "Headers tab focused on an empty row.",
    "1. Type 'auth' in key field.",
    "Partial: auth",
    "Autocomplete dropdown shows Authorization, Authentication-Info, etc.; selecting an item populates the row.",
    "Low")

add("RE", "Send", "Cancel in-flight request",
    "A slow endpoint is being requested.",
    "1. Send request to slow endpoint (e.g., 30s delay).\n2. While 'Sending…', click Cancel.",
    "Slow endpoint",
    "Request is aborted; response panel shows 'Cancelled' state; no response body is committed to history (or marked aborted).",
    "Medium")

add("RE", "Send", "Send via Ctrl+Enter",
    "URL is valid; cursor anywhere in editor.",
    "1. Press Ctrl/Cmd+Enter.",
    "Shortcut",
    "Request is sent regardless of focus (URL, body, headers); response panel updates.",
    "High")

add("RE", "Tabs", "Editor tab persists active sub-tab across requests",
    "Request A: Body tab open. Request B exists.",
    "1. Switch to Request B.\n2. Switch back to A.",
    "—",
    "Request A still shows Body tab active (last-used tab is remembered per-request, or globally as designed).",
    "Low")

# ============================================================
# BODY EDITOR (BE)
# ============================================================
add("BE", "Body Type", "Default body type for GET is none",
    "New GET request created.",
    "1. Open Body tab.",
    "—",
    "Body type selector defaults to 'none'; no editor surface is shown beyond the picker.",
    "Medium")

add("BE", "Form Data", "Add text field and submit",
    "POST request to httpbin.org/post; body type form-data.",
    "1. Add row name=username, value=alice.\n2. Send.",
    "field username=alice",
    "Request is sent with multipart/form-data; response echo shows form.username = alice; Content-Type header set automatically.",
    "High")

add("BE", "Form Data", "Upload file in form-data",
    "POST request; form-data body.",
    "1. Add row, type File.\n2. Pick a 1KB text file from disk.\n3. Send.",
    "File: sample.txt (1KB)",
    "File is uploaded; response shows files.<key> with content; file attachment is stored in IndexedDB and persists if request is re-opened.",
    "High")

add("BE", "URL-encoded", "Submit application/x-www-form-urlencoded",
    "POST request; body type x-www-form-urlencoded.",
    "1. Add a=1, b=hello.\n2. Send to httpbin.org/post.",
    "a=1&b=hello",
    "Body sent as 'a=1&b=hello'; Content-Type = application/x-www-form-urlencoded; response echoes form fields.",
    "High")

add("BE", "Raw JSON", "Submit JSON body",
    "POST request; body type raw → JSON.",
    "1. Paste {\"name\":\"alice\",\"age\":30}.\n2. Send.",
    "JSON body",
    "Content-Type = application/json; response.json shows the parsed object; pretty-print formats body in editor.",
    "High")

add("BE", "Raw JSON", "Invalid JSON shows inline error",
    "Body type raw JSON.",
    "1. Type {\"a\": } (invalid).",
    "Invalid JSON",
    "Monaco shows syntax-error squiggle; Send still works (server may reject); the editor does not block edits.",
    "Medium")

add("BE", "Raw XML", "XML body sends correctly",
    "POST request; body type raw → XML.",
    "1. Enter <root><x>1</x></root>.\n2. Send.",
    "XML body",
    "Content-Type defaults to application/xml; body is sent verbatim; XML highlighting visible in editor.",
    "Medium")

add("BE", "Binary", "Binary file upload sends raw bytes",
    "POST request; body type binary.",
    "1. Pick a 10KB PNG.\n2. Send.",
    "Binary file: 10KB PNG",
    "Body sent with the file's bytes; Content-Type from file extension; response echoes file size.",
    "Medium")

add("BE", "GraphQL Body", "Send simple GraphQL query",
    "POST request to a GraphQL endpoint.",
    "1. Body type → GraphQL.\n2. Query: { user(id: 1) { name } }.\n3. Send.",
    "GraphQL query",
    "Request body wraps the query as {\"query\": \"...\"}; Content-Type is application/json; response is returned.",
    "High")

add("BE", "GraphQL Body", "Variables panel sent alongside query",
    "GraphQL request open.",
    "1. Set query: query($id: ID!) { user(id:$id) { name } }.\n2. Variables: {\"id\": \"1\"}.\n3. Send.",
    "Variables JSON",
    "Body includes both query and variables fields; resolved correctly server-side.",
    "Medium")

add("BE", "Body Type Switch", "Switching body type clears incompatible content",
    "Body is raw JSON with content.",
    "1. Switch body type to 'none'.\n2. Switch back to raw JSON.",
    "—",
    "After switching away, body content is cleared (or stashed and restorable, per design); switching back gives an empty editor.",
    "Low")

# ============================================================
# AUTHENTICATION (AU)
# ============================================================
add("AU", "Auth Type Picker", "All 17 auth types are listed",
    "Auth tab open.",
    "1. Open auth-type dropdown.\n2. Inspect entries.",
    "—",
    "None, Inherit, Basic, Bearer, API Key, OAuth1, OAuth2 (with grants), Digest, NTLM, Hawk, AWS SigV4, JWT Bearer, ASAP, EdgeGrid (Akamai), Custom Header are visible.",
    "Medium")

add("AU", "Bearer", "Bearer token added to Authorization header",
    "Request to httpbin.org/bearer; auth type Bearer.",
    "1. Set token = 'abc.def.ghi'.\n2. Send.",
    "Token: abc.def.ghi",
    "Authorization: Bearer abc.def.ghi is sent; httpbin returns 200 with authenticated=true.",
    "High")

add("AU", "Basic", "Basic auth base64-encodes credentials",
    "httpbin.org/basic-auth/user/pass with basic auth.",
    "1. username=user, password=pass.\n2. Send.",
    "user/pass",
    "Authorization: Basic dXNlcjpwYXNz is sent; response 200.",
    "High")

add("AU", "API Key", "API key in header",
    "Auth type API Key; placement Header.",
    "1. key=X-Api-Key, value=secret-123.\n2. Send to httpbin.org/headers.",
    "X-Api-Key: secret-123",
    "Header is added to outgoing request; response echo confirms.",
    "High")

add("AU", "API Key", "API key in query",
    "Auth type API Key; placement Query.",
    "1. key=api_key, value=abc.\n2. Send.",
    "Query: api_key=abc",
    "URL has '?api_key=abc' appended on send; response confirms; Params tab does NOT show this row (auth-managed).",
    "Medium")

add("AU", "Digest", "Digest auth completes 401 challenge-response",
    "Endpoint requires Digest auth.",
    "1. Set username/password.\n2. Send.",
    "Valid digest credentials",
    "First request gets 401 with WWW-Authenticate; client computes response hash with nonce; second request returns 200; only the final response is shown.",
    "High")

add("AU", "Digest", "Digest stale=true rotates nonce",
    "Server returns 401 stale=true after a successful request.",
    "1. Send a request that triggers nonce rotation.",
    "—",
    "Client detects stale=true and retries with new nonce; succeeds without user intervention.",
    "Low")

add("AU", "NTLM", "NTLM 3-way handshake completes",
    "Endpoint requires NTLM; username/password/domain set.",
    "1. Send request.",
    "user/pass/domain",
    "Negotiate → Challenge → Authenticate exchange happens; final 200 response is shown; MIC is computed on Authenticate.",
    "High")

add("AU", "Hawk", "Hawk MAC signature accepted",
    "Endpoint accepts Hawk; hawkId/hawkKey configured.",
    "1. Send.",
    "Valid Hawk ID/key",
    "Authorization: Hawk … header sent; server verifies MAC; 200 response.",
    "Medium")

add("AU", "AWS SigV4", "AWS SigV4 signs GET request",
    "Auth type AWS SigV4; valid accessKeyId/secretAccessKey/region/service.",
    "1. Send GET to an AWS-signed endpoint (or mock).",
    "AWS creds + region",
    "Authorization header has 'AWS4-HMAC-SHA256 Credential=…' format; x-amz-date header is added; signing canonicalizes correctly.",
    "High")

add("AU", "AWS SigV4", "AWS SigV4 signs POST with body",
    "POST with JSON body to S3-style endpoint.",
    "1. Body: JSON.\n2. Send.",
    "POST + JSON",
    "x-amz-content-sha256 reflects the SHA256 of the body; signature validates server-side (or via local verifier).",
    "Medium")

add("AU", "JWT Bearer", "JWT signed with HS256",
    "Auth type JWT; alg HS256; secret set.",
    "1. Configure claims (iss, sub, exp).\n2. Send.",
    "HS256 secret",
    "Authorization: Bearer <signed-jwt> is sent; signature verifies; payload matches claims.",
    "Medium")

add("AU", "JWT Bearer", "JWT signed with RS256 using PEM key",
    "Auth type JWT; alg RS256; PEM private key.",
    "1. Configure claims.\n2. Send.",
    "RS256 PEM key",
    "JWT signed correctly; verifies with matching public key.",
    "Medium")

add("AU", "Inherit", "Folder-level auth inherited by child request",
    "Folder has Bearer auth set; child request has auth=Inherit.",
    "1. Send child request.",
    "Inherited bearer",
    "Folder's Bearer token is applied to the child request; Authorization header is present.",
    "High")

add("AU", "Inherit", "Request-level auth overrides folder",
    "Folder has Bearer; child request has Basic auth.",
    "1. Send child.",
    "—",
    "Child request uses Basic auth, NOT the folder's Bearer; folder auth is ignored.",
    "High")

# ============================================================
# OAUTH2 (O2)
# ============================================================
add("O2", "Client Credentials", "Acquire token via client_credentials",
    "Mock IdP configured; clientId/secret set.",
    "1. Auth tab → OAuth2 → Client Credentials.\n2. Click 'Get token'.",
    "Valid client creds",
    "Token request POST to /token; access_token returned and stored; UI shows 'Token acquired'; obtainedScope/expiresAt populated.",
    "High")

add("O2", "Authorization Code", "Auth code flow via web popup",
    "Web build; mock IdP; auth code grant configured.",
    "1. Click 'Get token'.\n2. Popup opens authorize URL.\n3. Approve.",
    "Auth code grant",
    "Popup completes; access_token stored; refresh_token (if returned) stored; popup closes; UI updates.",
    "High")

add("O2", "Authorization Code", "Desktop callback server captures code",
    "Desktop build; auth code flow.",
    "1. Click 'Get token'.\n2. Native browser opens authorize URL.\n3. Approve.",
    "Auth code grant",
    "localhost server on a free port receives the callback; access_token stored; redirect URI auto-set; user does not need to copy/paste anything.",
    "High")

add("O2", "PKCE", "Authorization code with PKCE (public client)",
    "Client without secret; PKCE enabled.",
    "1. Get token.",
    "PKCE S256",
    "code_verifier auto-generated (43–128 chars); code_challenge=S256(verifier); token exchange includes verifier; succeeds.",
    "High")

add("O2", "Password", "Resource owner password grant",
    "Grant=password; username/password set.",
    "1. Get token.",
    "user/pass",
    "POST /token with grant_type=password; access_token returned; stored.",
    "Medium")

add("O2", "Implicit", "Implicit grant returns access_token in URL fragment",
    "Grant=implicit; configured.",
    "1. Get token.",
    "—",
    "Authorize URL completes with token in fragment; client extracts and stores; no refresh_token expected.",
    "Medium")

add("O2", "Device Code", "Device flow polls token endpoint",
    "Grant=device_code.",
    "1. Get token.\n2. Approve on device verification URL in another tab.",
    "—",
    "User_code displayed in UI; polling interval respected; on approval, access_token is stored.",
    "Medium")

add("O2", "Refresh", "Manual token refresh",
    "Token already obtained with refresh_token.",
    "1. Click 'Refresh token'.",
    "—",
    "POST /token with grant_type=refresh_token; new access_token stored; expiresAt updated.",
    "High")

add("O2", "Auto-refresh", "Send auto-refreshes expired token",
    "Token expired; refresh_token present.",
    "1. Click Send on request using this auth.",
    "Expired token",
    "Token is refreshed transparently before the request goes out; outgoing request uses fresh token; no manual interaction.",
    "High")

add("O2", "Clear", "Clear token forces re-authentication",
    "Token present.",
    "1. Click 'Clear token'.\n2. Click Send.",
    "—",
    "Token removed; on send, user is required to re-acquire token (toast/prompt) or send fails with auth error.",
    "Medium")

add("O2", "CSRF/State", "State mismatch on callback is rejected",
    "Auth code flow in progress.",
    "1. Intercept callback URL.\n2. Replace state param with garbage.\n3. Return to client.",
    "Tampered state",
    "Client rejects callback with state-mismatch error; no token is stored; toast informs user.",
    "High")

add("O2", "Redirect URI Mismatch", "IdP error surfaces meaningful message",
    "Redirect URI in client does not match IdP config.",
    "1. Get token.",
    "Misconfigured redirect URI",
    "Authorize endpoint returns error; client surfaces a clear error message; no token stored.",
    "Medium")

add("O2", "Private Key JWT", "Client authenticates via signed JWT assertion",
    "Grant=client_credentials; client_assertion_type=jwt-bearer; private key set.",
    "1. Get token.",
    "JWT private key",
    "Token request includes client_assertion (signed JWT); succeeds; access_token returned.",
    "Low")

# ============================================================
# VARIABLES & ENVIRONMENTS (VR)
# ============================================================
add("VR", "Environment CRUD", "Create new environment",
    "Workspace open.",
    "1. Open Environments panel.\n2. Click + new environment.\n3. Name 'Dev'.",
    "Name: Dev",
    "Environment 'Dev' is created with empty variables list; activated automatically or selectable from picker.",
    "High")

add("VR", "Environment CRUD", "Activate environment from dropdown",
    "Environments 'Dev' and 'Prod' exist.",
    "1. Click env picker in top bar or panel.\n2. Select 'Prod'.",
    "—",
    "Active environment switches to Prod; all {{var}} resolutions now use Prod values.",
    "High")

add("VR", "Environment CRUD", "Duplicate environment clones all vars",
    "Env 'Dev' has 5 vars.",
    "1. Right-click Dev → Duplicate.",
    "—",
    "New env 'Dev (copy)' is created with all 5 vars cloned at the same values; original Dev unmodified.",
    "Medium")

add("VR", "Environment CRUD", "Delete environment with confirm",
    "Env 'Old' exists.",
    "1. Right-click 'Old' → Delete.\n2. Confirm.",
    "—",
    "Env is removed; if active, falls back to no env or next env in priority order.",
    "Medium")

add("VR", "Variable", "Add plaintext variable",
    "Active env 'Dev' open.",
    "1. Variables tab → add baseUrl = https://dev.api.test.",
    "baseUrl",
    "Variable saves; {{baseUrl}} now resolves to 'https://dev.api.test' in URL/body/headers.",
    "High")

add("VR", "Variable", "Add secret variable masks input",
    "Active env open.",
    "1. Add row, mark type=secret.\n2. Enter value 'topsecret'.",
    "Secret var",
    "Value field shows dots/asterisks; on focus, shows actual value; persisted encrypted; never pushed to git.",
    "High")

add("VR", "Variable", "Variable autocomplete in URL",
    "Env has baseUrl=…",
    "1. In URL bar, type '{{ba'.",
    "—",
    "Suggestion popup shows 'baseUrl'; selecting completes to '{{baseUrl}}'.",
    "Medium")

add("VR", "Scope Resolution", "Request-level var overrides environment",
    "Env has 'x' = 'env'; request has 'x' = 'req' in pre-request script via pm.variables.set.",
    "1. Send request with body referencing {{x}}.",
    "—",
    "Sent body uses 'req'; environment value is shadowed.",
    "Medium")

add("VR", "Scope Resolution", "Workspace var fallback",
    "Workspace has 'app' = 'studio'; no env var with same name.",
    "1. Reference {{app}} in headers.\n2. Send.",
    "—",
    "Header resolves to 'studio'; workspace fallback works.",
    "Medium")

add("VR", "Encryption", "Passphrase change re-encrypts secrets",
    "Workspace has secrets and passphrase set.",
    "1. Settings → change passphrase.\n2. Provide old + new.",
    "Old + new passphrase",
    "All secrets re-encrypted with new key; old passphrase no longer works.",
    "High")

add("VR", "Linked Env", "Linked workspace env priority drag-reorder",
    "Workspace has 1 local env + 2 linked envs.",
    "1. Drag the second linked env above the first.",
    "—",
    "Priority order updates; variable resolution now follows new order; reflected on next send.",
    "Low")

# ============================================================
# RESPONSE PANEL (RP)
# ============================================================
add("RP", "Status", "200 OK shows green badge",
    "Successful request sent.",
    "1. Send a GET that returns 200.",
    "—",
    "Status code 200 with green/positive badge; time in ms; size in human-readable bytes (KB/MB).",
    "High")

add("RP", "Status", "Non-2xx shows distinct color",
    "Endpoint returns 404.",
    "1. Send.",
    "—",
    "Status 404 shown with red/error badge; response body still rendered.",
    "Medium")

add("RP", "Body Viewer", "Pretty/Raw/Preview toggle",
    "JSON response present.",
    "1. Click Pretty.\n2. Click Raw.\n3. Click Preview.",
    "JSON response",
    "Pretty shows formatted JSON; Raw shows minified single-line; Preview shows syntax-highlighted (or HTML preview for HTML content).",
    "High")

add("RP", "Body Viewer", "Large response shows preview cap notice",
    "Response > 1 MB.",
    "1. Send to endpoint returning 5 MB.",
    "5 MB body",
    "Preview is capped (e.g., first 1 MB); banner notifies user; Download button is available.",
    "Medium")

add("RP", "Headers", "Response headers tab lists all headers",
    "Response with multiple headers.",
    "1. Open Headers tab in response.",
    "—",
    "All response headers listed in key-value table; copyable; case preserved.",
    "Medium")

add("RP", "Cookies", "Set-Cookie header populates cookie jar",
    "Endpoint returns Set-Cookie: session=abc.",
    "1. Send.\n2. Open Cookies tab.",
    "—",
    "Cookies tab shows the new cookie with domain/path/expiry; subsequent requests to the same domain include it.",
    "High")

add("RP", "Transformations", "TOON/YAML/CSV savings shown for JSON",
    "JSON response received.",
    "1. Look at size hint area.",
    "—",
    "Badges show percent size reduction for TOON / YAML / CSV vs minified JSON; click to preview in selected format.",
    "Medium")

add("RP", "Transformations", "Switch to YAML preview",
    "JSON response open.",
    "1. Click 'as YAML' link.",
    "—",
    "Body viewer renders YAML representation of the JSON; switching back to JSON works.",
    "Low")

add("RP", "Download", "Download response body to file",
    "Any response.",
    "1. Click Download.",
    "—",
    "File save dialog opens; saved file matches response body bytes exactly.",
    "Medium")

add("RP", "Time/Size", "Time and size update on each send",
    "—",
    "1. Send twice.",
    "—",
    "Each send produces fresh time/size values reflecting the latest response.",
    "Low")

# ============================================================
# PRE-REQUEST SCRIPTS & TESTS (SC)
# ============================================================
add("SC", "Pre-request", "pm.variables.set persists across send",
    "Pre-request script: pm.variables.set('token', 'xyz')",
    "1. Set body using {{token}}.\n2. Send.",
    "—",
    "Body contains 'xyz' on send; variable available for the duration of the run.",
    "High")

add("SC", "Pre-request", "Runtime error aborts send",
    "Pre-request script: throw new Error('boom').",
    "1. Send.",
    "—",
    "Script error is surfaced (toast or panel); HTTP request is NOT sent.",
    "High")

add("SC", "Pre-request", "Async pm.sendRequest helper (if supported)",
    "Pre-request: pm.sendRequest({url:...}, cb).",
    "1. Send.",
    "—",
    "Helper executes and resolves before main send; if not supported, surfaces a clear 'not available' error.",
    "Low")

add("SC", "Console", "console.log appears in script console",
    "Pre-request: console.log('hello').",
    "1. Send.\n2. Open script console.",
    "—",
    "'hello' line appears in the console panel with timestamp.",
    "Medium")

add("SC", "Tests Tab", "Add status assertion",
    "Tests tab open.",
    "1. Add assertion: status equals 200.\n2. Send to a 200 endpoint.",
    "Assertion: status==200",
    "After response, Tests panel shows a green pass row; counts as 1/1 passing.",
    "High")

add("SC", "Tests Tab", "Add JSON path assertion",
    "Response: {\"id\":42}.",
    "1. Add assertion: body.id equals 42.\n2. Send.",
    "—",
    "Assertion passes; actual=42, expected=42 displayed.",
    "High")

add("SC", "Tests Tab", "Failing assertion shows red",
    "Response status 404.",
    "1. Add status==200 assertion.\n2. Send.",
    "—",
    "Assertion fails; row in red with expected vs actual; aggregate counter shows 0/1.",
    "Medium")

add("SC", "Tests Tab", "Regex matches assertion",
    "Response body: {\"email\":\"a@b.com\"}.",
    "1. Add assertion body.email matches ^[^@]+@[^@]+$.\n2. Send.",
    "—",
    "Passes; regex evaluated against string value.",
    "Medium")

# ============================================================
# HISTORY (HS)
# ============================================================
add("HS", "Logging", "Each send creates a history entry",
    "History panel open.",
    "1. Send 3 different requests.",
    "—",
    "3 entries appear at top of history with method, URL, status, time; newest first.",
    "High")

add("HS", "Date Buckets", "Today/Yesterday/older buckets",
    "Runs across multiple days exist.",
    "1. Open history.",
    "—",
    "Runs grouped under 'Today', 'Yesterday', 'Last 7 days', 'Last month', or month/year buckets.",
    "Medium")

add("HS", "Filter", "Filter by status range",
    "Mixed 2xx, 4xx, 5xx history.",
    "1. Apply filter: status 4xx.",
    "—",
    "Only 4xx entries are visible; counts update.",
    "Medium")

add("HS", "Replay", "Replay re-executes saved request state",
    "A run with body and headers exists.",
    "1. Click 'Replay' on a history entry.",
    "—",
    "Editor opens with the historical request state restored; user can edit before re-sending.",
    "Medium")

add("HS", "Delete Run", "Delete single history entry",
    "Run exists.",
    "1. Right-click run → Delete.\n2. Confirm.",
    "—",
    "Run is removed from list; storage freed.",
    "Low")

add("HS", "Bulk Clear", "Clear all history",
    "Multiple runs.",
    "1. Click 'Clear history'.\n2. Confirm.",
    "—",
    "All run entries are deleted; list shows empty state with help text.",
    "Medium")

add("HS", "Persistence", "History persists across app restarts",
    "5 runs in current session.",
    "1. Quit and re-open app.",
    "—",
    "All 5 runs remain visible in history; counts and metadata intact.",
    "Medium")

# ============================================================
# COOKIES (CO)
# ============================================================
add("CO", "Auto-populate", "Set-Cookie response stored in jar",
    "Response with Set-Cookie: a=1.",
    "1. Send.\n2. Open Cookie manager.",
    "—",
    "Cookie 'a=1' visible under the response domain; expiry set if specified.",
    "High")

add("CO", "Auto-send", "Stored cookies sent on next matching request",
    "Cookie a=1 stored for example.com.",
    "1. Send another request to example.com.",
    "—",
    "Outgoing request includes 'Cookie: a=1'.",
    "High")

add("CO", "Expiry", "Expired cookies not sent",
    "Cookie a=1 with Expires=past.",
    "1. Send request to its domain.",
    "—",
    "Cookie is not included; jar may show it as expired/struck-through.",
    "Medium")

add("CO", "Manual Edit", "User adds cookie manually",
    "Cookie manager open.",
    "1. Add cookie domain=example.com, name=tk, value=abc.",
    "—",
    "Cookie persisted; sent on subsequent example.com requests.",
    "Medium")

add("CO", "Clear", "Clear domain cookies",
    "Multiple cookies for example.com.",
    "1. Clear all for example.com.\n2. Confirm.",
    "—",
    "All example.com cookies removed; cookies for other domains unaffected.",
    "Medium")

# ============================================================
# MOCK SERVERS (MK)
# ============================================================
add("MK", "Definition", "Create mock server from scratch",
    "Mocks panel open.",
    "1. Click 'Create mock server'.\n2. Name 'Users API'.\n3. Confirm.",
    "Name: Users API",
    "Mock server entry appears in the list; edit page opens with empty endpoints.",
    "High")

add("MK", "Endpoint", "Add endpoint GET /users/:id with 200 JSON",
    "Mock 'Users API' selected.",
    "1. Add endpoint GET /users/:id.\n2. Add 200 response with JSON body {\"id\":1,\"name\":\"alice\"}.",
    "Endpoint + response",
    "Endpoint saved; visible in list; status 200 response defined.",
    "High")

add("MK", "Runtime (Desktop)", "Start mock server",
    "Desktop build; mock with 1 endpoint defined.",
    "1. Click Start.",
    "—",
    "Mock server starts on a free local port; status badge shows 'Running' with port; logs panel opens.",
    "High")

add("MK", "Runtime (Desktop)", "GET /users/:id returns defined response",
    "Mock running on port X.",
    "1. From within the app (or curl), GET http://localhost:X/users/1.",
    "—",
    "Response: 200 with the defined body; logs show received request.",
    "High")

add("MK", "Runtime (Desktop)", "Stop mock server",
    "Mock running.",
    "1. Click Stop.",
    "—",
    "Server stops; port freed; status badge shows 'Stopped'; logs preserved until panel cleared.",
    "Medium")

add("MK", "Web Build", "Mock runtime disabled in web",
    "Web build only.",
    "1. Open mocks panel; click Start.",
    "—",
    "Start button is disabled or shows tooltip 'Available in desktop app only'; definitions remain editable.",
    "Medium")

add("MK", "Spec Import", "Import OpenAPI generates mock endpoints",
    "An OpenAPI YAML file exists.",
    "1. Mocks → Import spec.\n2. Pick the YAML.",
    "OpenAPI 3.0 spec",
    "Endpoints populated from spec; stub responses generated from examples/schemas; user can refine.",
    "Medium")

add("MK", "Delete Mock", "Delete mock server",
    "Mock exists.",
    "1. Right-click mock → Delete.\n2. Confirm.",
    "—",
    "Mock and all its endpoints removed.",
    "Medium")

# ============================================================
# SETTINGS & THEMING (ST)
# ============================================================
add("ST", "Theme", "Switch from dark to light theme",
    "Active theme dark.",
    "1. Settings → Appearance → choose 'workbench-light'.",
    "—",
    "UI immediately re-themes light; Monaco editor follows; no reload required.",
    "High")

add("ST", "Theme", "High-contrast theme passes a11y contrast",
    "Themes list visible.",
    "1. Select 'high-contrast' theme.",
    "—",
    "Text/background contrast ratio ≥ 7:1; focus rings clearly visible.",
    "Medium")

add("ST", "Font", "Change code font to JetBrains Mono",
    "Settings open.",
    "1. Pick font 'jetbrains-mono'.",
    "—",
    "Monaco editor font updates immediately; UI sans font unaffected.",
    "Low")

add("ST", "Font Size", "Increase UI text size shortcut",
    "Default UI size 100%.",
    "1. Press Ctrl/Cmd+Shift+=. Repeat 3 times.",
    "—",
    "UI text scales up to 115%; layout adapts; size persists across reloads.",
    "Medium")

add("ST", "Font Size", "Reset UI text size",
    "UI size 115%.",
    "1. Press Ctrl/Cmd+Shift+0.",
    "—",
    "UI size resets to 100%.",
    "Low")

add("ST", "Per-Workspace Settings", "Theme is workspace-scoped",
    "Workspaces A and B exist.",
    "1. Set theme dark in A.\n2. Switch to B and set light.\n3. Switch back to A.",
    "—",
    "A shows dark, B shows light; each is remembered independently.",
    "Medium")

# ============================================================
# IMPORT / EXPORT (IE)
# ============================================================
add("IE", "Postman Import", "Import Postman v2.1 collection",
    "Postman v2.1 JSON file available.",
    "1. Workspace → Import → choose file.",
    "Postman v2.1 JSON",
    "Collection imported with folders and requests; URLs/methods/headers/body/auth (where compatible) preserved; summary modal lists counts and warnings.",
    "High")

add("IE", "Postman Import", "Unsupported auth falls back gracefully",
    "Postman collection with NTLM-style auth.",
    "1. Import.",
    "—",
    "NTLM auth either fully imported or downgraded to 'none' with a warning listed; import does not fail entirely.",
    "Medium")

add("IE", "cURL Import", "Paste cURL command creates request",
    "Editor open.",
    "1. Paste 'curl -X POST -H Content-Type:application/json -d {\"a\":1} https://api.test/x' into URL bar (or dedicated import).",
    "cURL command",
    "New request configured: POST URL, header, JSON body; user is prompted or it imports automatically per design.",
    "High")

add("IE", "Insomnia Import", "Import Insomnia collection",
    "Insomnia JSON export.",
    "1. Import via the same modal.",
    "—",
    "Insomnia structure parsed; requests/envs imported.",
    "Medium")

add("IE", "Export Workspace", "Export workspace as JSON",
    "Workspace has 5 requests.",
    "1. Workspace → Export → JSON.",
    "—",
    "File saved; valid WorkspaceSynced JSON; secrets excluded.",
    "Medium")

add("IE", "Copy as cURL", "Copy request as cURL",
    "Request open.",
    "1. Click 'Copy as cURL' (menu).",
    "—",
    "Clipboard contains a valid cURL command with method, headers, body, auth flags; pastable to a terminal.",
    "Medium")

# ============================================================
# CLI (CL)
# ============================================================
add("CL", "Invocation", "apicircle --help prints usage",
    "CLI installed; terminal open.",
    "1. Run 'apicircle --help'.",
    "—",
    "Usage and command list (mock, mcp, import) displayed; exit code 0.",
    "Medium")

add("CL", "Mock", "apicircle mock starts mock server",
    "Workspace folder with mock defined.",
    "1. Run 'apicircle mock ./workspace.json' (or path).",
    "—",
    "Mock server starts on configured port; logs to stdout; Ctrl+C cleanly stops.",
    "High")

add("CL", "Import", "apicircle import OpenAPI spec",
    "Local workspace folder; spec.yaml exists.",
    "1. Run 'apicircle import ./spec.yaml ./workspace'.",
    "—",
    "workspace.json updated with imported endpoints; summary line printed; non-zero exit on schema error.",
    "Medium")

add("CL", "MCP", "apicircle mcp starts MCP stdio server",
    "Workspace folder ready.",
    "1. Run 'apicircle mcp ./workspace'.",
    "—",
    "Process reads stdio for MCP protocol messages; responds to handshake; suitable for editor integration.",
    "Low")

add("CL", "Secrets", "Vault key passed via env var decrypts secrets",
    "Workspace with encrypted secrets.",
    "1. APICIRCLE_VAULT_SECRET_KEY=… apicircle mock ./workspace",
    "—",
    "Secrets decrypted at runtime; CLI does not log secret material.",
    "High")

# ============================================================
# DESKTOP-SPECIFIC (DS)
# ============================================================
add("DS", "Auto-Updater", "Update check on startup",
    "Desktop app launched.",
    "1. Launch app.\n2. Observe update logic.",
    "—",
    "Background check for new release; if new version exists, banner/dialog prompts user; cancel does not block app.",
    "Medium")

add("DS", "Auto-Updater", "Apply update on restart",
    "Update downloaded.",
    "1. Quit and re-open.",
    "—",
    "App restarts on the new version; version string updates; rollback path defined if signing fails.",
    "Medium")

add("DS", "Native Secret Bridge", "Save to OS Keychain/Credential Manager",
    "Desktop build; first secret saved.",
    "1. Add a secret variable.\n2. Quit and reopen without entering passphrase (if platform secret is available).",
    "—",
    "Secret was stored under the OS credential manager; decrypts automatically on next launch.",
    "High")

add("DS", "Window State", "Window bounds persist",
    "Resize app to 1200×800 at (200,200).",
    "1. Quit and reopen.",
    "—",
    "Window opens at the same size/position; multi-monitor scenarios handled gracefully.",
    "Low")

add("DS", "System Menu", "File menu items work",
    "App in focus.",
    "1. File → New Workspace.\n2. File → Import.",
    "—",
    "Each menu action triggers the corresponding flow; keyboard accelerators advertised in menu match actual bindings.",
    "Medium")

add("DS", "IPC", "Untrusted IPC senders rejected",
    "Renderer harness or test calls a privileged IPC channel from unauthorized origin.",
    "1. Trigger a forbidden IPC call.",
    "—",
    "Main process rejects the call; assertTrustedSender path logs/aborts; no privileged action executes.",
    "High")

# ============================================================
# GIT INTEGRATION (GT)
# ============================================================
add("GT", "Push Conflict", "Concurrent push fails second push",
    "Two devices pushed from same branch.",
    "1. Device A push.\n2. Device B push without pull.",
    "—",
    "Device B push fails with 'remote has changes'; user prompted to pull/merge first.",
    "High")

add("GT", "Three-way Merge", "Auto-merge non-conflicting edits",
    "Local edits modify request URL; remote edits modify request headers.",
    "1. Pull.",
    "—",
    "Both changes merged automatically; final request has both new URL and new headers; sync snapshot updated.",
    "High")

add("GT", "Three-way Merge", "Conflict surfaced for same-field edits",
    "Local URL=A; remote URL=B for same request.",
    "1. Pull.",
    "—",
    "Conflict modal lists the conflicting fields; user chooses theirs/mine/manual; resolution stored in next commit.",
    "High")

add("GT", "Branch Switch", "Switch working branch redownloads synced doc",
    "Workspace currently tracks branch X.",
    "1. Switch to branch Y.",
    "—",
    "Local synced data is replaced with branch Y's content; UI updates; local edits warned about (offer stash/discard).",
    "Medium")

add("GT", "Retired Branch", "Cleanup prompt after PR merge",
    "PR merged into main; working branch deleted.",
    "1. Refresh.",
    "—",
    "Prompt offers (a) create new working branch from main, (b) abandon local changes; choice respected.",
    "Medium")

add("GT", "Commit Metadata", "Commit author matches GitHub OAuth identity",
    "Linked workspace.",
    "1. Push edits.\n2. Inspect commit on GitHub.",
    "—",
    "Commit author/email match the OAuthed GitHub identity; co-author trailers (if any) accurate.",
    "Low")

# ============================================================
# GRAPHQL (GQ)
# ============================================================
add("GQ", "Schema Fetch", "Fetch schema via introspection",
    "GraphQL endpoint accessible.",
    "1. Body → GraphQL.\n2. Click 'Fetch Schema'.",
    "Endpoint URL",
    "Introspection query sent; schema saved to globalAssets; reused in completions.",
    "High")

add("GQ", "Completions", "Field autocomplete uses schema",
    "Schema fetched.",
    "1. In query editor, type 'query { us'.",
    "—",
    "Suggestions show fields like 'user' from the loaded schema.",
    "Medium")

add("GQ", "Variables", "Variables object validated against query signature",
    "Query: query($id:ID!) { user(id:$id) { name } }",
    "1. Set variables {\"id\": 1}.\n2. Send.",
    "—",
    "Numeric id coerced to ID; request executes; mismatched types surface server error and remain debuggable.",
    "Medium")

add("GQ", "Schema Reuse", "Stored schema reused after reload",
    "Workspace reopened.",
    "1. Open a GraphQL request.",
    "—",
    "Completions still work without re-fetching; schema persisted in workspace synced state.",
    "Low")

# ============================================================
# ASSERTIONS & PLANS (AS)
# ============================================================
add("AS", "Plan Create", "Create execution plan with 3 steps",
    "Workspace has multiple requests.",
    "1. Execution panel → New plan.\n2. Add 3 requests as steps.",
    "—",
    "Plan saved; step order is set; each step shows its source request.",
    "High")

add("AS", "Plan Run", "Run plan sequentially",
    "Plan with 3 steps.",
    "1. Click Run.",
    "—",
    "Each step runs in order; aggregate progress shows; per-step status (pass/fail) appears.",
    "High")

add("AS", "Plan Run", "Disable a step skips it",
    "Plan with step 2 disabled.",
    "1. Run plan.",
    "—",
    "Step 2 is shown as skipped; steps 1 and 3 run; aggregate counts skipped separately from pass/fail.",
    "Medium")

add("AS", "Plan Run", "Stop on first failure (if option present)",
    "Plan with failing step 2 and 3 healthy steps after.",
    "1. Enable 'stop on failure'.\n2. Run.",
    "—",
    "Step 1 passes; step 2 fails; remaining steps marked 'not run'; plan summary reflects this.",
    "Medium")

add("AS", "Plan Report", "View per-step results in panel",
    "Run completed.",
    "1. Click into plan run.",
    "—",
    "Each step expandable to show request, response, assertion results.",
    "Medium")

# ============================================================
# DOCUMENTATION (DC)
# ============================================================
add("DC", "Help Panel", "Open help and search a topic",
    "Help panel available.",
    "1. Open Help panel.\n2. Type 'oauth' in search.",
    "—",
    "Topics matching 'oauth' surface (e.g., 'OAuth2 flows'); content renders as Markdown; links open intra-doc or externally.",
    "Medium")

add("DC", "Request Docs", "Markdown docs render correctly",
    "Request with docs field set to '# Hello\\n- bullet'.",
    "1. Open Docs tab.",
    "—",
    "Heading and bullet render with correct HTML semantics; XSS-safe (script tags escaped).",
    "Medium")

# ============================================================
# SEARCH & MARKETPLACE (SE)
# ============================================================
add("SE", "Marketplace Search", "Search returns matching public workspaces",
    "Marketplace reachable.",
    "1. Open marketplace search.\n2. Query 'stripe'.",
    "—",
    "Public workspaces with 'stripe' in name/desc returned; can be linked to current workspace.",
    "Low")

add("SE", "Marketplace Link", "Link a public workspace",
    "Marketplace result open.",
    "1. Click 'Add to workspace'.",
    "—",
    "Linked workspace appears in left sidebar; its collections become referenceable; consumer overrides allowed.",
    "Medium")

# ============================================================
# KEYBOARD SHORTCUTS (KB)
# ============================================================
add("KB", "Panel Switching", "Ctrl+1..9 switches panels",
    "All panels available.",
    "1. Press Ctrl/Cmd+1.\n2. Then Ctrl/Cmd+3.",
    "—",
    "Active panel changes accordingly; no shortcut conflicts with text editing in inputs (panel switching is suppressed while typing).",
    "Medium")

add("KB", "Vault", "Ctrl+K opens Vault tab",
    "Any panel.",
    "1. Press Ctrl/Cmd+K.",
    "—",
    "Right dock opens Vault tab; focus moves to it.",
    "Low")

add("KB", "Refresh", "Ctrl+Shift+R refreshes workspace",
    "Linked workspace.",
    "1. Press Ctrl/Cmd+Shift+R.",
    "—",
    "Refresh action runs (pull check); does NOT reload the browser/Electron window.",
    "Medium")

add("KB", "Send", "Ctrl+Enter sends from any focus",
    "Request open.",
    "1. Focus inside body.\n2. Press Ctrl/Cmd+Enter.",
    "—",
    "Request is sent; suppressing other input behaviors as needed.",
    "High")

# ============================================================
# ACCESSIBILITY & THEMING (AL)
# ============================================================
add("AL", "Keyboard Nav", "Tab order is logical",
    "Editor open.",
    "1. Press Tab repeatedly from URL bar.",
    "—",
    "Focus moves through Method → URL → Send → tabs → editor surfaces in expected reading order; no focus trap.",
    "High")

add("AL", "Focus Ring", "Visible focus on all interactive elements",
    "—",
    "1. Tab to each button and input.",
    "—",
    "Focus ring (var(--purple) accent) is clearly visible; meets 3:1 contrast with background.",
    "High")

add("AL", "Screen Reader", "Buttons have meaningful aria-labels",
    "—",
    "1. Audit with screen reader on key actions (Send, Save, Delete).",
    "—",
    "Each button announces its purpose; no 'button button' redundancy; state changes are announced.",
    "Medium")

add("AL", "Color Independence", "Status not conveyed by color alone",
    "Mixed test results.",
    "1. Inspect Tests panel.",
    "—",
    "Pass/Fail also conveyed by icon or text; users with color blindness can distinguish.",
    "Medium")

add("AL", "Reduced Motion", "Respect prefers-reduced-motion",
    "OS reduce-motion enabled.",
    "1. Open app.\n2. Trigger animations (panel slide, toast).",
    "—",
    "Animations are removed or shortened; no large transitions; functionality preserved.",
    "Low")

# ============================================================
# CROSS-CUTTING (CC)
# ============================================================
add("CC", "Toasts", "Success toast on successful push",
    "Linked workspace; local edits.",
    "1. Push.",
    "—",
    "Toast 'Workspace pushed' appears for ~5s; dismissible via X; non-blocking.",
    "Medium")

add("CC", "Toasts", "Error toast on network failure",
    "Disable network.",
    "1. Send request.",
    "—",
    "Toast describes the network error with actionable hint (e.g., 'Check connection'); not silent.",
    "High")

add("CC", "Toast Stacking", "Multiple toasts stack and dismiss independently",
    "—",
    "1. Trigger 3 toasts in quick succession.",
    "—",
    "All 3 visible stacked; each dismissible individually; auto-dismiss timers independent.",
    "Low")

add("CC", "Confirm Dialog", "Destructive confirm uses red action",
    "Delete a collection.",
    "1. Trigger delete.",
    "—",
    "Confirm dialog: clear title, body explaining consequences, destructive (red) confirm button, cancel is default.",
    "High")

add("CC", "Modal", "Escape closes non-critical modal",
    "Import modal open.",
    "1. Press Esc.",
    "—",
    "Modal closes; focus returns to the trigger button; partially entered data is discarded with warning if changes were made.",
    "Medium")

add("CC", "Modal", "Backdrop click does NOT close destructive modals",
    "Delete confirm modal open.",
    "1. Click outside the dialog.",
    "—",
    "Modal remains open (only explicit Cancel/Confirm dismiss); prevents accidental dismissal.",
    "Medium")

add("CC", "Drag-Drop", "Env priority drag reorders rows",
    "3 envs in priority list.",
    "1. Drag row 3 to position 1.",
    "—",
    "Order updates; lock-step keyboard alternative (Alt+Up/Down) optionally available; persisted.",
    "Medium")

add("CC", "Persistence", "Debounced save coalesces rapid edits",
    "Body editor open.",
    "1. Type 10 keystrokes within 250ms.",
    "—",
    "Only 1 disk write occurs in IndexedDB during the burst; final state matches what was typed.",
    "Medium")

add("CC", "Performance", "Switching workspaces under 1 second",
    "Two workspaces ≤ 50 requests each.",
    "1. Switch from A to B.",
    "—",
    "Switch completes within 1 second on a baseline machine; UI does not freeze.",
    "Medium")

add("CC", "Error Recovery", "Corrupted workspace.json shows recovery UI",
    "Manually corrupt workspace.json on disk.",
    "1. Reopen app.",
    "—",
    "App displays a recovery screen with detailed parse error and options to reset, restore from git, or contact support; does NOT silently crash.",
    "High")

# ============================================================
# Build the workbook
# ============================================================
wb = Workbook()

# --- Sheet 1: README ---
ws = wb.active
ws.title = "README"

ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 110

ws.merge_cells("A1:B1")
c = ws["A1"]
c.value = "APICircle Studio — Manual Test Cases"
c.font = TITLE_FONT
c.fill = HEADER_FILL
c.alignment = CENTER
ws.row_dimensions[1].height = 30

readme = [
    ("Document owner", "QA Engineering"),
    ("Build under test", "APICircle Studio (web + desktop) — fill in version before each cycle"),
    ("Test environment", "Specify OS, browser/Electron version, network, test data"),
    ("Last updated", "2026-05-14"),
    ("", ""),
    ("How to use this workbook", ""),
    ("1.", "Open the 'Test Cases' sheet. Filters are enabled on every column."),
    ("2.", "Pick a module (column B) or priority (column J) to scope your session."),
    ("3.", "For each row, follow the steps in 'Test Steps' (column F) with the data in 'Test Data' (column G)."),
    ("4.", "Compare what you see against 'Expected Result' (column H)."),
    ("5.", "Record what you saw in 'Actual Result' (column I)."),
    ("6.", "Set 'Status' (column J) to Pass / Fail / Blocked / Skipped / Not Run."),
    ("7.", "Add 'Tester' (column L), 'Test Date' (column M), and any defect link in 'Notes' (column N)."),
    ("", ""),
    ("Status values", "Pass, Fail, Blocked, Skipped, Not Run (drop-down on each row)"),
    ("Priority values", "High, Medium, Low"),
    ("", ""),
    ("Summary dashboard", "The 'Summary' sheet auto-aggregates counts by module and status via formulas. Do not edit it manually."),
    ("", ""),
    ("Adding new test cases", "Insert a new row above the last and continue the sequence in column A (e.g., TC-WS-018). Don't break the filter range."),
    ("Defect tracking", "Use column N (Notes) to link to the issue tracker (GitHub Issues, Linear, etc.) when a row fails."),
]

r = 2
for label, val in readme:
    ws.cell(row=r, column=1, value=label).font = BOLD_FONT
    ws.cell(row=r, column=1).alignment = WRAP
    ws.cell(row=r, column=2, value=val).font = BASE_FONT
    ws.cell(row=r, column=2).alignment = WRAP
    r += 1

# --- Sheet 3 (build first so summary can reference) : Test Cases ---
tc_ws = wb.create_sheet("Test Cases")

headers = [
    "TC ID", "Module", "Sub-Feature", "Test Case Title",
    "Pre-conditions", "Test Steps (How to Execute)", "Test Data",
    "Expected Result", "Actual Result", "Status",
    "Priority", "Tester", "Test Date", "Notes / Defect ID"
]

for col_idx, h in enumerate(headers, 1):
    c = tc_ws.cell(row=1, column=col_idx, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER

# Column widths
widths = [12, 24, 24, 38, 32, 50, 28, 50, 30, 12, 10, 14, 12, 28]
for i, w in enumerate(widths, 1):
    tc_ws.column_dimensions[get_column_letter(i)].width = w

tc_ws.row_dimensions[1].height = 32
tc_ws.freeze_panes = "E2"

# Number per-module
counters = {m[0]: 0 for m in MODULES}
module_names = {m[0]: m[1] for m in MODULES}

row = 2
for mod_code, sub, title, pre, steps, data, expected, prio in T:
    counters[mod_code] += 1
    tc_id = f"TC-{mod_code}-{counters[mod_code]:03d}"
    values = [
        tc_id, module_names[mod_code], sub, title,
        pre, steps, data, expected,
        "", "Not Run", prio, "", "", "",
    ]
    for col_idx, v in enumerate(values, 1):
        c = tc_ws.cell(row=row, column=col_idx, value=v)
        c.font = BASE_FONT
        c.alignment = WRAP
        c.border = BORDER

    # Row height auto-friendly; set a reasonable minimum
    tc_ws.row_dimensions[row].height = 75
    row += 1

last_row = row - 1

# Data validation for Status
status_dv = DataValidation(type="list", formula1='"Pass,Fail,Blocked,Skipped,Not Run"', allow_blank=True)
status_dv.error = "Choose one of: Pass, Fail, Blocked, Skipped, Not Run"
status_dv.errorTitle = "Invalid status"
status_dv.prompt = "Select run status"
status_dv.promptTitle = "Status"
tc_ws.add_data_validation(status_dv)
status_dv.add(f"J2:J{last_row}")

# Data validation for Priority
prio_dv = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
tc_ws.add_data_validation(prio_dv)
prio_dv.add(f"K2:K{last_row}")

# Conditional formatting on Status column
status_range = f"J2:J{last_row}"
tc_ws.conditional_formatting.add(status_range,
    CellIsRule(operator="equal", formula=['"Pass"'], fill=PASS_FILL))
tc_ws.conditional_formatting.add(status_range,
    CellIsRule(operator="equal", formula=['"Fail"'], fill=FAIL_FILL))
tc_ws.conditional_formatting.add(status_range,
    CellIsRule(operator="equal", formula=['"Blocked"'], fill=BLOCKED_FILL))
tc_ws.conditional_formatting.add(status_range,
    CellIsRule(operator="equal", formula=['"Not Run"'], fill=NOT_RUN_FILL))
tc_ws.conditional_formatting.add(status_range,
    CellIsRule(operator="equal", formula=['"Skipped"'], fill=SKIPPED_FILL))

# Conditional formatting on Priority column
prio_range = f"K2:K{last_row}"
tc_ws.conditional_formatting.add(prio_range,
    CellIsRule(operator="equal", formula=['"High"'], fill=PRIORITY_HIGH_FILL))
tc_ws.conditional_formatting.add(prio_range,
    CellIsRule(operator="equal", formula=['"Medium"'], fill=PRIORITY_MED_FILL))
tc_ws.conditional_formatting.add(prio_range,
    CellIsRule(operator="equal", formula=['"Low"'], fill=PRIORITY_LOW_FILL))

# Auto-filter the table
tc_ws.auto_filter.ref = f"A1:N{last_row}"

# --- Sheet 2: Summary ---
sm = wb.create_sheet("Summary", 1)  # insert as second sheet

sm.column_dimensions["A"].width = 32
for ch in "BCDEFGH":
    sm.column_dimensions[ch].width = 13

sm.merge_cells("A1:H1")
c = sm["A1"]
c.value = "Test Execution Summary"
c.font = TITLE_FONT
c.fill = HEADER_FILL
c.alignment = CENTER
sm.row_dimensions[1].height = 28

# Headers row
sm_headers = ["Module", "Total", "Not Run", "Pass", "Fail", "Blocked", "Skipped", "Pass %"]
for i, h in enumerate(sm_headers, 1):
    c = sm.cell(row=3, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER

# Per-module formulas
r = 4
for mod_code, mod_name in MODULES:
    sm.cell(row=r, column=1, value=mod_name).font = BOLD_FONT
    sm.cell(row=r, column=1).alignment = WRAP
    sm.cell(row=r, column=1).border = BORDER

    # Total
    sm.cell(row=r, column=2,
        value=f'=COUNTIF(\'Test Cases\'!B:B,"{mod_name}")')
    # Not Run
    sm.cell(row=r, column=3,
        value=f'=COUNTIFS(\'Test Cases\'!B:B,"{mod_name}",\'Test Cases\'!J:J,"Not Run")')
    # Pass
    sm.cell(row=r, column=4,
        value=f'=COUNTIFS(\'Test Cases\'!B:B,"{mod_name}",\'Test Cases\'!J:J,"Pass")')
    # Fail
    sm.cell(row=r, column=5,
        value=f'=COUNTIFS(\'Test Cases\'!B:B,"{mod_name}",\'Test Cases\'!J:J,"Fail")')
    # Blocked
    sm.cell(row=r, column=6,
        value=f'=COUNTIFS(\'Test Cases\'!B:B,"{mod_name}",\'Test Cases\'!J:J,"Blocked")')
    # Skipped
    sm.cell(row=r, column=7,
        value=f'=COUNTIFS(\'Test Cases\'!B:B,"{mod_name}",\'Test Cases\'!J:J,"Skipped")')
    # Pass % = IFERROR(D/(D+E), "")
    sm.cell(row=r, column=8,
        value=f'=IFERROR(D{r}/(D{r}+E{r}),"-")')

    for col in range(2, 9):
        cell = sm.cell(row=r, column=col)
        cell.font = BASE_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        if col == 8:
            cell.number_format = "0.0%;0.0%;-"
    r += 1

# Totals row
total_row = r
sm.cell(row=total_row, column=1, value="TOTAL").font = BOLD_FONT
sm.cell(row=total_row, column=1).fill = SECTION_FILL
sm.cell(row=total_row, column=1).font = SECTION_FONT
sm.cell(row=total_row, column=1).border = BORDER

for col in range(2, 8):
    col_letter = get_column_letter(col)
    sm.cell(row=total_row, column=col,
        value=f"=SUM({col_letter}4:{col_letter}{total_row-1})")
    cell = sm.cell(row=total_row, column=col)
    cell.font = BOLD_FONT
    cell.fill = PatternFill("solid", start_color="EDE9FE")
    cell.alignment = CENTER
    cell.border = BORDER

sm.cell(row=total_row, column=8,
    value=f"=IFERROR(D{total_row}/(D{total_row}+E{total_row}),\"-\")")
cell = sm.cell(row=total_row, column=8)
cell.font = BOLD_FONT
cell.fill = PatternFill("solid", start_color="EDE9FE")
cell.alignment = CENTER
cell.border = BORDER
cell.number_format = "0.0%;0.0%;-"

# Summary by Priority
r = total_row + 3
sm.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
c = sm.cell(row=r, column=1, value="Counts by Priority")
c.font = SECTION_FONT
c.fill = SECTION_FILL
c.alignment = CENTER
sm.row_dimensions[r].height = 22
r += 1

prio_headers = ["Priority", "Total", "Not Run", "Pass", "Fail", "Blocked", "Skipped", "Pass %"]
for i, h in enumerate(prio_headers, 1):
    c = sm.cell(row=r, column=i, value=h)
    c.font = HEADER_FONT
    c.fill = HEADER_FILL
    c.alignment = CENTER
    c.border = BORDER
r += 1

for prio in ["High", "Medium", "Low"]:
    sm.cell(row=r, column=1, value=prio).font = BOLD_FONT
    sm.cell(row=r, column=1).alignment = CENTER
    sm.cell(row=r, column=1).border = BORDER
    if prio == "High":
        sm.cell(row=r, column=1).fill = PRIORITY_HIGH_FILL
    elif prio == "Medium":
        sm.cell(row=r, column=1).fill = PRIORITY_MED_FILL
    else:
        sm.cell(row=r, column=1).fill = PRIORITY_LOW_FILL

    sm.cell(row=r, column=2, value=f'=COUNTIF(\'Test Cases\'!K:K,"{prio}")')
    sm.cell(row=r, column=3, value=f'=COUNTIFS(\'Test Cases\'!K:K,"{prio}",\'Test Cases\'!J:J,"Not Run")')
    sm.cell(row=r, column=4, value=f'=COUNTIFS(\'Test Cases\'!K:K,"{prio}",\'Test Cases\'!J:J,"Pass")')
    sm.cell(row=r, column=5, value=f'=COUNTIFS(\'Test Cases\'!K:K,"{prio}",\'Test Cases\'!J:J,"Fail")')
    sm.cell(row=r, column=6, value=f'=COUNTIFS(\'Test Cases\'!K:K,"{prio}",\'Test Cases\'!J:J,"Blocked")')
    sm.cell(row=r, column=7, value=f'=COUNTIFS(\'Test Cases\'!K:K,"{prio}",\'Test Cases\'!J:J,"Skipped")')
    sm.cell(row=r, column=8, value=f'=IFERROR(D{r}/(D{r}+E{r}),"-")')

    for col in range(2, 9):
        cell = sm.cell(row=r, column=col)
        cell.font = BASE_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        if col == 8:
            cell.number_format = "0.0%;0.0%;-"
    r += 1

# Save
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Total test cases: {len(T)}")
